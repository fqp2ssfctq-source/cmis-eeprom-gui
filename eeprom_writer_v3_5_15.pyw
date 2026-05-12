"""
EEPROM CMIS Writer  v2.9.4
CP2112-F03-GM USB-HID I2C 브릿지 기반 CMIS 5.2 EEPROM 관리 도구

구조:
  CP2112I2C    - Silicon Labs CP2112 USB-HID → I2C 통신 드라이버
  CanvasTable  - Canvas 기반 커스텀 테이블 위젯
                 (열 너비 드래그 조절 / 셀 선택 복사 / 바이트 팝업 편집)
  App          - 메인 GUI 애플리케이션 (tkinter)

탭 구성:
  연결   - CP2112 장치 선택 및 I2C 연결
  EEPROM 뷰어 - 페이지별 CMIS 메모리 맵 표시 및 편집
  쓰기   - EEPROM 실제 쓰기 및 검증
  로그   - 동작 이력
  정보   - 버전 및 시스템 정보
"""
# ── 표준 라이브러리 및 GUI 프레임워크 ────────────────────
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, time, os, sys, copy, colorsys, traceback, logging, math
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── 앱 메타 정보 ──────────────────────────────────────────
# TYPE_BADGE/TYPE_COLOR: 버전 히스토리 표시용 배지
#   [M]=major(하위 호환 불가), [+]=minor(기능 추가), [·]=patch(버그 수정)
APP_NAME    = "EEPROM CMIS Writer"
APP_VERSION = "3.5.15"
APP_DATE    = "2026-05-11"
# ── 참조 스펙 버전 ─────────────────────────────────────────
SPEC_CMIS   = "OIF-CMIS-05.2"       # April 27, 2022
SPEC_SFF8024= "SFF-8024 Rev 4.13"   # July 11, 2025
CHANGELOG = [
    ("3.5.15","2026-05-11","patch",  "A0h: SFF8024Identifier/CmisRevision/MemoryModel/ModuleState/FlagsSummaryBank/TempMon/VccMon/Aux*Mon/ModuleControls/ModuleFaultCause"),
    ("3.5.15","2026-05-11","patch",  "A0h: OutputDisableTx/Rx/OutputSquelchForceTx → CMIS 명칭, ModuleFaultCause 위치 수정 (idx 41)"),
    ("3.5.15","2026-05-11","patch",  "P00h: SFF8024IdentifierCopy/VendorPN/VendorSN/CLEICode/CableLinkLength/MediaConnectorType/PageChecksum"),
    ("3.5.15","2026-05-11","patch",  "P01h: InactiveFWVersion/HardwareRevision/LengthSMF/ImplementedMemoryPages/MaxDuration*/SupportedCDB*"),
    ("3.5.15","2026-05-11","patch",  "P02h: OpticalPowerHighAlarm*/LaserBiasHighAlarm*/Lane* → CMIS 5.2 공식 명칭, _P02H_DESC 동기화"),
    ("3.5.15","2026-05-11","patch",  "P10h: DPDeinit/OutputDisableTx/OutputSquelchForceTx/TxInputAmplitude*/TxInputEq*/RxOutputEq*/CDREnable*"),
    ("3.5.15","2026-05-11","patch",  "P11h: SCS1::AppSelCode*/DPConfig*/TxInputAmplitude*/TxInputEq*/RxOutputEq*/CDREnableRx*"),
    ("3.5.14","2026-05-11","patch",  "P11h _decoded_p11: SCS1 Rx CDR Bypass (idx 59~66) decoded (was returning '—')"),
    ("3.5.14","2026-05-11","patch",  "P11h defs: Reserved [0xDE~0xFE] length 33 (was 34, overlapping PageChecksum)"),
    ("3.5.14","2026-05-11","patch",  "P00h/P01h/P03h Reserved range labels corrected to match actual byte addresses"),
    ("3.5.13","2026-05-11","patch",  "CanvasTable: header row drawn last so it stays visible when scrolling vertically"),
    ("3.5.12","2026-05-03","minor",  "P10h 필드 정의: Lane SI Controls (DataPathDeinit/TxDisable/Amplitude/PreCursor/PostCursor/Equalization/CDR)"),
    ("3.5.12","2026-05-03","minor",  "P11h 필드 정의: SCS1 (AppSelLane/DataPathCtrl/SI Controls staged/RxCDRBypass)"),
    ("3.5.12","2026-05-03","patch",  "P10h/P11h Decoded: 비트맵 필드 2진수 표시, SI값 숫자 표시"),
    ("3.5.12","2026-05-03","patch",  "앱 아이콘: CMIS 텍스트 아이콘 (PIL 설치 시 자동 적용)"),
    ("3.5.11","2026-04-30","minor",  "설정 유지: .cmis_config.json — device/i2c/geometry/REF/페이지/예외 저장·복원"),
    ("3.5.11","2026-04-30","minor",  "CanvasTable 가로 스크롤 추가 (Shift+휠 / 가로 스크롤바)"),
    ("3.5.11","2026-04-30","patch",  "비교 예외: A0h 02h (MemoryModel) 추가 → 02~03h 범위로 통합"),
    ("3.5.11","2026-04-30","patch",  "타이틀바/연결탭 SPEC_CMIS, SPEC_SFF8024 버전 표기"),
    ("3.5.11","2026-04-30","patch",  "비교: P00h[80h] 누락 버그 수정 — 독립적으로 비교"),
    ("3.5.10","2026-04-29","patch",  "버그수정: _cmp_dut_loaded 플래그 / 검증실패 화면갱신 / CC orig동기화"),
    ("3.5.10","2026-04-29","major",  "쓰기: write_byte ACK polling — NVM busy(주소 NACK) 자동 재시도"),
    ("3.5.7","2026-04-28","patch",  "쓰기: tWR 보장 — burst 후 20ms 대기"),
    ("3.5.7","2026-04-28","patch",  "검증: 1바이트씩 read → 페이지 단위 read_page로 교체 (효율 개선)"),
    ("3.5.7","2026-04-28","patch",  "cs_map p00 제거 — 모듈이 checksum 자체 계산, 호스트 write 무시"),
    ("3.5.6","2026-04-28","patch",  "쓰기: Auto PW 재전송 제거 / P00h RO 필드 정확히 제외"),
    ("3.5.3","2026-04-28","patch",  "쓰기: P00h Reserved(DDh,E7~FFh) / P02h Reserved(E0~FEh) 제외"),
    ("3.5.2","2026-04-28","patch",  "팝업: x_root/y_root 스크린 절대좌표, PW 자동 재전송"),
    ("3.5.1","2026-04-28","patch",  "쓰기 확인창: RO 제외 바이트 수 표시 / P01h 쓰기 차단"),
    ("3.5.0","2026-04-28","major",  "팝업 편집 시 Decoded 실시간 미리보기 (팝업 하단 표시)"),
    ("3.5.0","2026-04-28","major",  "P00h AppDescriptor[0~7]/MediaLanes 누락 필드 추가"),
    ("3.5.0","2026-04-28","major",  "A0h ModuleFaultCause/LaneFlags/LaneMonitors/AppDesc[1~7] 누락 필드 추가"),
    ("3.5.0","2026-04-28","major",  "P01h Timing/CDB/MediaLane/SI Controls 누락 필드 추가"),
    ("3.5.0","2026-04-28","major",  "P02h Per-Lane Tx/Rx 임계값 누락 필드 추가"),
    ("3.4.4","2026-04-23","patch",  "CMIS 5.2 표기 수정, 스펙 버전 상수화, s16/u16 헬퍼 적용"),
    ("3.4.3","2026-04-14","patch",  "SFF8024 식별자 테이블 Rev 4.13 완전 업데이트 (7→26개)"),
    ("3.4.2","2026-04-14","major",  "비교 탭: REF 파일/DUT EEPROM읽기 분리, 자동비교, PASS/FAIL 배지"),
    ("3.4.2","2026-04-14","patch",  "비교 예외: PasswordEntry(A0h 7A~7D) 추가 / Identifier 중복FAIL 제거"),
    ("3.0.5","2026-04-07","patch",  "lambda 클로저 변수 캡처 오류 수정"),
    ("3.0.6","2026-04-07","major",  "USB 캡처 분석 기반: 실제 동작과 동일한 시퀀스로 교체"),
    ("3.0.6","2026-04-07","major",  "패킷 분석: 8bit 주소 사용 / 128바이트 일괄 읽기로 교체"),
    ("3.3.2","2026-04-09","minor",  "파일→전체쓰기 / EEPROM읽기→dirty쓰기 자동 구분"),
    ("3.3.3","2026-04-09","patch",  "P02h Tx/Rx Power: µW → dBm 단위 변환"),
    ("3.3.2","2026-04-09","patch",  "쓰기: 항상 선택 페이지 전체 쓰기로 단순화"),
    ("3.3.1","2026-04-09","minor",  "읽기/쓰기 페이지 선택 체크박스 / PW 자동 저장"),
    ("3.3.0","2026-04-09","major",  "쓰기 탭 제거 → 뷰어에서 수정 후 바로 쓰기 통합"),
    ("3.2.4","2026-04-08","minor",  "Password Unlock 기능 추가 (PW 주소/값 입력, 잠금 상태 표시)"),
    ("3.2.3","2026-04-08","minor",  "CP2112 GPIO Ready 설정 버튼 / Reset 버튼 추가"),
    ("3.2.2","2026-04-08","patch",  "_is_simulated 완전제거/_scan_i2c제거/미사용코드정리"),
    ("3.2.1","2026-04-08","patch",  "전페이지 체크섬 실시간/미사용코드 제거/8bit주소 일관성 수정"),
    ("3.2.0","2026-04-08","major",  "Clear버튼/시뮬레이션제거/스캔삭제/편집바삭제/드롭다운/Decoded직접편집/체크섬실시간"),
    ("3.1.2","2026-04-07","minor",  "P01h/P02h 핵심 필드 Decoded/Description 추가"),
    ("3.1.1","2026-04-07","patch",  "바이트팝업 인덱스 dec 주소 표시 / I2C 기본주소 0x50 수정"),
    ("3.1.0","2026-04-07","minor",  "I2C 기본주소 0x50 / 읽기 진행바 / 작업 시에만 포트 점유"),
    ("3.0.7","2026-04-07","patch",  "read_page: ForceReadResponse 추가, 읽기 시퀀스 LabVIEW VI와 동일하게"),
    ("3.0.6","2026-04-07","patch",  "버그수정: read_bytes→read_page, 8bit주소 변환, SimulatedI2C read_page 추가"),
    ("3.0.5","2026-04-07","minor",  "I2C 주소 스캔 기능 추가, 스캔 결과로 주소 자동 설정"),
    ("3.0.4","2026-04-07","patch",  "전체 DLL 함수 시그니처 헤더파일 기준으로 교정"),
    ("3.0.3","2026-04-07","patch",  "HidSmbus_Open 파라미터 순서 수정 (deviceNum, &handle)"),
    ("3.0.2","2026-04-07","patch",  "장치 이름 GetAttributes로 VID/PID/S/N 표시"),
    ("3.0.1","2026-04-07","patch",  "장치 이름 문자열 깨짐 수정 (unicode→bytes 버퍼)"),
    ("3.0.0","2026-04-07","major",  "hidapi raw HID -> SLABHIDtoSMBus.dll ctypes 방식으로 교체"),
    ("2.9.6","2026-04-07","patch",  "crash.log 인코딩 수정 / OSError: read error 처리 추가"),
    ("2.9.5","2026-04-07","patch",  "CP2112 장치 열림 상태 추적 / ValueError→IOError 변환"),
    ("2.9.4","2026-04-06","patch",  "코드 전체 주석 추가"),
    ("2.9.3","2026-04-06","patch",  "바이트별 개별 Entry 팝업으로 편집"),
    ("2.9.2","2026-04-06","patch",  "멀티바이트 Value 편집 시 전체 바이트 표시 및 수정"),
    ("2.9.1","2026-04-06","patch",  "Decoded 미해석 필드 N/A 표기, 편집 시 Decoded 동시 갱신"),
    ("2.9.0","2026-04-06","minor",  "_safe_call 제거 / Value[hex] 인라인 직접 편집"),
    ("2.8.1","2026-04-06","patch",  "crash.log: 실행 중 모든 예외 기록"),
    ("2.8.0","2026-04-06","minor",  "코드 정리: 중복/미사용 코드 제거"),
    ("2.7.4","2026-04-06","patch",  "hidapi 없을 때 자동 설치"),
    ("2.7.3","2026-04-06","patch",  "I2C 주소 0x50 고정"),
    ("2.7.2","2026-04-06","patch",  "ver_label 색상 덮어쓰기 방지"),
    ("2.7.1","2026-04-06","patch",  "HID_OK 참조 수정"),
    ("2.7.0","2026-04-06","minor",  "CP2112 USB-HID I2C 브릿지 지원"),
    ("2.6.5","2026-04-06","minor",  "셀/범위 드래그 선택 + Ctrl+C 복사"),
    ("2.6.4","2026-04-06","patch",  "기본 모드: 라이트 100%"),
    ("2.6.3","2026-04-06","patch",  "Value[hex] 전체 바이트 표시"),
    ("2.6.2","2026-04-06","patch",  "create_text width= 제거"),
    ("2.6.1","2026-04-06","minor",  "헤더 드래그 열 너비 조절"),
    ("2.6.0","2026-04-06","major",  "CanvasTable 교체: 열 너비 완전 제어"),
    ("2.5.0","2026-04-06","minor",  "P00h Value 해석 강화 / Decoded 열"),
    ("2.4.0","2026-04-06","minor",  "6컬럼 구성 / VendorSN DateCode 수정"),
    ("2.3.0","2026-04-06","minor",  "MultiColumn Treeview / 열 구분선"),
    ("2.2.0","2026-04-06","patch",  "Canvas 오버레이 제거"),
    ("2.1.0","2026-04-06","minor",  "열 간격 / 라이트모드 밝기 유지"),
    ("2.0.0","2026-04-06","major",  "버전 관리 / 능동 밝기 조절"),
    ("1.1.0","2026-04-05","minor",  "시뮬레이션 모드 추가"),
    ("1.0.0","2026-04-04","major",  "최초 릴리즈"),
]
TYPE_BADGE = {"major":"[M]","minor":"[+]","patch":"[·]"}
TYPE_COLOR = {"major":"#f07070","minor":"#7ab3f5","patch":"#888888"}

# ── crash.log 설정 ────────────────────────────────────────
def _setup_crash_log():
    """실행 파일과 같은 폴더에 crash.log 생성, 모든 예외 기록"""
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "crash.log")
    # FileHandler로 직접 생성 → Windows 한글 환경에서 인코딩 명시
    handler=logging.FileHandler(log_path, encoding="utf-8", mode="a")
    handler.setLevel(logging.ERROR)
    handler.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    ))
    root_logger=logging.getLogger()
    root_logger.setLevel(logging.ERROR)
    root_logger.addHandler(handler)
    # 미처리 예외 자동 기록
    def _exc_handler(exc_type, exc_value, exc_tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb)
            return
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_tb))
        # 원래 동작도 유지
        sys.__excepthook__(exc_type, exc_value, exc_tb)
    sys.excepthook = _exc_handler
    return log_path

CRASH_LOG = _setup_crash_log()

# ── SLABHIDtoSMBus.dll 로드 ──────────────────────────────
# Silicon Labs 전용 고수준 API (LabVIEW VI와 동일한 방식)
import ctypes

def _load_slab_dll():
    """SLABHIDtoSMBus.dll 로드. 스크립트 폴더 우선 탐색."""
    if sys.platform != "win32":
        return False, None
    dirs = [os.path.dirname(os.path.abspath(__file__)), os.getcwd()]
    for d in dirs:
        path = os.path.join(d, "SLABHIDtoSMBus.dll")
        if os.path.exists(path):
            try:
                return True, ctypes.WinDLL(path)
            except Exception as e:
                print(f"DLL 로드 실패: {e}")
    print("SLABHIDtoSMBus.dll 없음. 스크립트와 같은 폴더에 복사하세요.")
    return False, None

HID_OK, _SLAB_DLL = _load_slab_dll()
CP2112_VID = 0x10C4
CP2112_PID = 0xEA90

# ── 밝기 유틸 ──────────────────────────────────────────────
# _h2rgb : "#RRGGBB" → (r,g,b) 0~1 float 변환
# _rgb2h : (r,g,b) → "#RRGGBB" 변환
# _adj   : hex 색상을 HSV 공간에서 밝기(f) 조정
# make_theme: 베이스 테마 dict를 밝기 factor로 조정한 새 dict 반환
#             SKIP 키(강조색 등)는 밝기 조정 제외
def _h2rgb(h):
    h=h.lstrip("#")
    return tuple(int(h[i:i+2],16)/255 for i in(0,2,4))
def _rgb2h(r,g,b):
    return "#{:02X}{:02X}{:02X}".format(int(r*255),int(g*255),int(b*255))
def _adj(hc,f):
    r,g,b=_h2rgb(hc); hh,s,v=colorsys.rgb_to_hsv(r,g,b)
    return _rgb2h(*colorsys.hsv_to_rgb(hh,s,max(0.0,min(1.0,v*f))))
SKIP={"acc","grn","amb","red","dirty","sel_fg","entry_fg"}  # 밝기 조정 제외 키
def make_theme(base,f):
    return {k:(_adj(v,f) if isinstance(v,str) and v.startswith("#") and k not in SKIP else v)
            for k,v in base.items()}

# ── 테마 ──────────────────────────────────────────────────
# 다크/라이트 베이스 테마 dict
# bg0~bg3 : 배경 계층 (bg0=가장 어두움, bg3=가장 어두운 강조)
# t1~t4   : 텍스트 계층 (t1=주 텍스트, t4=비활성)
# acc      : 강조색 (탭 활성, 헤더 텍스트 등)
# col0_*   : Addr 열 배경 (짝수/홀수 행)
# col1_*   : Data 열 배경 (Value, Decoded, Field, Description)
# col_sep  : 열 구분선 색상
_DARK={
    "bg0":"#1a1a1a","bg1":"#252525","bg2":"#202020","bg3":"#161616",
    "bd":"#3a3a3a","bd2":"#505050",
    "t1":"#e0e0e0","t2":"#aaaaaa","t3":"#686868","t4":"#444444",
    "acc":"#7ab3f5","grn":"#6ac87a","amb":"#f0c060","red":"#f07070","dirty":"#f0a050",
    "sel_bg":"#1e3a5f","sel_fg":"#7ab3f5",
    "entry_bg":"#161616","entry_fg":"#f0c060",
    "btn_bg":"#2a2a2a","btn_fg":"#d0d0d0",
    # 열별 배경 (짝수열 / 홀수열)
    "col0_even":"#1c1c1c","col0_odd":"#232323",
    "col1_even":"#1c1c1c","col1_odd":"#232323",
    "col2_even":"#1c1c1c","col2_odd":"#232323",
    "col3_even":"#1c1c1c","col3_odd":"#232323",
    "col4_even":"#1c1c1c","col4_odd":"#232323",
    "col_head":"#2e2e2e","col_sep":"#404040",
    "row_hl":"#2a2a2a",
}
_LIGHT={
    "bg0":"#d2d2d2","bg1":"#c6c6c6","bg2":"#cccccc","bg3":"#c0c0c0",
    "bd":"#9a9a9a","bd2":"#787878",
    "t1":"#111111","t2":"#333333","t3":"#666666","t4":"#999999",
    "acc":"#185FA5","grn":"#2a5a0a","amb":"#7a4a00","red":"#8a2020","dirty":"#7a4000",
    "sel_bg":"#b8d4ee","sel_fg":"#0a3a70",
    "entry_bg":"#bbbbbb","entry_fg":"#7a4a00",
    "btn_bg":"#c8c8c8","btn_fg":"#111111",
    "col0_even":"#c0c0c0","col0_odd":"#cccccc",
    "col1_even":"#c0c0c0","col1_odd":"#cccccc",
    "col2_even":"#c0c0c0","col2_odd":"#cccccc",
    "col3_even":"#c0c0c0","col3_odd":"#cccccc",
    "col4_even":"#c0c0c0","col4_odd":"#cccccc",
    "col_head":"#b4b4b4","col_sep":"#909090",
    "row_hl":"#e8e8e8",
}

# ── CMIS 상수 ─────────────────────────────────────────────
# SFF8024 : SFF-8024 모듈 타입 식별자 테이블
# I2C_ADDR_A0 : CMIS 모듈 I2C 8bit 주소 (7bit=0x50)
# PAGE_KEYS   : 내부 데이터 딕셔너리 키 (소문자)
# PAGE_LABELS : GUI 표시용 레이블 (대문자)
# COL_CFG     : 테이블 열 정의
#   (헤더문자열, 기본폭px, 최소폭px, col_key, stretch여부)
#   col_key=0 → Addr 배경, col_key=1 → Data 배경
# SFF-8024 Rev 4.13 Table 4-1 Identifier Values (전체)
SFF8024={
    0x00:"Unknown/Unspecified",
    0x01:"GBIC",
    0x02:"Soldered to motherboard (SFF-8472)",
    0x03:"SFP/SFP+/SFP28 (SFF-8472)",
    0x04:"300 pin XBI",
    0x05:"XENPAK",
    0x06:"XFP",
    0x07:"XFF",
    0x08:"XFP-E",
    0x09:"XPAK",
    0x0A:"X2",
    0x0B:"DWDM-SFP/SFP+",
    0x0C:"QSFP (INF-8438)",
    0x0D:"QSFP+ or later (SFF-8636/8436)",
    0x0E:"CXP or later",
    0x0F:"Shielded Mini Multilane HD 4X",
    0x10:"Shielded Mini Multilane HD 8X",
    0x11:"QSFP28 or later (SFF-8636)",
    0x12:"CXP2 (CXP28) or later",
    0x13:"CDFP Style 1/2",
    0x14:"Shielded Mini Multilane HD 4X Fanout",
    0x15:"Shielded Mini Multilane HD 8X Fanout",
    0x16:"CDFP Style 3",
    0x17:"microQSFP",
    0x18:"QSFP-DD 8X (CMIS)",
    0x19:"OSFP 8X (CMIS)",
    0x1A:"SFP-DD (SFP-DD Mgmt I/F)",
    0x1B:"DSFP Dual SFP",
    0x1C:"x4 MiniLink/OcuLink",
    0x1D:"x8 MiniLink",
    0x1E:"QSFP+ or later (CMIS)",
    0x1F:"SFP-DD (CMIS)",
    0x20:"SFP+ and later (CMIS)",
    0x21:"OSFP-XD (CMIS)",
    0x22:"OIF-ELSFP (CMIS)",
    0x23:"CDFP x4 PCIe (CMIS)",
    0x24:"CDFP x8 PCIe (CMIS)",
    0x25:"CDFP x16 PCIe (CMIS)",
}
# 드롭다운용 목록 (value→label)
SFF8024_LIST=[(k,f"0x{k:02X} {v}") for k,v in sorted(SFF8024.items())]
MEDIA_TECH={0:"850nm VCSEL",1:"1310nm VCSEL",2:"1550nm VCSEL",
            3:"1310nm FP",4:"1310nm DFB",5:"1490nm DFB",
            6:"1310nm EML",7:"1550nm EML"}
MEDIA_TECH_LIST=[(k,f"0x{k:02X} {v}") for k,v in sorted(MEDIA_TECH.items())]
I2C_ADDR_A0=0xA0
PAGE_LABELS=["A0h","P00h","P01h","P02h","P03h","P10h","P11h"]
PAGE_KEYS  =["a0","p00","p01","p02","p03","p10","p11"]

# ── P02h Description 상수 테이블 (매 호출마다 재생성 방지) ────
# (이름, 단위 설명) 딕셔너리: key = byte index (0-based, 짝수만)
_P02H_DESC = {
    0: ("TempHighAlarm",   "S16/256 °C"),
    2: ("TempLowAlarm",    "S16/256 °C"),
    4: ("TempHighWarning", "S16/256 °C"),
    6: ("TempLowWarning",  "S16/256 °C"),
    8: ("VccHighAlarm",    "U16×100µV→V"),
   10: ("VccLowAlarm",     "U16×100µV→V"),
   12: ("VccHighWarning",  "U16×100µV→V"),
   14: ("VccLowWarning",   "U16×100µV→V"),
   16: ("Aux1HighAlarm",   "S16 (TEC Current/Custom)"),
   18: ("Aux1LowAlarm",    "S16 (TEC Current/Custom)"),
   20: ("Aux1HighWarning", "S16 (TEC Current/Custom)"),
   22: ("Aux1LowWarning",  "S16 (TEC Current/Custom)"),
   24: ("Aux2HighAlarm",   "S16 (TEC Current/Laser Temp)"),
   26: ("Aux2LowAlarm",    "S16 (TEC Current/Laser Temp)"),
   28: ("Aux2HighWarning", "S16 (TEC Current/Laser Temp)"),
   30: ("Aux2LowWarning",  "S16 (TEC Current/Laser Temp)"),
   32: ("Aux3HighAlarm",   "S16 (Laser Temp/Aux Voltage)"),
   34: ("Aux3LowAlarm",    "S16 (Laser Temp/Aux Voltage)"),
   36: ("Aux3HighWarning", "S16 (Laser Temp/Aux Voltage)"),
   38: ("Aux3LowWarning",  "S16 (Laser Temp/Aux Voltage)"),
   40: ("CustomMonHighAlarm",   "S16/U16"),
   42: ("CustomMonLowAlarm",    "S16/U16"),
   44: ("CustomMonHighWarning", "S16/U16"),
   46: ("CustomMonLowWarning",  "S16/U16"),
   48: ("OpticalPowerHighAlarmTx",   "U16×0.1µW→dBm"),
   50: ("OpticalPowerLowAlarmTx",    "U16×0.1µW→dBm"),
   52: ("OpticalPowerHighWarningTx", "U16×0.1µW→dBm"),
   54: ("OpticalPowerLowWarningTx",  "U16×0.1µW→dBm"),
   56: ("LaserBiasHighAlarmTx",   "U16×2µA"),
   58: ("LaserBiasLowAlarmTx",    "U16×2µA"),
   60: ("LaserBiasHighWarningTx", "U16×2µA"),
   62: ("LaserBiasLowWarningTx",  "U16×2µA"),
   64: ("OpticalPowerHighAlarmRx",   "U16×0.1µW→dBm"),
   66: ("OpticalPowerLowAlarmRx",    "U16×0.1µW→dBm"),
   68: ("OpticalPowerHighWarningRx", "U16×0.1µW→dBm"),
   70: ("OpticalPowerLowWarningRx",  "U16×0.1µW→dBm"),
   72: ("LaneOpticalPowerHighAlarmTx",   "U16×0.1µW→dBm"),
   74: ("LaneOpticalPowerLowAlarmTx",    "U16×0.1µW→dBm"),
   76: ("LaneOpticalPowerHighWarningTx", "U16×0.1µW→dBm"),
   78: ("LaneOpticalPowerLowWarningTx",  "U16×0.1µW→dBm"),
   80: ("LaneLaserBiasHighAlarmTx",    "U16×2µA"),
   82: ("LaneLaserBiasLowAlarmTx",     "U16×2µA"),
   84: ("LaneLaserBiasHighWarningTx",  "U16×2µA"),
   86: ("LaneLaserBiasLowWarningTx",   "U16×2µA"),
   88: ("LaneOpticalPowerHighAlarmRx",   "U16×0.1µW→dBm"),
   90: ("LaneOpticalPowerLowAlarmRx",    "U16×0.1µW→dBm"),
   92: ("LaneOpticalPowerHighWarningRx", "U16×0.1µW→dBm"),
   94: ("LaneOpticalPowerLowWarningRx",  "U16×0.1µW→dBm"),
   96: ("Reserved", "—"),
  127: ("PageChecksum", "Bytes 128~254 sum mod256"),
}

# ── CMIS Revision 버전명 맵 (A0h Byte 1 디코딩용) ─────────
# CMIS: 상위 4bit = Major, 하위 4bit = Minor (예: 0x52 → v5.2)
_CMIS_REV_MAP = {
    0x30: "v3.0",
    0x40: "v4.0",
    0x50: "v5.0",
    0x51: "v5.1",
    0x52: "v5.2",  # OIF-CMIS-05.2 (2022-04-27) ← 현재 참조 버전
}

# ── CMIS 5.2 쓰기 가능(RW/WO) 바이트 마스크 ─────────────────
# set에 포함된 idx만 실제 I2C write 시도.
# 나머지는 RO — 모듈이 NACK 반환하여 S1=0x00 오류 발생.
#
# A0h (Lower Memory) — 대부분 RO, 일부만 RW
_RW_A0 = set([
    26,          # ModuleControls          (RW)
    27,          # ModuleFaultCause        (WO, 쓰면 clear)
    65,          # TxDisable per-lane      (RW)
    66,          # RxOutputDisable         (RW)
    67,          # TxForcedSquelch         (RW)
    80,81,82,83, # AppDescSelector Lane1~4 (RW)
    84,          # AppDescSelector Lane5~8 (RW)
    # 122~125 (7A~7Dh) PasswordEntry 제외:
    #   읽기 시 항상 00h로 반환되므로 일반 쓰기 시 00h 덮어쓰면 재잠금됨
    #   → Password Unlock 버튼 전용, 여기서는 절대 쓰지 않음
    127,         # PageSelect              (RW)
])
# P00h (Page 00h) — CMIS 5.2 기준 필드별 RW 권한
# RO: idx 0(Identifier NVM copy), 72~74, 76~84, 85~93, 94(PageChecksum-모듈자체계산)
# RW: idx 1~71(VendorName~CLEICode), 75(ConnectorType)
# ※ idx 94(PageChecksum): 모듈이 자체 계산하므로 호스트 write 무시 → 제외
_RW_P00 = (set(range(1,  72))   # VendorName ~ CLEICode
         | {75})                 # ConnectorType
# P01h (Page 01h) — 광고 레지스터, 모두 RO (벤더 펌웨어가 기록)
_RW_P01 = set()
# P03h (Page 03h) — 레인 모니터 실시간값, 모두 RO
_RW_P03 = set()
# P02h (Page 02h) — 임계값, Reserved 바이트 제외
# idx 96~126 (E0~FEh): Reserved → NACK
# idx 127 (FFh): PageChecksum (RW, 포함)
_RW_P02 = set(range(0, 96)) | {127}
# P10h / P11h (Lane Control/Config) — 전부 RW
_RW_P10 = set(range(128))
_RW_P11 = set(range(128))

_CMIS_RW = {
    "a0":  _RW_A0,
    "p00": _RW_P00,
    "p01": _RW_P01,
    "p02": _RW_P02,
    "p03": _RW_P03,
    "p10": _RW_P10,
    "p11": _RW_P11,
}

# 열 정의: (헤더, 기본폭, 최소폭, col_key, stretch)
COL_CFG=[
    ("Addr(dec)",  80,  55, 0, False),
    ("Addr(hex)",  90,  65, 0, False),
    ("Value[hex]",140,  90, 1, False),
    ("Decoded",   220, 140, 1, False),
    ("Field",     185, 115, 1, False),
    ("Description",320,140, 1, True),
]

# ── 비교 탭 상수 ──────────────────────────────────────────
# _CMP_FIELD_A0/P00: (pk, byte idx) → 필드명  (Treeview Field 열 표시용)
_CMP_FIELD_A0 = {
    0:"SFF8024Identifier",   1:"CmisRevision",    2:"MemoryModel",
    3:"GlobalStatus",
    4:"FlagsSummary[Bank0]", 5:"FlagsSummary[Bank1]",
    6:"FlagsSummary[Bank2]", 7:"FlagsSummary[Bank3]",
    8:"ModuleFlags[CDB/FW]", 9:"ModuleFlags[Vcc/Temp]",
    10:"ModuleFlags[0Ah]",  11:"ModuleFlags[0Bh]",
    12:"ModuleFlags[0Ch]",  13:"ModuleFlags[0Dh]",
    14:"TempMon[MSB]",      15:"TempMon[LSB]",
    16:"VccMon[MSB]",       17:"VccMon[LSB]",
    18:"Aux1Mon[MSB]",      19:"Aux1Mon[LSB]",
    20:"Aux2Mon[MSB]",      21:"Aux2Mon[LSB]",
    22:"Aux3Mon[MSB]",      23:"Aux3Mon[LSB]",
    24:"CustomMon[MSB]",    25:"CustomMon[LSB]",
    26:"ModuleControls",
    37:"CdbStatus1",        38:"CdbStatus2",
    39:"ActiveFW[Major]",   40:"ActiveFW[Minor]",
    85:"AppDesc[0]HostIfID",86:"AppDesc[0]MediaIfID",
    87:"AppDesc[0]LaneCnt", 88:"AppDesc[0]LaneAssign",
    122:"PasswordEntry[0]", 123:"PasswordEntry[1]",
    124:"PasswordEntry[2]", 125:"PasswordEntry[3]",
    126:"PageSelectBank",   127:"PageSelectPage",
}
_CMP_FIELD_P00 = {
    0:"SFF8024Identifier(NVM)",
    **{i:"VendorName"  for i in range(1,17)},
    17:"VendorOUI[0]", 18:"VendorOUI[1]", 19:"VendorOUI[2]",
    **{i:"VendorPN"    for i in range(20,36)},
    36:"VendorRev[0]", 37:"VendorRev[1]",
    **{i:"VendorSN"    for i in range(38,54)},
    **{i:"DateCode"    for i in range(54,62)},
    **{i:"CLEICode"    for i in range(62,72)},
    72:"ModulePowerClass", 73:"MaxPower",
    74:"CableLinkLength",  75:"ConnectorType",
    84:"MediaInterfaceTech", 94:"PageChecksum",
}
# 기본 예외: (pk, start_idx, end_idx, 설명 라벨, 기본 체크 여부)
_CMP_DEFAULT_EXC = [
    ("a0",   2,  3,  "MemoryModel/GlobalStatus (메모리 구성 및 실시간 모듈 상태)", True),
    ("a0",   4, 13,  "FlagsSum/ModuleFlags (인터럽트 플래그)",       True),
    ("a0",  14, 17,  "Temp/Vcc Monitor (실시간 측정값)",             True),
    ("a0",  18, 25,  "Aux1/2/3/Custom Monitor (실시간 측정값)",      True),
    ("a0",  26, 26,  "ModuleControls (상태 의존 필드)",               False),
    ("a0",  37, 38,  "CdbStatus1/2 (실시간 CDB 상태)",               True),
    ("a0",  39, 40,  "ActiveFWVersion (현재 활성 FW 버전)",           False),
    ("a0", 122,125,  "PasswordEntry 0x7A~0x7D (잠금 상태에 따라 변동)", True),
    ("a0", 126,127,  "PageSelectBank/Page (현재 페이지 번호)",        True),
    ("p00", 38, 53,  "VendorSN (DUT마다 고유 시리얼번호)",            True),
    ("p00", 54, 61,  "DateCode (제조일, DUT마다 다름)",               True),
    ("p00", 94, 94,  "P00h PageChecksum",                            False),
    ("p01",127,127,  "P01h PageChecksum",                            False),
    ("p02",127,127,  "P02h PageChecksum",                            False),
    ("p03",  0,127,  "P03h LaneMonitors (실시간 레인 모니터 값)",      True),
    ("p10",127,127,  "P10h PageChecksum",                            False),
    ("p11",127,127,  "P11h PageChecksum",                            False),
]

# h2         : 정수 → 2자리 대문자 16진수 문자열 (None이면 "--")
# asc_display: 바이트 배열 → ASCII 표시 문자열
#              0x00=·, 0x20~0x7E=그대로, 나머지=?
# parse_txt  : 탭 구분 txt 파일 → 페이지별 바이트 배열 dict
# build_txt  : 페이지별 바이트 배열 dict → txt 파일 문자열
def h2(v): return "--" if v is None else f"{int(v):02X}"

# ── 공통 디코딩 헬퍼 ──────────────────────────────────────
# s16 / u16 : P01h, P02h Decoded에서 중복 정의되던 로컬 함수를 모듈 레벨로 추출
def s16(hi, lo):
    """2바이트 Signed 16-bit 정수 변환 (CMIS big-endian)"""
    raw = (hi << 8) | lo
    return raw if (hi & 0x80) == 0 else -((~raw & 0xFFFF) + 1)

def u16(hi, lo):
    """2바이트 Unsigned 16-bit 정수 변환 (CMIS big-endian)"""
    return (hi << 8) | lo

def uw_to_dbm(raw_u16):
    """CMIS 광 파워: U16 × 0.1µW → dBm 변환"""
    uw = raw_u16 * 0.1          # 단위: µW
    if uw <= 0: return "No signal"
    return f"{10 * math.log10(uw * 1e-3):+.2f} dBm"

def i2c_8bit(addr_var_str):
    """i2c_addr_var 문자열(7bit hex) → 8bit 주소 정수
    예: "50" → 0xA0, "51" → 0xA2
    """
    return (int(addr_var_str.strip(), 16) & 0x7F) << 1
def asc_display(a,s,e):
    """표시용: 유효문자 + 공백은 그대로, 0x00은 · 로 표시"""
    r=""
    for i in range(s,e+1):
        c=a[i] if i<len(a) else 0
        if c==0x00: r+="·"
        elif 0x20<=c<=0x7E: r+=chr(c)
        else: r+="?"
    return r.rstrip("·").rstrip() or "(empty)"
_LBL_TO_KEY={"A0":"a0","A0h":"a0","P00h":"p00","P01h":"p01","P02h":"p02",
             "P03h":"p03","P10h":"p10","P11h":"p11"}
def _parse_col_map(header_parts):
    """헤더 열 이름 → {col_index: page_key} 매핑 생성"""
    col_map={}
    for ci,h in enumerate(header_parts):
        if h in _LBL_TO_KEY: col_map[ci]=_LBL_TO_KEY[h]
    if not col_map:  # 헤더 없는 구형 파일 fallback
        col_map={1:"a0",3:"p00",4:"p01",5:"p02",6:"p10",7:"p11",8:"p03"}
    return col_map
def parse_txt(text):
    data={k:[0]*128 for k in PAGE_KEYS}
    lines=text.strip().splitlines()
    if not lines: return data
    hdr=lines[0].strip().split()
    col_map=_parse_col_map(hdr)
    for line in lines[1:]:
        p=line.strip().split()
        if not p or p[0].lower()=="addr": continue
        try: a=int(p[0],16)
        except: continue
        if not(0<=a<128): continue
        # A0h 열 (col 1)
        if 1 in col_map and len(p)>1 and p[1]!="--":
            try: data[col_map[1]][a]=int(p[1],16)
            except: pass
        # 상위 페이지 열 (col 3+), 주소는 col 2로부터
        if len(p)>2:
            try: idx=int(p[2],16)-0x80
            except: continue
            if not(0<=idx<128): continue
            for ci,pk in col_map.items():
                if ci>=3 and ci<len(p) and p[ci]!="--":
                    try: data[pk][idx]=int(p[ci],16)
                    except: pass
    return data
def build_txt(data):
    lines=["Addr\tA0\tAddr\tP00h\tP01h\tP02h\tP10h\tP11h\tP03h"]
    for i in range(128):
        row=[h2(i),h2(data["a0"][i]),h2(i+0x80)]
        for pk in["p00","p01","p02","p10","p11","p03"]: row.append(h2(data[pk][i]))
        lines.append("\t".join(row))
    return "\n".join(lines)

def parse_xlsx(path):
    if not _OPENPYXL_OK:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")
    data={k:[0]*128 for k in PAGE_KEYS}
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
    ws=wb.active
    col_map=None
    for ri,row in enumerate(ws.iter_rows(values_only=True)):
        if not row: continue
        if ri==0:  # 헤더 행으로 열 매핑 생성
            hdr=[str(c).strip() if c is not None else "" for c in row]
            col_map=_parse_col_map(hdr)
            continue
        if row[0] is None: continue
        try: a=int(str(row[0]).strip(),16)
        except: continue
        if not(0<=a<128): continue
        if col_map is None: col_map=_parse_col_map([])
        if 1 in col_map and len(row)>1 and row[1] is not None:
            try: data[col_map[1]][a]=int(str(row[1]).strip(),16)
            except: pass
        if len(row)>2 and row[2] is not None:
            try: idx=int(str(row[2]).strip(),16)-0x80
            except: continue
            if not(0<=idx<128): continue
            for ci,pk in col_map.items():
                if ci>=3 and ci<len(row) and row[ci] is not None:
                    v=str(row[ci]).strip()
                    if v and v!="--":
                        try: data[pk][idx]=int(v,16)
                        except: pass
    wb.close()
    return data

def build_xlsx(data,path):
    if not _OPENPYXL_OK:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="EEPROM"
    headers=["Addr","A0","Addr","P00h","P01h","P02h","P10h","P11h","P03h"]
    hdr_font=Font(bold=True,color="FFFFFF",name="Consolas")
    hdr_fill=PatternFill("solid",fgColor="2F5496")
    ctr=Alignment(horizontal="center")
    for ci,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=ci,value=h)
        cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=ctr
    for i in range(128):
        row_vals=[h2(i),h2(data["a0"][i]),h2(i+0x80)]
        for pk in["p00","p01","p02","p10","p11","p03"]: row_vals.append(h2(data[pk][i]))
        fill=PatternFill("solid",fgColor="F2F2F2" if i%2==0 else "FFFFFF")
        for ci,v in enumerate(row_vals,1):
            cell=ws.cell(row=i+2,column=ci,value=v)
            cell.alignment=ctr; cell.font=Font(name="Consolas"); cell.fill=fill
    for ci in range(1,len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=8
    ws.freeze_panes="A2"
    wb.save(path)

# ── CP2112 USB-HID I2C (SLABHIDtoSMBus.dll ctypes) ───────
# 패킷 분석 결과 기반:
#   - slaveAddress: 8bit 주소 그대로 전달 (0xA0)
#   - 128바이트 일괄 읽기 (AddressReadRequest 1회)
#   - GetReadResponse를 반복 호출하여 61바이트씩 수신
class CP2112I2C:
    """CP2112-F03-GM SLABHIDtoSMBus.dll ctypes 드라이버"""
    HID_SMBUS_SUCCESS = 0x00
    XFER_IDLE     = 0x00
    XFER_BUSY     = 0x01
    XFER_COMPLETE = 0x02
    XFER_ERROR    = 0x03

    def __init__(self, device_index=0):
        if not HID_OK or _SLAB_DLL is None:
            raise RuntimeError("SLABHIDtoSMBus.dll 없음. 스크립트와 같은 폴더에 복사하세요.")
        self._dll = _SLAB_DLL
        self._handle = ctypes.c_void_p(0)
        self._is_open = False
        # HidSmbus_Open(HID_SMBUS_DEVICE* device, DWORD deviceNum, WORD vid, WORD pid)
        status = self._dll.HidSmbus_Open(
            ctypes.byref(self._handle),
            ctypes.c_uint32(device_index),
            ctypes.c_uint16(CP2112_VID),
            ctypes.c_uint16(CP2112_PID))
        if status != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_Open 실패 (status=0x{status:02X})")
        self._is_open = True
        self._configure()

    def _configure(self):
        """HidSmbus_SetSmbusConfig: 400kHz, writeTimeout=1000, readTimeout=1000, retries=3
        sclLowTimeout=False: EEPROM NVM write 중 모듈이 SCL 클럭 스트레칭을 사용하므로
        CP2112가 SCL HIGH까지 대기하도록 함. writeTimeout(1000ms)이 안전망 역할.
        (True 시 25ms 초과하면 S1=0x00 오류 → NVM write 중 false alarm 발생)
        """
        st = self._dll.HidSmbus_SetSmbusConfig(
            self._handle,
            ctypes.c_uint32(400000),
            ctypes.c_uint8(0x02),
            ctypes.c_int(0),       # autoReadRespond=False
            ctypes.c_uint16(1000),
            ctypes.c_uint16(1000),
            ctypes.c_int(0),       # sclLowTimeout=False (NVM 클럭 스트레칭 허용)
            ctypes.c_uint16(3))
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_SetSmbusConfig 실패 (0x{st:02X})")

    def _check_open(self):
        if not self._is_open:
            raise IOError("CP2112 미연결. 재연결하세요.")

    def _wait_complete(self, timeout_ms=2000):
        """TransferStatusRequest + GetTransferStatusResponse 폴링"""
        self._check_open()
        deadline = time.time() + timeout_ms/1000.0
        ts = ctypes.c_uint8(0)
        ds = ctypes.c_uint8(0)
        nr = ctypes.c_uint16(0)
        br = ctypes.c_uint16(0)
        while time.time() < deadline:
            self._dll.HidSmbus_TransferStatusRequest(self._handle)
            st = self._dll.HidSmbus_GetTransferStatusResponse(
                self._handle,
                ctypes.byref(ts), ctypes.byref(ds),
                ctypes.byref(nr), ctypes.byref(br))
            if st != self.HID_SMBUS_SUCCESS:
                raise IOError(f"GetTransferStatusResponse 실패 (0x{st:02X})")
            v = ts.value
            if v == self.XFER_COMPLETE: return br.value
            if v == self.XFER_ERROR:
                raise IOError(
                    f"I2C 전송 오류 (S1=0x{ds.value:02X}). "
                    f"배선/주소/풀업저항 확인 필요")
            time.sleep(0.002)
        raise IOError(f"CP2112 타임아웃 ({timeout_ms}ms)")

    def close(self):
        self._is_open = False
        try: self._dll.HidSmbus_Close(self._handle)
        except: pass

    def write_byte(self, i2c_addr_8bit, reg, value):
        """1바이트 쓰기 + ACK polling
        CMIS NVM 모듈은 기입 중 주소 NACK(S1=0x00)으로 busy를 표현.
        ACK polling: NACK 시 재시도 → 모듈이 준비되면 ACK → 성공.
        """
        self._check_open()
        buf = (ctypes.c_uint8 * 2)(reg & 0xFF, value & 0xFF)
        addr_byte = ctypes.c_uint8(i2c_addr_8bit & 0xFE)
        NVM_POLL_INTERVAL = 0.002   # 2ms 간격 폴링
        NVM_POLL_RETRIES  = 50      # 최대 50회 = 100ms 대기
        for attempt in range(NVM_POLL_RETRIES):
            st = self._dll.HidSmbus_WriteRequest(
                self._handle, addr_byte, buf, ctypes.c_uint8(2))
            if st != self.HID_SMBUS_SUCCESS:
                raise IOError(f"HidSmbus_WriteRequest 실패 (0x{st:02X})")
            try:
                self._wait_complete()
                return True   # ACK → 성공
            except IOError as e:
                if "S1=0x00" in str(e):
                    # NVM 기입 중 주소 NACK → 잠시 대기 후 재시도
                    time.sleep(NVM_POLL_INTERVAL)
                    continue
                raise   # 다른 오류는 즉시 raise
        raise IOError(
            f"NVM write timeout: reg=0x{reg:02X} "
            f"({NVM_POLL_RETRIES}회 ACK polling 실패)")

    def write_burst(self, i2c_addr_8bit, start_reg, data_list):
        """연속 주소 쓰기 — byte-by-byte, ACK polling 내장
        write_byte에 ACK polling이 포함되어 있으므로
        NVM 기입 중 주소 NACK 시 자동으로 대기 후 재시도.
        """
        self._check_open()
        for i, val in enumerate(data_list):
            reg = (start_reg + i) & 0xFF
            try:
                self.write_byte(i2c_addr_8bit, reg, val)
            except IOError as e:
                raise IOError(
                    f"I2C 전송 오류 reg=0x{reg:02X} "
                    f"(byte {i}/{len(data_list)}): {e}")
        return True

    def read_page(self, i2c_addr_8bit, start_reg, num_bytes=128):
        """
        여러 바이트 읽기 (LabVIEW VI와 동일한 시퀀스)
        1. HidSmbus_AddressReadRequest  - 읽기 요청
        2. TransferStatusRequest + GetTransferStatusResponse - 완료 대기
        3. HidSmbus_ForceReadResponse   - 읽기 응답 강제 요청
        4. HidSmbus_GetReadResponse     - 61바이트씩 반복 수신
        """
        self._check_open()
        target_addr = (ctypes.c_uint8 * 16)(start_reg & 0xFF)
        st = self._dll.HidSmbus_AddressReadRequest(
            self._handle,
            ctypes.c_uint8(i2c_addr_8bit & 0xFE),
            ctypes.c_uint16(num_bytes),
            ctypes.c_uint8(1),
            target_addr)
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_AddressReadRequest 실패 (0x{st:02X})")
        # 전송 완료 대기
        self._wait_complete()
        # ForceReadResponse: 장치에게 데이터를 보내달라고 요청
        # LabVIEW GetReadData.vi: TransferStatus 후 ForceReadResponse 호출
        remaining = num_bytes
        result = []
        while remaining > 0:
            chunk = min(remaining, 61)
            st_f = self._dll.HidSmbus_ForceReadResponse(
                self._handle,
                ctypes.c_uint16(chunk))
            if st_f != self.HID_SMBUS_SUCCESS:
                break
            # GetReadResponse로 수신
            rs  = ctypes.c_uint8(0)
            buf = (ctypes.c_uint8 * 61)()
            nr  = ctypes.c_uint8(0)
            st2 = self._dll.HidSmbus_GetReadResponse(
                self._handle,
                ctypes.byref(rs), buf,
                ctypes.c_uint8(61), ctypes.byref(nr))
            if st2 != self.HID_SMBUS_SUCCESS or nr.value == 0:
                break
            result.extend(buf[j] for j in range(nr.value))
            remaining -= nr.value
        # 부족하면 0으로 채움
        while len(result) < num_bytes:
            result.append(0)
        return result[:num_bytes]

    def read_byte(self, i2c_addr_8bit, reg):
        """1바이트 읽기 (호환성 유지)"""
        result = self.read_page(i2c_addr_8bit, reg, 1)
        if not result:
            raise IOError(f"읽기 데이터 없음 reg=0x{reg:02X}")
        return result[0]

    def set_page(self, page_num):
        """CMIS 페이지 전환: Byte 0x7F에 페이지 번호 쓰기"""
        self.write_byte(I2C_ADDR_A0, 0x7F, page_num)
        time.sleep(0.01)

    def set_gpio_config(self, direction, mode, function=0x00, clk_div=0x00):
        """GPIO 설정
        HidSmbus_SetGpioConfig(device, BYTE direction, BYTE mode,
                               BYTE function, BYTE clkDiv)
        direction: bit=1→Output, bit=0→Input (GPIO0=bit0 ... GPIO7=bit7)
        mode:      bit=1→Push-Pull, bit=0→Open-Drain
        function:  특수기능 (TX/RX Toggle, Clock Out)
        clk_div:   Clock Output Divider (0~255)
        """
        self._check_open()
        st = self._dll.HidSmbus_SetGpioConfig(
            self._handle,
            ctypes.c_uint8(direction),
            ctypes.c_uint8(mode),
            ctypes.c_uint8(function),
            ctypes.c_uint8(clk_div))
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_SetGpioConfig 실패 (0x{st:02X})")

    def write_latch(self, latch_value, latch_mask=0xFF):
        """GPIO Latch 값 설정
        HidSmbus_WriteLatch(device, BYTE latchValue, BYTE latchMask)
        latch_value: 설정할 GPIO 값 (bit=1→High, bit=0→Low)
        latch_mask:  변경할 GPIO 비트 마스크
        """
        self._check_open()
        st = self._dll.HidSmbus_WriteLatch(
            self._handle,
            ctypes.c_uint8(latch_value),
            ctypes.c_uint8(latch_mask))
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_WriteLatch 실패 (0x{st:02X})")

    def reset(self):
        """CP2112 리셋
        HidSmbus_Reset(device) - USB 재열거 발생
        """
        self._check_open()
        self._dll.HidSmbus_Reset(self._handle)
        self._is_open = False  # 리셋 후 핸들 무효화

    def scan_addresses(self, start=0x08, end=0x77):
        """I2C 주소 스캔 (7bit 범위 → 8bit로 변환하여 시도)"""
        self._check_open()
        found = []
        for addr_7bit in range(start, end+1):
            addr_8bit = addr_7bit << 1
            try:
                buf = (ctypes.c_uint8 * 1)(0x00)
                st = self._dll.HidSmbus_WriteRequest(
                    self._handle,
                    ctypes.c_uint8(addr_8bit),
                    buf, ctypes.c_uint8(1))
                if st != self.HID_SMBUS_SUCCESS:
                    continue
                ts = ctypes.c_uint8(0)
                ds = ctypes.c_uint8(0)
                nr = ctypes.c_uint16(0)
                br = ctypes.c_uint16(0)
                deadline = time.time() + 0.15
                while time.time() < deadline:
                    self._dll.HidSmbus_TransferStatusRequest(self._handle)
                    self._dll.HidSmbus_GetTransferStatusResponse(
                        self._handle,
                        ctypes.byref(ts), ctypes.byref(ds),
                        ctypes.byref(nr), ctypes.byref(br))
                    if ts.value in (0x02, 0x03, 0x00): break
                    time.sleep(0.002)
                if ts.value == 0x02:
                    found.append(addr_7bit)
            except Exception:
                continue
        return found

    @staticmethod
    def list_devices():
        """연결된 CP2112 장치 목록 반환"""
        if not HID_OK or _SLAB_DLL is None: return []
        num = ctypes.c_uint32(0)
        _SLAB_DLL.HidSmbus_GetNumDevices(
            ctypes.byref(num),
            ctypes.c_uint16(CP2112_VID),
            ctypes.c_uint16(CP2112_PID))
        result = []
        for i in range(num.value):
            vid = ctypes.c_uint16(0)
            pid = ctypes.c_uint16(0)
            rel = ctypes.c_uint16(0)
            _SLAB_DLL.HidSmbus_GetAttributes(
                ctypes.c_uint32(i),
                ctypes.c_uint16(CP2112_VID), ctypes.c_uint16(CP2112_PID),
                ctypes.byref(vid), ctypes.byref(pid), ctypes.byref(rel))
            sn_buf = ctypes.create_string_buffer(260)
            _SLAB_DLL.HidSmbus_GetString(
                ctypes.c_uint32(i),
                ctypes.c_uint16(CP2112_VID), ctypes.c_uint16(CP2112_PID),
                sn_buf, ctypes.c_uint32(0x04))
            sn = sn_buf.value.decode("ascii","ignore").strip()
            prod = f"CP2112"
            if sn: prod += f" S/N:{sn}"
            result.append({"product_string": prod,
                           "serial_number":  sn,
                           "manufacturer_string": "Silicon Labs"})
        return result

# ── CanvasTable ───────────────────────────────────────────
class CanvasTable(tk.Frame):
    """
    Canvas 기반 커스텀 테이블 위젯.
    tkinter 기본 Treeview 대신 사용하는 이유:
      - Treeview는 열별 배경색 독립 제어 불가
      - Treeview는 pack 레이아웃에서 열 너비가 자동 축소됨
      - Canvas는 픽셀 단위 완전 제어 가능

    주요 기능:
      - 헤더 경계 드래그 → 열 너비 조절
      - 데이터 셀 클릭/드래그 → 범위 선택
      - Ctrl+C → 선택 범위 클립보드 복사
      - Value[hex] 열 더블클릭 → 바이트별 편집 팝업

    상수:
      ROW_H    : 데이터 행 높이 (px)
      HEAD_H   : 헤더 행 높이 (px)
      PAD_X    : 텍스트 좌측 여백 (px)
      SEP_W    : 열 구분선 두께 (px)
      DRAG_TOL : 열 경계 감지 허용 범위 (px)
    """
    ROW_H=24; HEAD_H=28; PAD_X=6; SEP_W=1; DRAG_TOL=5

    def __init__(self,parent,col_cfg,theme,**kw):
        super().__init__(parent,**kw)
        self._col_cfg=col_cfg; self._theme=theme; self.theme=theme
        self._rows=[]; self._col_w=[c[1] for c in col_cfg]
        self._sel_idx=-1; self._dblclick_cb=None; self._scroll_y=0
        self._scroll_x=0   # 가로 스크롤 오프셋
        self._sel_r0=self._sel_r1=self._sel_c0=self._sel_c1=None
        self._drag_ci=-1; self._drag_x0=0; self._drag_w0=0
        self._col_dragging=False; self._cell_dragging=False
        self._decoded_edit_cb=None
        self._col_resize_cb=None
        self._row_heights=[]; self._row_ys=[]; self._total_rows_h=0
        self._build()

    def _build(self):
        t=self._theme
        self.config(bg=t.get("col_sep","#404040"))
        self._vsb=ttk.Scrollbar(self,orient=tk.VERTICAL,command=self._on_vscroll)
        self._vsb.pack(side=tk.RIGHT,fill=tk.Y)
        self._hsb=ttk.Scrollbar(self,orient=tk.HORIZONTAL,command=self._on_hscroll)
        self._hsb.pack(side=tk.BOTTOM,fill=tk.X)
        self._cv=tk.Canvas(self,highlightthickness=0,bd=0)
        self._cv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        cv=self._cv
        cv.bind("<Configure>",self._on_resize)
        cv.bind("<Motion>",self._on_motion)
        cv.bind("<ButtonPress-1>",self._on_press)
        cv.bind("<B1-Motion>",self._on_drag)
        cv.bind("<ButtonRelease-1>",self._on_release)
        cv.bind("<Double-1>",self._on_dbl)
        cv.bind("<MouseWheel>",self._on_wheel)
        cv.bind("<Button-4>",self._on_wheel)
        cv.bind("<Button-5>",self._on_wheel)
        cv.bind("<Shift-MouseWheel>",self._on_hwheel)
        cv.bind("<Shift-Button-4>",self._on_hwheel)
        cv.bind("<Shift-Button-5>",self._on_hwheel)
        cv.bind("<Control-c>",self._copy_sel)
        cv.bind("<Control-C>",self._copy_sel)
        cv.config(takefocus=True)
        # 바이트별 편집 팝업 (Value[hex] 더블클릭 시)
        self._byte_popup = None   # tk.Toplevel
        self._byte_entries = []   # 바이트별 Entry 리스트
        self._byte_vals = []      # 현재 바이트 값 리스트
        self._edit_row = -1
        self._edit_cb  = None
        self._decode_preview_cb = None   # Decoded 미리보기 콜백
        self._popup_decoded_var = None   # 팝업 Decoded 라벨 StringVar
        self._popup_pk  = ""             # 팝업 열린 페이지
        self._popup_start_idx = 0        # 팝업 필드 시작 idx

    # ── 스크롤 ──────────────────────────────────────────
    def _on_vscroll(self,*args):
        total=max(1,self._total_rows_h+self.HEAD_H)
        cv_h=max(self._cv.winfo_height(),1)
        max_y=max(0,total-cv_h)
        if args[0]=="moveto": self._scroll_y=int(float(args[1])*total)
        elif args[0]=="scroll":
            u=self.ROW_H if args[2]=="units" else cv_h
            self._scroll_y+=int(args[1])*u
        self._scroll_y=max(0,min(self._scroll_y,max_y))
        self._upd_sb(); self._redraw()

    def _on_hscroll(self,*args):
        total_w=max(1,sum(self._col_w)+self.SEP_W*len(self._col_w))
        cv_w=max(self._cv.winfo_width(),1)
        max_x=max(0,total_w-cv_w)
        if args[0]=="moveto": self._scroll_x=int(float(args[1])*total_w)
        elif args[0]=="scroll":
            u=30 if args[2]=="units" else cv_w
            self._scroll_x+=int(args[1])*u
        self._scroll_x=max(0,min(self._scroll_x,max_x))
        self._upd_sb(); self._redraw()

    def _on_wheel(self,event):
        d=-3 if(event.num==4 or event.delta>0)else 3
        self._scroll_y=max(0,self._scroll_y+d*self.ROW_H)
        self._upd_sb(); self._redraw()

    def _on_hwheel(self,event):
        d=-3 if(event.num==4 or event.delta>0)else 3
        cv_w=max(self._cv.winfo_width(),1)
        total_w=max(1,sum(self._col_w)+self.SEP_W*len(self._col_w))
        max_x=max(0,total_w-cv_w)
        self._scroll_x=max(0,min(self._scroll_x+d*20,max_x))
        self._upd_sb(); self._redraw()

    def _upd_sb(self):
        total=max(1,self._total_rows_h+self.HEAD_H)
        h=max(self._cv.winfo_height(),1)
        self._vsb.set(self._scroll_y/total,min(1.0,(self._scroll_y+h)/total))
        # 가로 스크롤바
        total_w=max(1,sum(self._col_w)+self.SEP_W*len(self._col_w))
        w=max(self._cv.winfo_width(),1)
        self._hsb.set(self._scroll_x/total_w,min(1.0,(self._scroll_x+w)/total_w))

    def _on_resize(self,event): self._recalc_stretch(); self._redraw()

    # ── 열 너비 계산 ─────────────────────────────────────
    # _recalc_stretch: stretch=True 열에 남은 공간 배분
    # fit_col_widths : 데이터 기준 각 열 최적 너비 계산
    def _recalc_stretch(self):
        cw=self._cv.winfo_width()
        if cw<10: return
        fixed=sum(self._col_w[ci] for ci,c in enumerate(self._col_cfg) if not c[4])
        nst=sum(1 for c in self._col_cfg if c[4])
        if nst:
            rem=max(80,cw-fixed-self.SEP_W*len(self._col_cfg))
            sw=rem//nst
            for ci,c in enumerate(self._col_cfg):
                if c[4]: self._col_w[ci]=max(sw,c[2])

    def fit_col_widths(self,rows_data):
        CW=8; PAD=20; n=len(self._col_cfg)
        cw=[len(self._col_cfg[ci][0])*CW+PAD for ci in range(n)]
        for row in rows_data:
            for ci in range(n):
                txt=str(row[ci]) if ci<len(row) else ''
                w=len(txt)*CW+PAD
                if w>cw[ci]: cw[ci]=w
        for ci,c in enumerate(self._col_cfg):
            if not c[4]: self._col_w[ci]=max(cw[ci],c[2])
        self._recalc_stretch()

    def _calc_row_heights(self):
        """열 너비와 텍스트 길이 기반 행 높이 계산 (줄바꿈 지원)"""
        CHAR_W=7.5; LINE_H=15; V_PAD=6; min_h=self.ROW_H
        ys=[]; heights=[]; y=0
        for vals,tag in self._rows:
            max_lines=1
            for ci in range(len(self._col_cfg)):
                avail=max(1,self._col_w[ci]-2*self.PAD_X)
                text=str(vals[ci]) if ci<len(vals) else ''
                if not text: continue
                chars_per_line=max(1,avail/CHAR_W)
                n=max(1,math.ceil(len(text)/chars_per_line))
                if n>max_lines: max_lines=n
            h=max(min_h,max_lines*LINE_H+V_PAD)
            ys.append(y); heights.append(h); y+=h
        self._row_ys=ys; self._row_heights=heights; self._total_rows_h=y

    def bind_col_resize(self,cb): self._col_resize_cb=cb

    # ── 데이터 관리 ──────────────────────────────────────
    # delete_all  : 모든 행 삭제 및 상태 초기화
    # insert_row  : 행 추가 (commit() 전까지 화면 갱신 없음)
    # commit      : insert_row 완료 후 한번 호출 → 화면 갱신
    def delete_all(self):
        self._rows=[]; self._sel_idx=-1; self._scroll_y=0
        self._sel_r0=self._sel_r1=self._sel_c0=self._sel_c1=None
        self._cell_dragging=False
        self._row_heights=[]; self._row_ys=[]; self._total_rows_h=0
        self._cv.delete("all"); self._upd_sb()

    def insert_row(self,values,tag='even'): self._rows.append((values,tag))

    def commit(self):
        self._calc_row_heights(); self._recalc_stretch(); self._upd_sb(); self._redraw()

    def get_selected_values(self):
        if 0<=self._sel_idx<len(self._rows): return self._rows[self._sel_idx]
        return None

    def bind_dblclick(self,cb): self._dblclick_cb=cb

    # ── 좌표 헬퍼 ────────────────────────────────────────
    # _col_x  : ci번 열의 시작 x 픽셀 좌표
    # _row_at : y 픽셀 → 행 인덱스 (-1이면 행 없음)
    # _ci_at  : x 픽셀 → 열 인덱스
    # _sep_at : x 픽셀 근처 열 구분선 존재 여부 (열 인덱스 반환)
    def _col_x(self,ci):
        x=0
        for i in range(ci): x+=self._col_w[i]+self.SEP_W
        return x - self._scroll_x

    def _row_at(self,y):
        dy=y+self._scroll_y-self.HEAD_H
        if dy<0 or not self._row_ys: return -1
        lo,hi=0,len(self._row_ys)-1
        while lo<=hi:
            mid=(lo+hi)//2
            h=self._row_heights[mid] if mid<len(self._row_heights) else self.ROW_H
            if dy<self._row_ys[mid]: hi=mid-1
            elif dy>=self._row_ys[mid]+h: lo=mid+1
            else: return mid
        return -1

    def _ci_at(self,x):
        for ci in range(len(self._col_cfg)):
            cx=self._col_x(ci)
            if cx<=x<cx+self._col_w[ci]: return ci
        return len(self._col_cfg)-1

    def _sep_at(self,x):
        for ci in range(len(self._col_cfg)-1):
            sx=self._col_x(ci)+self._col_w[ci]
            if abs(x-sx)<=self.DRAG_TOL: return ci
        return -1

    # ── 마우스 이벤트 ─────────────────────────────────────
    # _on_motion  : 커서 모양 변경 (구분선 근처=↔, 데이터=I빔)
    # _on_press   : 헤더=열 크기 드래그 시작 / 데이터=셀 선택 시작
    # _on_drag    : 열 크기 조절 또는 셀 범위 확장
    # _on_release : 드래그 상태 해제
    # _on_dbl     : Value[hex] 열=바이트 팝업 / 나머지=편집바 사용
    def _on_motion(self,event):
        if self._col_dragging: return
        if event.y<=self.HEAD_H and self._sep_at(event.x)>=0:
            self._cv.config(cursor="sb_h_double_arrow")
        elif event.y>self.HEAD_H:
            self._cv.config(cursor="xterm")
        else:
            self._cv.config(cursor="arrow")

    def _on_press(self,event):
        self._cv.focus_set()
        if event.y<=self.HEAD_H:
            ci=self._sep_at(event.x)
            if ci>=0:
                self._drag_ci=ci; self._drag_x0=event.x
                self._drag_w0=self._col_w[ci]; self._col_dragging=True
        else:
            ri=self._row_at(event.y); ci=self._ci_at(event.x)
            if ri>=0:
                self._sel_r0=self._sel_r1=ri
                self._sel_c0=self._sel_c1=ci
                self._sel_idx=ri; self._cell_dragging=True
                self._redraw()

    def _on_drag(self,event):
        if self._col_dragging:
            ci=self._drag_ci; mn=self._col_cfg[ci][2]
            self._col_w[ci]=max(mn,self._drag_w0+event.x-self._drag_x0)
            self._recalc_stretch(); self._redraw()
        elif self._cell_dragging:
            ri=self._row_at(event.y)
            if ri<0: ri=max(0,min(len(self._rows)-1,
                (event.y+self._scroll_y-self.HEAD_H)//self.ROW_H))
            self._sel_r1=ri; self._sel_c1=self._ci_at(event.x)
            self._redraw()

    def _on_release(self,event):
        was_col=self._col_dragging
        self._col_dragging=False; self._cell_dragging=False
        self._cv.config(cursor="arrow")
        if was_col:
            self._calc_row_heights(); self._upd_sb(); self._redraw()
            if self._col_resize_cb: self._col_resize_cb(list(self._col_w))

    def _on_dbl(self,event):
        if event.y<self.HEAD_H or self._sep_at(event.x)>=0: return
        ri=self._row_at(event.y)
        if ri<0: return
        self._sel_idx=ri; self._redraw()
        ci=self._ci_at(event.x)
        if ci==2:  # Value[hex] 열 → 인라인 편집
            self._start_inline_edit(ri, event)
        elif ci==3:  # Decoded 열 → 드롭다운 or 텍스트 편집
            if self._decoded_edit_cb: self._decoded_edit_cb(ri)
        else:
            if self._dblclick_cb: self._dblclick_cb(ri)

    # ── 바이트 팝업 편집 ──────────────────────────────────
    # _start_inline_edit : Value[hex] 셀 위치에 Toplevel 팝업 생성
    #   - 바이트별 Entry 그리드 (한 줄 8바이트)
    #   - Enter=단일 바이트 확정+다음이동 / Tab=이동 / Esc=취소
    # _byte_commit_one   : 단일 바이트 즉시 반영 (byte_offset 포함)
    # _byte_commit_all   : 전체 바이트 일괄 반영
    def _start_inline_edit(self, ri, event=None):
        """바이트별 Entry 팝업 표시 — event가 있으면 마우스 근처에 열림"""
        if ri<0 or ri>=len(self._rows): return
        self._close_byte_popup()
        vals=self._rows[ri][0]
        cur_val=str(vals[2]) if len(vals)>2 else "00"
        bytes_list=cur_val.split()
        if not bytes_list: return
        self._edit_row=ri
        self._byte_vals=list(bytes_list)
        # 시작 주소 파싱
        try:
            addr_str=str(vals[0]).split("~")[0].strip()
            base_addr=int(addr_str)
        except:
            base_addr=0
        # 팝업 pk / start_idx 저장 (Decoded 미리보기용)
        try:
            addr_hex=str(vals[1]).replace("h","").split("~")[0].strip()
            abs_addr=int(addr_hex,16)
            self._popup_start_idx=abs_addr
        except:
            self._popup_start_idx=base_addr
        # ── 팝업 기준 좌표: 마우스 스크린 절대 좌표 사용 ──────
        # event.x_root / y_root = 스크린 절대 좌표 (가장 정확)
        # fallback: canvas rootx + col2_x
        if event is not None:
            rx  = event.x_root
            ry2 = event.y_root + self.ROW_H
        else:
            col2_x  = self._col_x(2)
            rh = self._row_heights[ri] if ri<len(self._row_heights) else self.ROW_H
            ry_row  = self.HEAD_H + (self._row_ys[ri] if ri<len(self._row_ys) else ri*self.ROW_H) - self._scroll_y
            rx  = self._cv.winfo_rootx() + col2_x
            ry2 = self._cv.winfo_rooty() + ry_row + rh
        # Toplevel 팝업 생성
        t=self.theme if hasattr(self,"theme") else {}
        bg=t.get("bg1","#f0f0f0") if t else "#f0f0f0"
        fg_c=t.get("t1","#111111") if t else "#111111"
        acc=t.get("acc","#185FA5") if t else "#185FA5"
        t2_c=t.get("t2","#444444") if t else "#444444"
        pop=tk.Toplevel(self)
        pop.overrideredirect(True)
        pop.config(bg=acc)
        pop.attributes("-topmost",True)
        self._byte_popup=pop
        # 헤더 라벨
        hdr=tk.Frame(pop,bg=acc); hdr.pack(fill=tk.X,padx=1,pady=(1,0))
        end_addr=base_addr+len(bytes_list)-1
        addr_range=f"{base_addr}" if len(bytes_list)==1 else f"{base_addr}~{end_addr}"
        tk.Label(hdr,text=f"Byte 편집  Addr:{addr_range} ({len(bytes_list)}바이트)  ⠿드래그 이동",
                 bg=acc,fg="white",font=("",8),cursor="fleur").pack(side=tk.LEFT,padx=4)
        tk.Button(hdr,text="✕",bg=acc,fg="white",relief=tk.FLAT,
                  font=("",8),cursor="hand2",
                  command=self._close_byte_popup).pack(side=tk.RIGHT,padx=2)
        # 헤더 드래그로 팝업 이동
        self._popup_drag_x=0; self._popup_drag_y=0
        def _drag_start(e):
            self._popup_drag_x=e.x_root; self._popup_drag_y=e.y_root
        def _drag_move(e):
            dx=e.x_root-self._popup_drag_x; dy=e.y_root-self._popup_drag_y
            x=pop.winfo_x()+dx; y=pop.winfo_y()+dy
            pop.geometry(f"+{x}+{y}")
            self._popup_drag_x=e.x_root; self._popup_drag_y=e.y_root
        hdr.bind("<Button-1>",_drag_start)
        hdr.bind("<B1-Motion>",_drag_move)
        for w in hdr.winfo_children():
            if isinstance(w, tk.Label):
                w.bind("<Button-1>",_drag_start)
                w.bind("<B1-Motion>",_drag_move)
        # 바이트 Entry 그리드
        gf=tk.Frame(pop,bg=bg); gf.pack(padx=1,pady=(0,1))
        self._byte_entries=[]
        COLS=8
        for bi,bv in enumerate(bytes_list):
            row_f=tk.Frame(gf,bg=bg)
            if bi%COLS==0: row_f.grid(row=bi//COLS*2+1,column=0,
                                       columnspan=COLS*2,sticky="w")
            col=bi%COLS
            r=bi//COLS
            tk.Label(gf,text=str(base_addr+bi),bg=bg,fg=acc,
                     font=("Consolas",8)).grid(row=r*2,column=col,padx=2,pady=(4,0))
            var=tk.StringVar(value=bv)
            e=tk.Entry(gf,textvariable=var,width=3,
                       font=("Consolas",11),justify="center",
                       bg=bg,fg=fg_c,
                       relief="solid",bd=1,
                       highlightthickness=1,
                       highlightcolor=acc,
                       highlightbackground="#aaaaaa")
            e.grid(row=r*2+1,column=col,padx=2,pady=(0,4))
            e.bind("<Return>",    lambda ev,i=bi: self._byte_commit_one(i))
            e.bind("<KP_Enter>",  lambda ev,i=bi: self._byte_commit_one(i))
            e.bind("<Tab>",       lambda ev,i=bi: self._byte_tab(i))
            e.bind("<Escape>",    lambda ev: self._close_byte_popup())
            e.bind("<KeyRelease>",lambda ev: self._update_decoded_preview())
            self._byte_entries.append((e,var))
        # OK 버튼
        bf=tk.Frame(pop,bg=bg); bf.pack(fill=tk.X,padx=4,pady=(0,2))
        tk.Button(bf,text="전체 적용",font=("",9),cursor="hand2",
                  bg=acc,fg="white",relief=tk.FLAT,
                  command=self._byte_commit_all).pack(side=tk.RIGHT,padx=4)
        # ── Decoded 미리보기 패널 ──────────────────────────
        if self._decode_preview_cb:
            sep=tk.Frame(pop,bg=acc,height=1); sep.pack(fill=tk.X,padx=1)
            df=tk.Frame(pop,bg=bg); df.pack(fill=tk.X,padx=1,pady=(0,1))
            tk.Label(df,text="Decoded:",bg=bg,fg=t2_c,
                     font=("",8)).pack(side=tk.LEFT,padx=(6,2),pady=3)
            self._popup_decoded_var=tk.StringVar(value="—")
            tk.Label(df,textvariable=self._popup_decoded_var,
                     bg=bg,fg=acc,font=("Consolas",9,"bold"),
                     anchor=tk.W).pack(side=tk.LEFT,fill=tk.X,expand=True,
                                        padx=(0,6),pady=3)
            self._update_decoded_preview()
        # 위치 설정 — 화면 경계 처리
        pop.update_idletasks()
        pw=pop.winfo_reqwidth(); ph=pop.winfo_reqheight()
        sw=pop.winfo_screenwidth(); sh=pop.winfo_screenheight()
        # X: 팝업이 화면 오른쪽 밖으로 나가면 왼쪽으로 당김
        px = min(rx, sw-pw-4)
        px = max(0, px)
        # Y: 아래에 공간이 부족하면 클릭 위치 위쪽에 표시
        if ry2 + ph > sh - 4:
            py = ry2 - self.ROW_H - ph   # 행 위로 올림
        else:
            py = ry2
        py = max(0, py)
        pop.geometry(f"+{px}+{py}")
        if self._byte_entries:
            self._byte_entries[0][0].focus_set()
            self._byte_entries[0][0].select_range(0,"end")

    def _update_decoded_preview(self):
        """Entry 값 변경 시 Decoded 미리보기 라벨 갱신"""
        if not self._decode_preview_cb or not self._popup_decoded_var:
            return
        byte_list=[]
        for e,var in self._byte_entries:
            try:    byte_list.append(int(var.get().strip(),16)&0xFF)
            except: byte_list.append(0)
        try:
            txt=self._decode_preview_cb(
                self._popup_pk, self._popup_start_idx, byte_list)
            self._popup_decoded_var.set(txt or "—")
        except:
            self._popup_decoded_var.set("—")

    def _close_byte_popup(self,event=None):
        """바이트 팝업 닫기"""
        if self._byte_popup:
            try: self._byte_popup.destroy()
            except: pass
            self._byte_popup=None
        self._byte_entries=[]
        self._edit_row=-1
        self._cv.focus_set()

    def _byte_tab(self,bi):
        """Tab키로 다음 Entry 이동"""
        next_i=bi+1
        if next_i<len(self._byte_entries):
            self._byte_entries[next_i][0].focus_set()
            self._byte_entries[next_i][0].select_range(0,"end")

    def _byte_commit_one(self,bi):
        """단일 바이트 Enter: 현재 바이트만 즉시 반영 후 다음 이동"""
        if bi>=len(self._byte_entries): return
        e,var=self._byte_entries[bi]
        try: val=int(var.get().strip(),16)&0xFF
        except ValueError: return
        # 해당 바이트만 콜백 (offset 포함)
        if self._edit_cb and self._edit_row>=0:
            self._edit_cb(self._edit_row,[val],byte_offset=bi)
        # 다음 Entry로 이동 또는 팝업 닫기
        next_i=bi+1
        if next_i<len(self._byte_entries):
            self._byte_entries[next_i][0].focus_set()
            self._byte_entries[next_i][0].select_range(0,"end")
        else:
            self._close_byte_popup()

    def _byte_commit_all(self):
        """전체 적용 버튼: 모든 바이트 한번에 반영"""
        vals=[]
        for e,var in self._byte_entries:
            try: vals.append(int(var.get().strip(),16)&0xFF)
            except ValueError: vals.append(0)
        if self._edit_cb and self._edit_row>=0:
            self._edit_cb(self._edit_row,vals,byte_offset=0)
        self._close_byte_popup()

    def bind_edit(self,cb):
        """commit 콜백 등록: cb(row_index, new_byte_value)"""
        self._edit_cb=cb

    def bind_decode_preview(self,cb):
        """팝업 Decoded 미리보기 콜백 등록
        cb(pk, start_idx, byte_list) → decoded_string
        """
        self._decode_preview_cb=cb

    def bind_decoded_edit(self,cb):
        """Decoded 열 편집 콜백: cb(row_index)"""
        self._decoded_edit_cb=cb

    # ── 클립보드 복사 ─────────────────────────────────────
    # Ctrl+C: 선택된 셀 범위를 탭/개행으로 클립보드 복사
    #         선택 없으면 현재 행 전체 복사
    def _copy_sel(self,event=None):
        r0=self._sel_r0; r1=self._sel_r1
        c0=self._sel_c0; c1=self._sel_c1
        if r0 is None:
            if 0<=self._sel_idx<len(self._rows):
                vals=self._rows[self._sel_idx][0]
                self._clip(chr(9).join(str(v) for v in vals))
            return
        lines=[]
        for ri in range(min(r0,r1),max(r0,r1)+1):
            if ri>=len(self._rows): break
            vals=self._rows[ri][0]
            cells=[str(vals[ci]) if ci<len(vals) else ''
                   for ci in range(min(c0,c1),max(c0,c1)+1)]
            lines.append(chr(9).join(cells))
        self._clip(chr(10).join(lines))

    def _clip(self,text):
        try:
            self.clipboard_clear(); self.clipboard_append(text); self.update()
        except Exception: pass

    # ── 렌더링 ───────────────────────────────────────────
    # _redraw: Canvas 전체를 다시 그림
#   1) 전체 배경 → 2) 헤더(텍스트+구분선) → 3) 데이터 행(배경+텍스트+구분선)
    # 화면 밖 행은 스킵하여 성능 최적화
    def _redraw(self):
        t=self._theme; cv=self._cv
        cv.delete("all")
        cw=cv.winfo_width(); ch=cv.winfo_height()
        if cw<2 or ch<2: return
        n=len(self._col_cfg); sep=t.get('col_sep','#404040')
        cv.create_rectangle(0,0,cw,ch,fill=t.get('bg2','#cccccc'),outline='')
        r0=self._sel_r0; r1=self._sel_r1
        c0=self._sel_c0; c1=self._sel_c1; has_sel=r0 is not None
        a0e=t.get('col0_even','#ccd4e8'); a0o=t.get('col0_odd','#d8deee')
        d1e=t.get('col1_even','#c0c0c0'); d1o=t.get('col1_odd','#cccccc')
        sbg=t.get('sel_bg','#b8d4ee'); sfg=t.get('sel_fg','#0a3a70')
        dfg=t.get('dirty','#7a4000'); fg=t.get('t1','#111111')
        for ri,(vals,tag) in enumerate(self._rows):
            rh=self._row_heights[ri] if ri<len(self._row_heights) else self.ROW_H
            ry=self.HEAD_H+(self._row_ys[ri] if ri<len(self._row_ys) else ri*self.ROW_H)-self._scroll_y
            if ry+rh<0 or ry>ch: continue
            isdirt=(tag=='dirty'); iseven=(ri%2==0)
            row_in_sel=(has_sel and min(r0,r1)<=ri<=max(r0,r1))
            for ci in range(n):
                x=self._col_x(ci); w=self._col_w[ci]
                isaddr=(self._col_cfg[ci][3]==0)
                cell_sel=(row_in_sel and c0 is not None
                          and min(c0,c1)<=ci<=max(c0,c1))
                if cell_sel:    bg=sbg
                elif isdirt:    bg=sbg
                elif isaddr:    bg=a0e if iseven else a0o
                else:           bg=d1e if iseven else d1o
                cv.create_rectangle(x,ry,x+w,ry+rh,fill=bg,outline='')
                txt=str(vals[ci]) if ci<len(vals) else ''
                tfg=sfg if cell_sel else(dfg if isdirt else fg)
                wrap_w=max(1,w-2*self.PAD_X)
                cv.create_text(x+self.PAD_X,ry+rh//2,
                    text=txt,anchor='w',fill=tfg,font=('Consolas',10),
                    width=wrap_w)
                if ci<n-1:
                    lx=x+w; cv.create_line(lx,ry,lx,ry+rh,fill=sep,width=self.SEP_W)
            cv.create_line(0,ry+rh-1,cw,ry+rh-1,fill=sep,width=1)
        # ── 헤더를 마지막에 그려 스크롤 시에도 항상 최상단 고정 ──
        hbg=t.get('col_head','#b4b4b4'); hfg=t.get('acc','#185FA5')
        cv.create_rectangle(0,0,cw,self.HEAD_H,fill=hbg,outline='')
        for ci in range(n):
            x=self._col_x(ci); w=self._col_w[ci]
            cv.create_text(x+self.PAD_X,self.HEAD_H//2,
                text=self._col_cfg[ci][0],anchor='w',fill=hfg,font=('',10,'bold'))
            if ci<n-1:
                lx=x+w; cv.create_line(lx,0,lx,self.HEAD_H,fill=sep,width=2)
        cv.create_line(0,self.HEAD_H,cw,self.HEAD_H,fill=sep,width=2)

    def apply_theme(self,t):
        self._theme=t; self.config(bg=t.get('col_sep','#404040')); self._redraw()

    def set_col_width(self,ci,w):
        if 0<=ci<len(self._col_w): self._col_w[ci]=w


# ── 앱 아이콘 ─────────────────────────────────────────────
def _make_app_icon():
    """CMIS 텍스트 아이콘 생성. PIL 없으면 None 반환."""
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io, base64
        SIZE = 64
        img = Image.new("RGBA", (SIZE, SIZE), (0, 0, 0, 0))
        d = ImageDraw.Draw(img)
        bg, border, text_col = "#1e2a3a", "#7ab3f5", "#e0e0e0"
        try:
            d.rounded_rectangle([0, 0, SIZE-1, SIZE-1], radius=10, fill=bg, outline=border, width=2)
        except AttributeError:
            d.rectangle([0, 0, SIZE-1, SIZE-1], fill=bg, outline=border, width=2)
        font = None
        for fname in ["consola.ttf", "cour.ttf", "arial.ttf"]:
            try:
                font = ImageFont.truetype(fname, 20); break
            except: pass
        if font is None:
            font = ImageFont.load_default()
        text = "CMIS"
        try:
            bb = d.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
        except AttributeError:
            tw, th = d.textsize(text, font=font)
        d.text(((SIZE - tw) // 2, (SIZE - th) // 2 - 2), text, fill=text_col, font=font)
        cy = (SIZE - th) // 2 - 2 + th + 7
        for cx in [SIZE//2 - 8, SIZE//2, SIZE//2 + 8]:
            d.ellipse([cx-2, cy-2, cx+2, cy+2], fill=border)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return tk.PhotoImage(data=base64.b64encode(buf.getvalue()).decode())
    except Exception:
        return None


# ── 메인 앱 ───────────────────────────────────────────────
class App(tk.Tk):
    """
    메인 GUI 애플리케이션.

    상태 관리:
      self.data      : 현재 편집 중인 EEPROM 데이터 {page_key: [0]*128}
      self.orig_data : 파일 로드/EEPROM 읽기 직후 원본 데이터 (dirty 판별용)
      self.conn      : 연결된 I2C 객체 (CP2112I2C)
      self.current_page : 현재 뷰어에 표시 중인 페이지 키

    테마: 다크 모드 고정
    """
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME}  v{APP_VERSION}  |  {SPEC_CMIS}  |  {SPEC_SFF8024}")
        self.geometry("1280x820")
        self.minsize(1000,640)
        self._icon = _make_app_icon()
        if self._icon:
            self.iconphoto(True, self._icon)
        self._dark=False
        self.theme=make_theme(_LIGHT,1.0)
        self.data={k:[0]*128 for k in PAGE_KEYS}
        self.orig_data={k:[0]*128 for k in PAGE_KEYS}
        self.filename=""
        self.conn=None
        self._device_index=-1
        self.current_page="a0"
        self._col_w_cache={}  # 페이지별 열 너비 캐시 {pk: [widths]}
        # ── 비교 탭 상태 ──────────────────────────────────
        self.cmp_ref={k:[0]*128 for k in PAGE_KEYS}  # 기준 EEPROM
        self.cmp_dut={k:[0]*128 for k in PAGE_KEYS}  # 검사 대상 EEPROM
        self.cmp_ref_name="미로드"
        self._cmp_ref_loaded=False
        self._cmp_dut_loaded=False   # B1: DUT EEPROM 로드 여부
        self._cmp_result_cache=[]  # 내보내기용 전체 결과
        self._build_ui()
        self._apply_theme_full()
        self._refresh_ports()
        self.after(100, self._load_pw_config)
        self.after(150, self._load_app_config)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        # hidapi 자동 설치 완료 시 팝업 안내
        if HID_OK:
            self.after(800, self._check_hid_installed)

    def _on_close(self):
        """종료 시 설정 저장 후 창 닫기"""
        self._save_pw_config()
        self._save_app_config()
        self.destroy()

    # ── UI 빌드 ──────────────────────────────────────────
    def _build_ui(self):
        tb=tk.Frame(self,height=40); tb.pack(fill=tk.X); tb.pack_propagate(False)
        for c,fc in[("●","#ED6A5E"),("●","#F4BF4F"),("●","#61C554")]:
            tk.Label(tb,text=c,fg=fc,font=("",10)).pack(side=tk.LEFT,padx=(8,1),pady=9)
        tk.Label(tb,text=f"  {APP_NAME}",font=("",11,"bold")).pack(side=tk.LEFT,padx=4)
        self.ver_label=tk.Label(tb,text=f"v{APP_VERSION}",font=("",9,"bold"),
                                 fg="#185FA5",cursor="hand2")
        self.ver_label.pack(side=tk.LEFT,padx=2)
        self.ver_label.bind("<Button-1>",lambda e:self._show_ver_dialog())

        self.tab_frame=tk.Frame(self,height=34)
        self.tab_frame.pack(fill=tk.X); self.tab_frame.pack_propagate(False)
        self._tab_btns={}
        for name,label in[("connect","연결"),("viewer","EEPROM 뷰어"),
                           ("compare","비교"),
                           ("log","로그"),("about","정보")]:
            btn=tk.Button(self.tab_frame,text=label,relief=tk.FLAT,bd=0,
                           padx=14,pady=6,cursor="hand2",
                           command=lambda n=name:self._switch_tab(n))
            btn.pack(side=tk.LEFT); self._tab_btns[name]=btn

        self.page_container=tk.Frame(self)
        self.page_container.pack(fill=tk.BOTH,expand=True)
        self._pages={}
        for name in["connect","viewer","compare","log","about"]:
            f=tk.Frame(self.page_container); self._pages[name]=f
            {"connect":self._build_connect,"viewer":self._build_viewer,
             "compare":self._build_compare,
             "log":self._build_log,
             "about":self._build_about}[name](f)
        self._switch_tab("connect")

    # ════════════════════════════════════════════════════
    # 탭 빌드 메서드
    # _build_connect : CP2112 장치 선택, I2C 속도, 연결 버튼
    # _build_viewer  : 페이지 선택 버튼, CanvasTable, 편집바
    # _build_write   : 쓰기 페이지 선택, 범위, 진행바, 검증
    # _build_log     : 동작 이력 텍스트
    # _build_about   : 버전 히스토리, 시스템 정보
    # ════════════════════════════════════════════════════
    def _build_connect(self,p):
        pad=dict(padx=10,pady=7)
        g1=tk.LabelFrame(p,text=" CP2112-F03-GM USB-HID I2C 브릿지 ",padx=8,pady=8)
        g1.pack(fill=tk.X,padx=16,pady=(16,8))
        tk.Label(g1,text="장치:").grid(row=0,column=0,sticky=tk.W,**pad)
        self.port_var=tk.StringVar()
        self.port_cb=ttk.Combobox(g1,textvariable=self.port_var,width=36,state="readonly")
        self.port_cb.grid(row=0,column=1,sticky=tk.W,**pad)
        tk.Button(g1,text="새로고침",command=self._refresh_ports,width=8,cursor="hand2"
                  ).grid(row=0,column=2,**pad)
        tk.Label(g1,text="I2C 주소:").grid(row=1,column=0,sticky=tk.W,**pad)
        self.i2c_addr_var=tk.StringVar(value="50")  # CMIS 기본값 7bit
        fr=tk.Frame(g1); fr.grid(row=1,column=1,sticky=tk.W,**pad)
        tk.Label(fr,text="0x").pack(side=tk.LEFT)
        tk.Entry(fr,textvariable=self.i2c_addr_var,width=5,
                 font=("Consolas",10)).pack(side=tk.LEFT)
        tk.Label(fr,text="(7-bit, CMIS: 0x50 고정)").pack(side=tk.LEFT,padx=6)
        tk.Label(g1,text="I2C 속도:").grid(row=2,column=0,sticky=tk.W,**pad)
        self.i2c_speed_var=tk.StringVar(value="400kHz")
        ttk.Combobox(g1,textvariable=self.i2c_speed_var,width=10,state="readonly",
                     values=["100kHz","400kHz","1MHz"]
                     ).grid(row=2,column=1,sticky=tk.W,**pad)
        self.conn_btn=tk.Button(g1,text="연결",width=10,cursor="hand2",command=self._toggle_connect)
        self.conn_btn.grid(row=0,column=3,rowspan=2,padx=8)
        btn_frame=tk.Frame(g1); btn_frame.grid(row=0,column=4,rowspan=3,padx=(0,8))
        tk.Button(btn_frame,text="Ready",width=9,cursor="hand2",
                  command=self._gpio_ready).pack(pady=(0,4))
        tk.Button(btn_frame,text="Reset",width=9,cursor="hand2",
                  command=self._cp2112_reset).pack()
        self.conn_status=tk.Label(g1,text="● 미연결",fg="#f07070",font=("",10,"bold"))
        self.conn_status.grid(row=2,column=3,**pad)
        hid_fg="#2a7a2a" if HID_OK else "#cc3300"
        hid_msg="✓ SLABHIDtoSMBus.dll 로드됨" if HID_OK else "⚠ SLABHIDtoSMBus.dll 없음 (스크립트 폴더에 복사 필요)"
        tk.Label(g1,text=hid_msg,fg=hid_fg,font=("",9)
                 ).grid(row=3,column=0,columnspan=4,sticky=tk.W,padx=10,pady=(0,4))
        g2=tk.LabelFrame(p,text=" 연결 정보 ",padx=8,pady=8)
        g2.pack(fill=tk.X,padx=16,pady=8)
        self.conn_info=tk.Text(g2,height=5,state=tk.DISABLED,relief=tk.FLAT)
        self.conn_info.pack(fill=tk.X)
        # ── Password Unlock ──────────────────────────
        gp=tk.LabelFrame(p,text=" Password Unlock ",padx=8,pady=8)
        gp.pack(fill=tk.X,padx=16,pady=(0,8))
        padp=dict(padx=8,pady=4)
        # PW Address
        tk.Label(gp,text="PW Address:",width=12,anchor=tk.W).grid(row=0,column=0,sticky=tk.W,**padp)
        self.pw_addr_vars=[tk.StringVar(value=v) for v in ["7A","7B","7C","7D"]]
        pw_addr_fr=tk.Frame(gp); pw_addr_fr.grid(row=0,column=1,sticky=tk.W,**padp)
        for i,var in enumerate(self.pw_addr_vars):
            tk.Entry(pw_addr_fr,textvariable=var,width=3,font=("Consolas",10),
                     justify="center").pack(side=tk.LEFT,padx=2)
        tk.Label(gp,text="(hex, 4바이트)").grid(row=0,column=2,sticky=tk.W,padx=4)
        # PW Value
        tk.Label(gp,text="PW Value:",width=12,anchor=tk.W).grid(row=1,column=0,sticky=tk.W,**padp)
        self.pw_val_vars=[tk.StringVar(value=v) for v in ["00","00","00","00"]]
        pw_val_fr=tk.Frame(gp); pw_val_fr.grid(row=1,column=1,sticky=tk.W,**padp)
        for i,var in enumerate(self.pw_val_vars):
            tk.Entry(pw_val_fr,textvariable=var,width=3,font=("Consolas",10),
                     justify="center").pack(side=tk.LEFT,padx=2)
        tk.Label(gp,text="(hex, 4바이트)").grid(row=1,column=2,sticky=tk.W,padx=4)
        # Unlock 버튼 + 잠금 상태
        btn_pw_fr=tk.Frame(gp); btn_pw_fr.grid(row=0,column=3,rowspan=2,padx=12)
        tk.Button(btn_pw_fr,text="Unlock",width=10,cursor="hand2",
                  command=self._pw_unlock).pack(pady=(0,4))
        tk.Button(btn_pw_fr,text="Lock",width=10,cursor="hand2",
                  command=self._pw_lock).pack()
        self.pw_status=tk.Label(gp,text="🔒 상태 미확인",font=("",10),fg="#888888")
        self.pw_status.grid(row=2,column=0,columnspan=4,sticky=tk.W,padx=8,pady=(0,4))

        g3=tk.LabelFrame(p,text=" CP2112 사용 안내 ",padx=8,pady=8)
        g3.pack(fill=tk.BOTH,expand=True,padx=16,pady=8)
        st=scrolledtext.ScrolledText(g3,height=12,relief=tk.FLAT)
        lines=[
            "[ CP2112-F03-GM USB-HID I2C Bridge ]",
            "",
            "연결 방법:",
            "  1. CP2112 보드를 USB로 PC에 연결",
            "  2. EEPROM SDA/SCL을 CP2112에 연결",
            "  3. I2C 주소 설정 (CMIS Lower Memory: 0x50)",
            "  4. pip install hid",
            "  5. 새로고침 클릭 후 장치 선택 -> 연결",
            "",
            "I2C 주소 (7-bit):",
            "  CMIS Lower Memory (A0h) = 0x50",
            "  CMIS A2h               = 0x51",
            "",
            "드라이버:",
            "  Windows: Silicon Labs CP2112 HID USB-to-SMBus",
            "  Linux  : hid-cp2112 커널 모듈 (내장)",
            "  macOS  : 추가 드라이버 불필요",
            "",
            "설치 확인:",
            "  python -c \"import hid; print(hid.enumerate(0x10C4,0xEA90))\"",
        ]
        st.insert(tk.END,"\n".join(lines))
        st.config(state=tk.DISABLED); st.pack(fill=tk.BOTH,expand=True)


    # ── 뷰어 탭 ──────────────────────────────────────────
    def _build_viewer(self,p):
        top=tk.Frame(p); top.pack(fill=tk.X,padx=12,pady=(10,2))
        for txt,cmd,w in[("파일 열기",self._open_file,9),
                          ("파일 저장",self._save_file,9),
                          ("EEPROM 읽기",self._read_eeprom,11),
                          ("EEPROM 쓰기",self._start_write,11),
                          ("Clear",self._clear_eeprom,7)]:
            tk.Button(top,text=txt,width=w,cursor="hand2",command=cmd
                      ).pack(side=tk.LEFT,padx=(0,5))
        self.file_label=tk.Label(top,text="파일 없음",anchor=tk.W)
        self.file_label.pack(side=tk.LEFT,padx=8)
        self.dirty_label=tk.Label(top,text="")
        self.dirty_label.pack(side=tk.LEFT)
        # 페이지 선택 체크박스
        sel=tk.Frame(p); sel.pack(fill=tk.X,padx=12,pady=(2,0))
        tk.Label(sel,text="대상:",font=("",8)).pack(side=tk.LEFT,padx=(0,4))
        self.page_sel_vars={pk:tk.BooleanVar(value=True) for pk in PAGE_KEYS}
        for pk,lbl in zip(PAGE_KEYS,PAGE_LABELS):
            tk.Checkbutton(sel,text=lbl,variable=self.page_sel_vars[pk],
                           font=("",8),cursor="hand2").pack(side=tk.LEFT,padx=2)
        tk.Button(sel,text="전체",font=("",8),cursor="hand2",width=4,
                  command=lambda:self._set_all_pages(True)
                  ).pack(side=tk.LEFT,padx=(8,2))
        tk.Button(sel,text="해제",font=("",8),cursor="hand2",width=4,
                  command=lambda:self._set_all_pages(False)
                  ).pack(side=tk.LEFT,padx=2)

        pg=tk.Frame(p); pg.pack(fill=tk.X,padx=12,pady=(0,4))
        tk.Label(pg,text="페이지:").pack(side=tk.LEFT,padx=(0,6))
        self._page_btns={}
        for pk,lbl in zip(PAGE_KEYS,PAGE_LABELS):
            btn=tk.Button(pg,text=lbl,width=8,cursor="hand2",
                           command=lambda k=pk:self._show_page(k))
            btn.pack(side=tk.LEFT,padx=2); self._page_btns[pk]=btn

        # CanvasTable
        self.mct=CanvasTable(p, COL_CFG, self.theme)
        self.mct.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        self.mct.bind_dblclick(self._on_row_dblclick)
        self.mct.bind_col_resize(self._on_col_resize)
        self.mct.bind_edit(self._on_inline_edit)
        self.mct.bind_decoded_edit(self._on_decoded_edit)
        self.mct.bind_decode_preview(self._decoded_preview_cb)

        # 편집 안내 라벨만 유지
        self.edit_var=tk.StringVar()  # 내부 호환용
        self.edit_entry=tk.Entry(p,width=1)  # 숨김 (호환용)
        self._editing_row=None
        edit_info=tk.Label(p,text="Value[hex] 더블클릭: 직접 편집",
                           anchor=tk.W,font=("",8))
        edit_info.pack(fill=tk.X,padx=14,pady=(0,2))
        # 쓰기 진행 상태
        self.progress_var=tk.DoubleVar()
        self.write_progress=ttk.Progressbar(p,variable=self.progress_var,maximum=100)
        self.write_progress.pack(fill=tk.X,padx=14,pady=(0,2))
        self.progress_label=tk.Label(p,text="",anchor=tk.W,font=("",8))
        self.progress_label.pack(fill=tk.X,padx=14,pady=(0,4))

    # ── 비교 탭 ──────────────────────────────────────────
    def _build_compare(self, p):
        """
        비교(검사) 탭
        워크플로우:
          1) REF 파일 로드  → cmp_ref_name 표시
          2) DUT EEPROM 읽기 → 읽기 완료 즉시 자동 비교
          3) 결과: PASS/FAIL 배지 + Treeview (FAIL·SKIP 행만)
          4) CSV/TXT 내보내기
        """
        # ── 상단 2패널: REF(파일만) / DUT(읽기만) ────────
        top = tk.Frame(p); top.pack(fill=tk.X, padx=12, pady=(10,4))

        # REF 패널
        gref = tk.LabelFrame(top, text=" 기준 EEPROM (REF) — 파일 로드 ",
                             padx=8, pady=6)
        gref.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0,6))
        rf = tk.Frame(gref); rf.pack(fill=tk.X)
        tk.Button(rf, text="📂  파일 열기", width=13, cursor="hand2",
                  font=("",9,"bold"),
                  command=self._cmp_load_ref).pack(side=tk.LEFT, padx=(0,8))
        self.cmp_ref_lbl = tk.Label(rf, text="📄 미로드", anchor=tk.W,
                                     font=("",9))
        self.cmp_ref_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # DUT 패널
        gdut = tk.LabelFrame(top, text=" 검사 대상 (DUT) — EEPROM 읽기 ",
                             padx=8, pady=6)
        gdut.pack(side=tk.LEFT, fill=tk.X, expand=True)
        df = tk.Frame(gdut); df.pack(fill=tk.X)
        tk.Button(df, text="📡  EEPROM 읽기", width=14, cursor="hand2",
                  font=("",9,"bold"),
                  command=self._cmp_read_dut).pack(side=tk.LEFT, padx=(0,8))
        self.cmp_dut_lbl = tk.Label(df, text="📄 미읽음", anchor=tk.W,
                                     font=("",9))
        self.cmp_dut_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # ── 예외 필드 설정 ────────────────────────────────
        mid = tk.Frame(p); mid.pack(fill=tk.X, padx=12, pady=(0,4))

        gexc = tk.LabelFrame(mid,
            text=" 검사 예외 필드 (체크 = 비교 제외) ",
            padx=8, pady=6)
        gexc.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,6))

        # 스크롤 가능 체크박스
        ecv = tk.Canvas(gexc, height=110, highlightthickness=0)
        evsb = ttk.Scrollbar(gexc, orient=tk.VERTICAL, command=ecv.yview)
        ecv.configure(yscrollcommand=evsb.set)
        evsb.pack(side=tk.RIGHT, fill=tk.Y)
        ecv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ei = tk.Frame(ecv); ecv.create_window((0,0), window=ei, anchor="nw")
        ei.bind("<Configure>",
            lambda e: ecv.configure(scrollregion=ecv.bbox("all")))
        self._exc_inner = ei

        self._exc_vars = []
        for pk,s,e,lbl,default in _CMP_DEFAULT_EXC:
            var = tk.BooleanVar(value=default)
            self._exc_vars.append((pk, s, e, var))
            tk.Checkbutton(ei,
                text=f"[{pk.upper()} {s:02X}~{e:02X}h]  {lbl}",
                variable=var, font=("Consolas",9), cursor="hand2",
                anchor=tk.W).pack(fill=tk.X, padx=2, pady=1)

        # 커스텀 예외 추가
        gcust = tk.LabelFrame(mid, text=" 커스텀 예외 추가 ",
                              padx=8, pady=6)
        gcust.pack(side=tk.LEFT, fill=tk.Y)
        r0 = tk.Frame(gcust); r0.pack(fill=tk.X, pady=2)
        tk.Label(r0, text="Page:", width=6, anchor=tk.W).pack(side=tk.LEFT)
        self._exc_page_var = tk.StringVar(value="a0")
        ttk.Combobox(r0, textvariable=self._exc_page_var,
                     values=PAGE_KEYS, state="readonly", width=6
                     ).pack(side=tk.LEFT, padx=2)
        r1 = tk.Frame(gcust); r1.pack(fill=tk.X, pady=2)
        tk.Label(r1, text="Start:", width=6, anchor=tk.W).pack(side=tk.LEFT)
        self._exc_s_var = tk.StringVar(value="00")
        tk.Entry(r1, textvariable=self._exc_s_var, width=4,
                 font=("Consolas",10), justify="center"
                 ).pack(side=tk.LEFT, padx=2)
        tk.Label(r1, text="End:", width=4).pack(side=tk.LEFT)
        self._exc_e_var = tk.StringVar(value="00")
        tk.Entry(r1, textvariable=self._exc_e_var, width=4,
                 font=("Consolas",10), justify="center"
                 ).pack(side=tk.LEFT, padx=2)
        tk.Label(r1, text="(hex)").pack(side=tk.LEFT, padx=2)
        tk.Button(gcust, text="추가", width=8, cursor="hand2",
                  command=self._cmp_add_exc).pack(pady=(4,0))

        # ── 결과 배지 + 요약 ──────────────────────────────
        act = tk.Frame(p); act.pack(fill=tk.X, padx=12, pady=(0,4))
        self.cmp_result_badge = tk.Label(act, text="  ---  ",
            font=("",13,"bold"), relief=tk.RIDGE, padx=10, pady=3)
        self.cmp_result_badge.pack(side=tk.LEFT, padx=(0,12))
        self.cmp_summary = tk.Label(act, text="", anchor=tk.W, font=("",9))
        self.cmp_summary.pack(side=tk.LEFT)
        tk.Button(act, text="CSV 내보내기", width=11, cursor="hand2",
                  command=lambda: self._export_compare("csv")
                  ).pack(side=tk.RIGHT, padx=(4,0))
        tk.Button(act, text="TXT 내보내기", width=11, cursor="hand2",
                  command=lambda: self._export_compare("txt")
                  ).pack(side=tk.RIGHT, padx=4)

        # ── 결과 Treeview ─────────────────────────────────
        tf = tk.Frame(p); tf.pack(fill=tk.BOTH, expand=True, padx=12,
                                   pady=(0,4))
        cols = ("page","addr","field","ref","dut","result")
        self.cmp_tv = ttk.Treeview(tf, columns=cols, show="headings",
                                    selectmode="browse")
        for col,hdr,w,st in [
            ("page",  "Page",     62, False),
            ("addr",  "Addr",     80, False),
            ("field", "Field",   200, True),
            ("ref",   "REF[hex]",120, False),
            ("dut",   "DUT[hex]",120, False),
            ("result","결과",     72, False),
        ]:
            self.cmp_tv.heading(col, text=hdr)
            self.cmp_tv.column(col, width=w, minwidth=50, stretch=st)
        vsb_tv = ttk.Scrollbar(tf, orient=tk.VERTICAL,
                                command=self.cmp_tv.yview)
        hsb_tv = ttk.Scrollbar(tf, orient=tk.HORIZONTAL,
                                command=self.cmp_tv.xview)
        self.cmp_tv.configure(yscrollcommand=vsb_tv.set,
                               xscrollcommand=hsb_tv.set)
        vsb_tv.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_tv.pack(side=tk.BOTTOM, fill=tk.X)
        self.cmp_tv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 행 색상 태그
        self.cmp_tv.tag_configure("fail",
            foreground="#c00000", background="#fde8e0")
        self.cmp_tv.tag_configure("skip",
            foreground="#7a5500", background="#fffbe0")

    # ── REF 파일 로드 ─────────────────────────────────────
    def _cmp_load_ref(self):
        """기준 EEPROM 파일 로드 (파일만, 장치 읽기 없음)"""
        path = filedialog.askopenfilename(
            title="기준(REF) EEPROM 파일 선택",
            filetypes=[("EEPROM 파일","*.txt *.xlsx *.xls"),
                       ("Text","*.txt"),("Excel","*.xlsx *.xls"),("All","*.*")])
        if not path: return
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in (".xlsx", ".xls"):
                self.cmp_ref = parse_xlsx(path)
            else:
                with open(path, "r", encoding="utf-8") as f:
                    self.cmp_ref = parse_txt(f.read())
            self.cmp_ref_name = os.path.basename(path)
            self._cmp_ref_loaded = True
            self._cmp_ref_path = path   # 설정 저장용
            self.cmp_ref_lbl.config(
                text=f"✅ {self.cmp_ref_name}",
                fg=self.theme.get("grn","#2a7a2a"))
            self._log(f"비교 REF 로드: {path}")
        except Exception as e:
            messagebox.showerror("로드 오류", str(e))

    # ── DUT EEPROM 읽기 → 자동 비교 ─────────────────────
    def _cmp_read_dut(self):
        """DUT EEPROM 읽기 (장치만, 파일 없음) → 완료 후 자동 비교"""
        if self._device_index < 0:
            messagebox.showwarning("연결 필요",
                "연결 탭에서 CP2112를 먼저 연결하세요.")
            return
        if not self._cmp_ref_loaded:
            messagebox.showwarning("REF 없음",
                "먼저 기준(REF) 파일을 로드하세요.")
            return
        def task():
            conn = None
            try:
                conn = self._open_conn()
                i2c = i2c_8bit(self.i2c_addr_var.get())
                tmp = {k:[0]*128 for k in PAGE_KEYS}
                self._log("비교 DUT EEPROM 읽기 시작...")
                self.after(0, lambda: self.cmp_dut_lbl.config(
                    text="⏳ 읽기 중...",
                    fg=self.theme.get("amb","#7a4a00")))
                # Lower Memory
                data = conn.read_page(i2c, 0x00, 128)
                for reg,val in enumerate(data): tmp["a0"][reg] = val
                # Upper Pages
                for pk,pn in zip(["p00","p01","p02","p03","p10","p11"],
                                  [0,1,2,3,0x10,0x11]):
                    conn.set_page(pn)
                    data = conn.read_page(i2c, 0x80, 128)
                    for reg,val in enumerate(data): tmp[pk][reg] = val
                ts = time.strftime("%H:%M:%S")
                self.cmp_dut = tmp
                self._cmp_dut_loaded = True   # B1
                self._log(f"비교 DUT 읽기 완료 ({ts}) → 자동 비교 실행")
                self.after(0, lambda: self.cmp_dut_lbl.config(
                    text=f"✅ 장치 읽기 완료 ({ts})",
                    fg=self.theme.get("grn","#2a7a2a")))
                # 읽기 완료 후 자동으로 비교 실행
                self.after(100, self._cmp_run)
            except Exception as e:
                logging.error("비교 DUT 읽기 오류", exc_info=True)
                self._log(f"비교 DUT 읽기 오류: {e}")
                self.after(0, lambda e=e: (
                    messagebox.showerror("읽기 오류", str(e)),
                    self.cmp_dut_lbl.config(
                        text="❌ 읽기 실패",
                        fg=self.theme.get("red","#8a2020"))))
            finally:
                self._close_conn(conn)
        threading.Thread(target=task, daemon=True).start()

    # ── 예외 셋 수집 ──────────────────────────────────────
    def _get_cmp_exceptions(self):
        """체크된 예외 → set{(pk, idx)} 반환"""
        skip = set()
        for pk,s,e,var in self._exc_vars:
            if var.get():
                for i in range(s, e+1):
                    skip.add((pk, i))
        return skip

    def _cmp_add_exc(self):
        """커스텀 예외 범위 추가"""
        try:
            pk = self._exc_page_var.get()
            s  = int(self._exc_s_var.get().strip(), 16)
            e  = int(self._exc_e_var.get().strip(), 16)
            if s > e or not (0<=s<=127) or not (0<=e<=127):
                raise ValueError("범위 오류: 0x00~0x7F, Start ≤ End")
        except Exception as ex:
            messagebox.showerror("입력 오류", str(ex)); return
        lbl = f"Custom [{pk.upper()} {s:02X}~{e:02X}h]"
        var = tk.BooleanVar(value=True)
        self._exc_vars.append((pk, s, e, var))
        tk.Checkbutton(self._exc_inner,
            text=lbl, variable=var,
            font=("Consolas",9), cursor="hand2",
            anchor=tk.W).pack(fill=tk.X, padx=2, pady=1)
        self._log(f"커스텀 예외 추가: {lbl}")

    # ── 필드명 조회 ───────────────────────────────────────
    def _cmp_field_name(self, pk, idx):
        if pk == "a0":  return _CMP_FIELD_A0.get(idx,  f"A0[{idx:02X}h]")
        if pk == "p00": return _CMP_FIELD_P00.get(idx, f"P00[{idx+0x80:02X}h]")
        base = 0x80
        return f"{pk.upper()}[{idx+base:02X}h]"

    # ── 비교 실행 ─────────────────────────────────────────
    def _cmp_run(self):
        """바이트 단위 전체 비교"""
        if not self._cmp_ref_loaded:
            messagebox.showwarning("REF 없음","기준(REF) 파일을 먼저 로드하세요.")
            return
        if not self._cmp_dut_loaded:
            messagebox.showwarning("DUT 없음","DUT EEPROM 읽기를 먼저 실행하세요.")
            return
        skip = self._get_cmp_exceptions()
        self.cmp_tv.delete(*self.cmp_tv.get_children())
        self._cmp_result_cache = []

        n_pass = 0; n_fail = 0; n_skip = 0

        for pk, page_lbl in zip(PAGE_KEYS, PAGE_LABELS):
            base = 0 if pk == "a0" else 0x80
            for i in range(128):
                # ※ P00h idx=0 (0x80h, SFF8024Identifier):
                # A0h[00h]와 물리적으로 같은 레지스터이지만
                # 파일 기반 비교에서는 독립적으로 비교해야 함 → skip 제거

                abs_addr = i + base
                ref_val  = self.cmp_ref[pk][i]
                dut_val  = self.cmp_dut[pk][i]
                field    = self._cmp_field_name(pk, i)

                if (pk, i) in skip:
                    n_skip += 1
                    if ref_val != dut_val:   # 값이 다를 때만 SKIP 행 표시
                        row = (page_lbl, f"{abs_addr:02X}h", field,
                               h2(ref_val), h2(dut_val), "SKIP")
                        self.cmp_tv.insert("", tk.END, values=row,
                                           tags=("skip",))
                        self._cmp_result_cache.append(row)
                    continue

                if ref_val == dut_val:
                    n_pass += 1
                    self._cmp_result_cache.append(
                        (page_lbl, f"{abs_addr:02X}h", field,
                         h2(ref_val), h2(dut_val), "PASS"))
                else:
                    n_fail += 1
                    row = (page_lbl, f"{abs_addr:02X}h", field,
                           h2(ref_val), h2(dut_val), "✗ FAIL")
                    self.cmp_tv.insert("", tk.END, values=row,
                                       tags=("fail",))
                    self._cmp_result_cache.append(row)

        # 배지 + 요약
        verdict  = "PASS" if n_fail == 0 else "FAIL"
        badge_bg = self.theme.get("grn","#2a7a2a") if n_fail == 0 \
                   else self.theme.get("red","#8a2020")
        self.cmp_result_badge.config(
            text=f"  {verdict}  ", fg="white", bg=badge_bg)
        self.cmp_summary.config(
            text=(f"PASS: {n_pass}  |  ✗ FAIL: {n_fail}"
                  f"  |  SKIP: {n_skip}"
                  f"  |  합계: {n_pass+n_fail+n_skip}"))
        self._log(
            f"비교 완료 [{self.cmp_ref_name}] "
            f"→ PASS={n_pass}  FAIL={n_fail}  SKIP={n_skip}  [{verdict}]")

        # 탭을 비교 탭으로 전환
        self._switch_tab("compare")

        if n_fail == 0:
            messagebox.showinfo("비교 결과",
                f"모든 비교 바이트 일치\n\n"
                f"PASS: {n_pass}  |  SKIP: {n_skip}\n\n✅  PASS")
        else:
            messagebox.showwarning("비교 결과",
                f"{n_fail}개 바이트 불일치\n\n"
                f"PASS: {n_pass}  |  FAIL: {n_fail}  |  SKIP: {n_skip}\n\n"
                f"❌  FAIL  — 아래 테이블에서 상세 확인")

    # ── 비교 결과 내보내기 ────────────────────────────────
    def _export_compare(self, fmt):
        if not self._cmp_result_cache:
            messagebox.showwarning("결과 없음","먼저 DUT를 읽어 비교를 실행하세요.")
            return
        ext = ".csv" if fmt == "csv" else ".txt"
        path = filedialog.asksaveasfilename(
            defaultextension=ext,
            initialfile=f"compare_{time.strftime('%Y%m%d_%H%M%S')}{ext}",
            filetypes=[(fmt.upper(), f"*{ext}"), ("All","*.*")])
        if not path: return
        sep = "," if fmt == "csv" else "\t"
        header = sep.join(["Page","Addr","Field",
                            f"REF ({self.cmp_ref_name})",
                            "DUT (EEPROM읽기)", "Result"])
        lines = [header]
        for row in self._cmp_result_cache:
            if fmt == "csv":
                vals = [f'"{v}"' if ("," in str(v) or sep in str(v))
                        else str(v) for v in row]
            else:
                vals = [str(v) for v in row]
            lines.append(sep.join(vals))
        with open(path, "w", encoding="utf-8-sig") as f:
            f.write("\n".join(lines))
        self._log(f"비교 결과 내보내기: {path}")
        messagebox.showinfo("저장 완료", f"저장됨:\n{path}")

    # ── 로그 탭 ──────────────────────────────────────────
    def _build_log(self,p):
        bar=tk.Frame(p); bar.pack(fill=tk.X,padx=12,pady=8)
        tk.Button(bar,text="로그 지우기",cursor="hand2",command=self._clear_log
                  ).pack(side=tk.LEFT)
        tk.Button(bar,text="로그 저장",cursor="hand2",command=self._save_log
                  ).pack(side=tk.LEFT,padx=6)
        self.log_text=scrolledtext.ScrolledText(p,relief=tk.FLAT,wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,12))
        self.log_text.config(state=tk.DISABLED)

    # ── 정보 탭 ──────────────────────────────────────────
    def _build_about(self,p):
        hf=tk.Frame(p); hf.pack(fill=tk.X,padx=24,pady=(24,8))
        tk.Label(hf,text=APP_NAME,font=("",16,"bold")).pack(anchor=tk.W)
        tk.Label(hf,text=f"Version {APP_VERSION}  ({APP_DATE})",font=("",10)
                 ).pack(anchor=tk.W,pady=2)
        tk.Label(hf,text=f"{SPEC_CMIS} 기반 EEPROM 관리 도구  |  {SPEC_SFF8024}  |  CP2112 USB-HID",
                 font=("",9)).pack(anchor=tk.W,pady=4)
        lf=tk.Frame(p); lf.pack(anchor=tk.W,padx=24,pady=(0,4))
        for typ,badge in TYPE_BADGE.items():
            tk.Label(lf,text=f"{badge} = {typ}",font=("",9),
                     fg=TYPE_COLOR[typ]).pack(side=tk.LEFT,padx=(0,16))
        ttk.Separator(p,orient=tk.HORIZONTAL).pack(fill=tk.X,padx=24,pady=6)
        tk.Label(p,text="변경 이력",font=("",11,"bold")).pack(anchor=tk.W,padx=24,pady=(0,4))
        vf=tk.Frame(p); vf.pack(fill=tk.BOTH,expand=True,padx=24,pady=(0,8))
        vt=ttk.Treeview(vf,columns=("typ","ver","date","note"),show="headings",
                          height=8,selectmode="none")
        for c,h,w in zip(("typ","ver","date","note"),["","버전","날짜","변경 내용"],
                          [36,80,100,520]):
            vt.heading(c,text=h); vt.column(c,width=w,minwidth=20,stretch=(c=="note"))
        for ver,date,typ,note in CHANGELOG:
            tag="latest" if ver==APP_VERSION else typ
            vt.insert("",tk.END,values=(TYPE_BADGE.get(typ,""),ver,date,note),tags=(tag,))
        vsb2=ttk.Scrollbar(vf,orient=tk.VERTICAL,command=vt.yview)
        vt.configure(yscrollcommand=vsb2.set)
        vt.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); vsb2.pack(side=tk.RIGHT,fill=tk.Y)
        self._about_tree=vt
        ttk.Separator(p,orient=tk.HORIZONTAL).pack(fill=tk.X,padx=24,pady=4)
        sf=tk.Frame(p); sf.pack(fill=tk.X,padx=24,pady=(4,16))
        pys="로드됨" if HID_OK else "없음 (스크립트 폴더에 DLL 복사 필요)"
        for lbl,val in[("Python",sys.version.split()[0]),
                        ("SLABHIDtoSMBus.dll",pys),
                        ("플랫폼",sys.platform),
                        ("참조 CMIS",SPEC_CMIS),
                        ("참조 SFF-8024",SPEC_SFF8024)]:
            r=tk.Frame(sf); r.pack(fill=tk.X,pady=1)
            tk.Label(r,text=lbl+":",width=10,anchor=tk.W,font=("",9)).pack(side=tk.LEFT)
            tk.Label(r,text=val,anchor=tk.W,font=("",9)).pack(side=tk.LEFT)

    # ── 버전 팝업 ────────────────────────────────────────
    def _show_ver_dialog(self):
        d=tk.Toplevel(self); d.title("버전 정보")
        d.geometry("540x340"); d.resizable(False,False); d.grab_set()
        t=self.theme; d.config(bg=t["bg1"])
        tk.Label(d,text=APP_NAME,font=("",14,"bold"),bg=t["bg1"],fg=t["t1"]
                 ).pack(pady=(18,3))
        tk.Label(d,text=f"Version {APP_VERSION}  ·  {APP_DATE}",font=("",10),
                 bg=t["bg1"],fg=t["t2"]).pack()
        lf2=tk.Frame(d,bg=t["bg1"]); lf2.pack(pady=(4,2))
        for typ,badge in TYPE_BADGE.items():
            tk.Label(lf2,text=f"{badge}={typ}",font=("",9),
                     bg=t["bg1"],fg=TYPE_COLOR[typ]).pack(side=tk.LEFT,padx=8)
        ttk.Separator(d,orient=tk.HORIZONTAL).pack(fill=tk.X,padx=20,pady=8)
        fr=tk.Frame(d,bg=t["bg1"]); fr.pack(fill=tk.BOTH,expand=True,padx=20,pady=(0,8))
        vt=ttk.Treeview(fr,columns=("typ","ver","date","note"),show="headings",
                          height=7,selectmode="none")
        for c,h,w in zip(("typ","ver","date","note"),["","버전","날짜","내용"],
                          [36,72,92,280]):
            vt.heading(c,text=h); vt.column(c,width=w,stretch=(c=="note"))
        for ver,date,typ,note in CHANGELOG:
            tag="latest" if ver==APP_VERSION else typ
            vt.insert("",tk.END,values=(TYPE_BADGE.get(typ,""),ver,date,note),tags=(tag,))
        vt.tag_configure("latest",foreground=t["acc"],font=("Consolas",10,"bold"))
        vt.tag_configure("major",foreground=t["red"])
        vt.tag_configure("minor",foreground=t["acc"])
        vt.tag_configure("patch",foreground=t["t3"])
        vsb3=ttk.Scrollbar(fr,orient=tk.VERTICAL,command=vt.yview)
        vt.configure(yscrollcommand=vsb3.set)
        vt.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); vsb3.pack(side=tk.RIGHT,fill=tk.Y)
        tk.Button(d,text="닫기",width=10,command=d.destroy,cursor="hand2",
                  bg=t["btn_bg"],fg=t["btn_fg"],relief=tk.FLAT).pack(pady=(0,14))

    # ── 탭 전환 ──────────────────────────────────────────
    def _switch_tab(self,name):
        for n,f in self._pages.items(): f.place_forget()
        self._pages[name].place(relx=0,rely=0,relwidth=1,relheight=1)
        t=self.theme
        for n,btn in self._tab_btns.items():
            active=(n==name)
            btn.config(fg=t["acc"] if active else t["t2"],
                       bg=t["bg2"] if active else t["bg1"],relief=tk.FLAT)

    def _apply_theme_full(self):
        t=self.theme
        self.config(bg=t["bg0"])
        self._walk(self,t)
        self._style_misc(t)
        try: self.mct.apply_theme(t)
        except: pass
        cur=[n for n,f in self._pages.items() if str(f.place_info())!="{}"]
        if cur: self._switch_tab(cur[0])

    def _walk(self,widget,t):
        cls=widget.winfo_class()
        try:
            if cls in("Frame","LabelFrame","Toplevel"):
                widget.config(bg=t["bg1"])
            elif cls=="Label":
                # ver_label은 acc 색상 고정 (버전 표시)
                if widget is getattr(self,"ver_label",None):
                    widget.config(bg=t["bg1"],fg=t["acc"])
                else:
                    widget.config(bg=t["bg1"],fg=t["t1"])
            elif cls=="Button":
                widget.config(bg=t["btn_bg"],fg=t["btn_fg"],
                               activebackground=t["bg0"],relief=tk.FLAT)
            elif cls=="Entry":
                widget.config(bg=t["entry_bg"],fg=t["entry_fg"],
                               insertbackground=t["t1"],relief=tk.FLAT,
                               highlightthickness=1,highlightcolor=t["acc"],
                               highlightbackground=t["bd"])
            elif cls=="Text":
                widget.config(bg=t["bg2"],fg=t["t1"],
                               insertbackground=t["t1"],relief=tk.FLAT)
            elif cls=="Checkbutton":
                widget.config(bg=t["bg1"],fg=t["t1"],
                               activebackground=t["bg1"],selectcolor=t["bg3"])
        except: pass
        for child in widget.winfo_children():
            self._walk(child,t)

    def _style_misc(self,t):
        s=ttk.Style(self); s.theme_use("default")
        s.configure("TCombobox",fieldbackground=t["entry_bg"],background=t["bg2"],
                    foreground=t["t1"],selectbackground=t["sel_bg"])
        s.configure("Vertical.TScrollbar",background=t["bg1"],
                    troughcolor=t["bg0"],arrowcolor=t["t3"])
        s.configure("Horizontal.TScrollbar",background=t["bg1"],
                    troughcolor=t["bg0"],arrowcolor=t["t3"])
        s.configure("TProgressbar",background=t["acc"],troughcolor=t["bg2"])
        # 정보탭 버전 트리
        s.configure("Treeview",background=t["bg2"],foreground=t["t1"],
                    fieldbackground=t["bg2"],rowheight=22,font=("Consolas",10))
        s.configure("Treeview.Heading",background=t["col_head"],foreground=t["acc"],
                    relief="groove",borderwidth=2,font=("",10,"bold"))
        s.map("Treeview",background=[("selected",t["sel_bg"])],
              foreground=[("selected",t["sel_fg"])])
        try:
            self._about_tree.tag_configure("latest",foreground=t["acc"],
                                            font=("Consolas",10,"bold"))
            self._about_tree.tag_configure("major",foreground=t["red"])
            self._about_tree.tag_configure("minor",foreground=t["acc"])
            self._about_tree.tag_configure("patch",foreground=t["t3"])
        except: pass


    def _check_hid_installed(self):
        """자동 설치 후 안내 팝업 (최초 1회)"""
        # 연결 탭의 hid 상태 레이블 갱신
        try:
            self._refresh_ports()
        except Exception:
            pass

    def _refresh_ports(self):
        if not HID_OK:
            self.port_cb["values"]=["SLABHIDtoSMBus.dll 없음"]; return
        devs=CP2112I2C.list_devices()
        if devs:
            items=[f"[{i}] {d.get('product_string','CP2112')}"
                   for i,d in enumerate(devs)]
            self.port_cb["values"]=items; self.port_var.set(items[0])
        else:
            self.port_cb["values"]=["CP2112 장치 없음 (USB 연결 확인)"]


    def _toggle_connect(self):
        # 연결 해제
        if self.conn_btn.cget("text") == "연결 해제":
            if self.conn:
                self.conn.close(); self.conn=None
            self._device_index = -1
            self.conn_status.config(text="● 미연결",fg=self.theme["red"])
            self.conn_btn.config(text="연결")
            self._log("연결 해제"); return
        # DLL 없으면 연결 불가
        if not HID_OK:
            messagebox.showerror("DLL 없음",
                "SLABHIDtoSMBus.dll을 스크립트 폴더에 복사하세요.")
            return
        # 장치 인덱스 파싱 후 저장 (포트는 실제 사용 시에만 열기)
        sel=self.port_var.get()
        try: idx=int(sel.split("]")[0].replace("[","").strip())
        except: idx=0
        # 연결 테스트 (정상이면 즉시 닫고 인덱스만 기억)
        self._log(f"연결 시도: 장치[{idx}] ({sel})")
        try:
            test=CP2112I2C(device_index=idx)
            test.close()
            self._device_index=idx
            self.conn=None
            self.conn_status.config(text="● 대기 중 (포트 해제됨)",fg=self.theme["grn"])
            self.conn_btn.config(text="연결 해제")
            self._log(f"✓ CP2112 장치[{idx}] 연결 확인됨")
            self._update_conn_info()
        except Exception as e:
            logging.error("CP2112 연결 실패", exc_info=True)
            self._log(f"연결 실패 (장치[{idx}]): {e}")
            messagebox.showerror("연결 실패",
                str(e)+"\n\n"
                "※ 다른 프로그램(HidSmbus Example 등)이\n"
                "   CP2112 포트를 점유 중일 수 있습니다.\n"
                "   다른 프로그램을 종료 후 다시 시도하세요.")


    def _open_conn(self):
        """작업 전 포트 열기."""
        if self._device_index < 0:
            raise IOError("장치가 선택되지 않았습니다. 먼저 연결하세요.")
        return CP2112I2C(device_index=self._device_index)

    # ── Password Unlock ──────────────────────────────
    def _set_all_pages(self, val):
        for v in self.page_sel_vars.values(): v.set(val)

    def _pw_config_path(self):
        import os
        return os.path.join(
            os.path.dirname(os.path.abspath(__file__)),".eeprom_pw.json")

    def _app_config_path(self):
        import os
        return os.path.join(
            os.path.dirname(os.path.abspath(__file__)),".cmis_config.json")

    def _save_app_config(self):
        """GUI 설정 저장 — 종료 시 자동 호출"""
        import json, os
        try:
            geo=self.geometry()   # "WxH+X+Y"
            cfg={
                "geometry":       geo,
                "device_index":   self._device_index,
                "i2c_addr":       self.i2c_addr_var.get(),
                "page_sel":       {pk:v.get() for pk,v in self.page_sel_vars.items()},
                "ref_path":       getattr(self,"_cmp_ref_path",""),
                "exc_states":     [(pk,s,e,var.get())
                                   for pk,s,e,var in self._exc_vars
                                   if not hasattr(var,"_default")],
                "col_widths":     self._col_w_cache,
            }
            with open(self._app_config_path(),"w",encoding="utf-8") as f:
                json.dump(cfg,f,indent=2)
        except Exception as e:
            self._log(f"설정 저장 실패: {e}")

    def _load_app_config(self):
        """GUI 설정 복원 — 시작 시 자동 호출"""
        import json, os
        path=self._app_config_path()
        if not os.path.exists(path): return
        try:
            with open(path,"r",encoding="utf-8") as f: cfg=json.load(f)
            # 창 크기/위치
            if "geometry" in cfg:
                try: self.geometry(cfg["geometry"])
                except: pass
            # CP2112 device index
            if "device_index" in cfg and cfg["device_index"]>=0:
                self._device_index=cfg["device_index"]
                # 콤보박스 선택도 동기화
                try:
                    self.device_combo.current(cfg["device_index"])
                except: pass
            # I2C 주소
            if "i2c_addr" in cfg:
                self.i2c_addr_var.set(cfg["i2c_addr"])
            # 페이지 선택 체크박스
            for pk,val in cfg.get("page_sel",{}).items():
                if pk in self.page_sel_vars:
                    self.page_sel_vars[pk].set(val)
            # 비교 예외 체크박스 상태 복원
            exc_map={(pk,s,e):var for pk,s,e,var in self._exc_vars}
            for pk,s,e,val in cfg.get("exc_states",[]):
                if (pk,s,e) in exc_map:
                    exc_map[(pk,s,e)].set(val)
            # 페이지별 열 너비
            for pk,widths in cfg.get("col_widths",{}).items():
                if isinstance(widths,list): self._col_w_cache[pk]=widths
            self._log("설정 복원 완료")
        except Exception as e:
            self._log(f"설정 복원 실패: {e}")

    def _save_pw_config(self):
        import json
        try:
            cfg={"pw_addr":[v.get() for v in self.pw_addr_vars],
                 "pw_val" :[v.get() for v in self.pw_val_vars]}
            with open(self._pw_config_path(),"w") as f:
                json.dump(cfg,f)
        except Exception as e:
            self._log(f"PW 설정 저장 실패: {e}")

    def _load_pw_config(self):
        import json, os
        path=self._pw_config_path()
        if not os.path.exists(path): return
        try:
            with open(path,"r") as f: cfg=json.load(f)
            for var,val in zip(self.pw_addr_vars,cfg.get("pw_addr",[])):
                var.set(val)
            for var,val in zip(self.pw_val_vars,cfg.get("pw_val",[])):
                var.set(val)
            self._log("PW 설정 복원 완료")
        except Exception as e:
            self._log(f"PW 설정 복원 실패: {e}")

    # CMIS 5.2 기준: Password는 I2C로 4바이트 직접 쓰기

    # PW_ADDRESS: 쓸 레지스터 주소 (기본 7Ah~7Dh)
    # PW_VALUE  : 4바이트 패스워드 값
    def _parse_pw_bytes(self, vars_list):
        """StringVar 리스트에서 hex 바이트 파싱"""
        result=[]
        for var in vars_list:
            try:
                val=int(var.get().strip(),16)&0xFF
            except ValueError:
                raise ValueError(f"올바른 hex 값을 입력하세요: {var.get()!r}")
            result.append(val)
        return result

    def _pw_unlock(self):
        """Password Unlock: PW_ADDRESS에 PW_VALUE 4바이트 쓰기"""
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        try:
            addrs=self._parse_pw_bytes(self.pw_addr_vars)
            vals =self._parse_pw_bytes(self.pw_val_vars)
        except ValueError as e:
            messagebox.showerror("입력 오류",str(e)); return
        conn=None
        try:
            conn=self._open_conn()
            i2c=i2c_8bit(self.i2c_addr_var.get())
            self._log("Password Unlock 시도...")
            self._log(f"  PW Addr : {" ".join(f"{a:02X}h" for a in addrs)}")
            self._log(f"  PW Value: {" ".join(f"{v:02X}h" for v in vals)}")
            # 4바이트를 각 주소에 순서대로 쓰기
            for addr,val in zip(addrs,vals):
                conn.write_byte(i2c, addr, val)
            self._log("✓ Password 쓰기 완료")
            self._save_pw_config()
            self.after(0,lambda: self.pw_status.config(
                text="🔓 Unlock 완료 (쓰기 성공)",
                fg="#2a7a2a"))
        except Exception as e:
            logging.error("Password Unlock 오류",exc_info=True)
            self._log(f"Unlock 오류: {e}")
            self.after(0,lambda e=e: self.pw_status.config(
                text=f"❌ 오류: {e}",fg="#cc3300"))
        finally:
            self._close_conn(conn)

    def _pw_lock(self):
        """Password Lock: PW_VALUE 전체를 00으로 쓰기"""
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        try:
            addrs=self._parse_pw_bytes(self.pw_addr_vars)
        except ValueError as e:
            messagebox.showerror("입력 오류",str(e)); return
        if not messagebox.askyesno("Lock 확인",
            "PW_ADDRESS에 0x00을 써서 잠금 상태로 전환합니다.\n계속하시겠습니까?"): return
        conn=None
        try:
            conn=self._open_conn()
            i2c=i2c_8bit(self.i2c_addr_var.get())
            self._log("Password Lock 시도...")
            for addr in addrs:
                conn.write_byte(i2c, addr, 0x00)
            self._log("✓ Password Lock 완료")
            self.after(0,lambda: self.pw_status.config(
                text="🔒 Lock 완료",fg="#cc3300"))
        except Exception as e:
            logging.error("Password Lock 오류",exc_info=True)
            self._log(f"Lock 오류: {e}")
            self.after(0,lambda e=e: self.pw_status.config(
                text=f"❌ 오류: {e}",fg="#cc3300"))
        finally:
            self._close_conn(conn)

    # GPIO 설정값 (HidSmbus Example 그림 기준)
    # GPIO 0,3,7 = Input  / Open-Drain
    # GPIO 1,2,4,5,6 = Output / Push-Pull
    _GPIO_DIRECTION = 0x76  # 0b01110110
    _GPIO_MODE      = 0x76  # 0b01110110
    _GPIO_FUNCTION  = 0x00
    _GPIO_CLKDIV    = 0x00
    _GPIO_LATCH     = 0xB1  # 0b10110001 (초기 출력값)

    def _gpio_ready(self):
        """CP2112 GPIO를 Ready 상태로 설정"""
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        conn=None
        try:
            conn=self._open_conn()
            conn.set_gpio_config(
                self._GPIO_DIRECTION,
                self._GPIO_MODE,
                self._GPIO_FUNCTION,
                self._GPIO_CLKDIV)
            conn.write_latch(self._GPIO_LATCH, 0xFF)
            self._log(
                f"GPIO Ready 설정 완료\n"
                f"  Direction : 0x{self._GPIO_DIRECTION:02X} (0b{self._GPIO_DIRECTION:08b})\n"
                f"  Mode      : 0x{self._GPIO_MODE:02X} (0b{self._GPIO_MODE:08b})\n"
                f"  Latch     : 0x{self._GPIO_LATCH:02X} (0b{self._GPIO_LATCH:08b})")
            messagebox.showinfo("GPIO Ready",
                "GPIO 설정 완료\n\n"
                f"Direction : 0x{self._GPIO_DIRECTION:02X}\n"
                f"Mode      : 0x{self._GPIO_MODE:02X}\n"
                f"Latch     : 0x{self._GPIO_LATCH:02X}")
        except Exception as e:
            logging.error("GPIO 설정 오류", exc_info=True)
            messagebox.showerror("GPIO 오류", str(e))
            self._log(f"GPIO 오류: {e}")
        finally:
            self._close_conn(conn)

    def _cp2112_reset(self):
        """CP2112 Reset - USB 재열거 발생"""
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        if not messagebox.askyesno("Reset 확인",
            "CP2112를 Reset합니다.\n"
            "USB 재열거가 발생하며 연결이 해제됩니다.\n\n"
            "계속하시겠습니까?"): return
        conn=None
        try:
            conn=self._open_conn()
            conn.reset()
            self._log("CP2112 Reset 실행 - USB 재열거 중...")
        except Exception as e:
            self._log(f"Reset 실행: {e}")
        finally:
            # 리셋 후 연결 상태 초기화
            try:
                if conn: conn._is_open=False
            except: pass
            self.conn=None
            self._device_index=-1
            self.after(0,lambda:(
                self.conn_status.config(text="● 미연결 (Reset됨)",
                                        fg=self.theme["red"]),
                self.conn_btn.config(text="연결")))
            self._log("연결 해제됨. 새로고침 후 재연결하세요.")
            self.after(1500, self._refresh_ports)

    def _close_conn(self, conn):
        """작업 후 포트 닫기."""
        try:
            if conn: conn.close()
        except: pass
        self.conn=None
        self.after(0, lambda: self.conn_status.config(
            text="● 대기 중 (포트 해제됨)", fg=self.theme["grn"]))

    def _update_conn_info(self):
        self.conn_info.config(state=tk.NORMAL); self.conn_info.delete("1.0",tk.END)
        if True:
            devs=CP2112I2C.list_devices()
            sel=self.port_var.get()
            try: idx=int(sel.split("]")[0].replace("[","").strip())
            except: idx=0
            d=devs[idx] if idx<len(devs) else {}
            txt=(f"장치: {d.get('product_string','CP2112')}\n"
                 f"S/N : {d.get('serial_number','N/A')}\n"
                 f"I2C 주소: 0x{self.i2c_addr_var.get()} (7-bit)\n"
                 f"I2C 속도: {self.i2c_speed_var.get()}\n"
                 "상태: 연결됨")
        self.conn_info.insert(tk.END,txt)
        self.conn_info.config(state=tk.DISABLED)


    # ── 파일 I/O ─────────────────────────────────────────
    def _clear_eeprom(self):
        """EEPROM 뷰어 데이터 초기화"""
        if not messagebox.askyesno("Clear 확인","뷰어의 모든 데이터를 지우겠습니까?"): return
        self.data={k:[0]*128 for k in PAGE_KEYS}
        self.orig_data={k:[0]*128 for k in PAGE_KEYS}
        self.filename=""
        self.file_label.config(text="파일 없음")
        self.dirty_label.config(text="")
        self._show_page(self.current_page)
        self._log("뷰어 데이터 Clear")

    def _open_file(self):
        ft=[("EEPROM 파일","*.txt *.xlsx *.xls"),("Text","*.txt"),
            ("Excel","*.xlsx *.xls"),("All","*.*")]
        path=filedialog.askopenfilename(filetypes=ft)
        if not path: return
        try:
            ext=os.path.splitext(path)[1].lower()
            if ext in(".xlsx",".xls"):
                self.data=parse_xlsx(path)
            else:
                with open(path,"r",encoding="utf-8") as f: self.data=parse_txt(f.read())
        except ImportError as e:
            messagebox.showerror("라이브러리 없음",str(e)); return
        except Exception as e:
            messagebox.showerror("파일 열기 오류",str(e)); return
        # B4 수정: 파일 로드 직후 orig_data를 파일값과 동기화
        # (이전: [0]*128 → 파일값과 달라서 모든 행이 dirty 표시됨)
        self.orig_data=copy.deepcopy(self.data)
        self.filename=path; self.file_label.config(text=os.path.basename(path))
        self._show_page(self.current_page)
        self._log(f"파일 로드: {path}")

    def _save_file(self):
        base=(os.path.splitext(os.path.basename(self.filename))[0]
              if self.filename else "eeprom")
        path=filedialog.asksaveasfilename(
            defaultextension=".xlsx",initialfile=base+"_edited",
            filetypes=[("Excel","*.xlsx"),("Text","*.txt"),("All","*.*")])
        if not path: return
        try:
            ext=os.path.splitext(path)[1].lower()
            if ext==".xlsx":
                build_xlsx(self.data,path)
            else:
                with open(path,"w",encoding="utf-8") as f: f.write(build_txt(self.data))
        except ImportError as e:
            messagebox.showerror("라이브러리 없음",str(e)); return
        except Exception as e:
            messagebox.showerror("파일 저장 오류",str(e)); return
        self._log(f"파일 저장: {path}")

    # ── 뷰어 ─────────────────────────────────────────────
    # ── 뷰어 데이터 갱신 ─────────────────────────────────
    # _show_page   : 페이지 전환 시 CanvasTable 전체 재구성
    #   1) 페이지 버튼 강조 갱신
    #   2) mct.delete_all() → _get_rows() → mct.insert_row()
    #   3) fit_col_widths() → commit()
    # _get_rows    : 페이지 데이터 → 6컬럼 tuple 리스트
    # _decoded_a0  : A0h 바이트 → Decoded 열 문자열
    # _desc_a0     : A0h 바이트 → Description 열 한국어 설명
    # _decoded_p00 : P00h 바이트 → Decoded 열 문자열
    # _desc_p00    : P00h 바이트 → Description 열 한국어 설명
    # _is_dirty    : 원본 대비 변경 여부 (dirty=파란 배경+주황 텍스트)
    def _show_page(self,pk):
        # 현재 페이지 열 너비 저장
        if hasattr(self,'mct') and self.mct._col_w:
            self._col_w_cache[self.current_page]=list(self.mct._col_w)
        self.current_page=pk
        self.mct._popup_pk=pk   # 팝업 Decoded 미리보기용 현재 페이지 동기화
        t=self.theme
        for k,btn in self._page_btns.items():
            btn.config(fg=t["acc"] if k==pk else t["t2"],
                       bg=t["sel_bg"] if k==pk else t["btn_bg"])
        self.mct.delete_all()
        rows=self._get_rows(pk)
        for i,row in enumerate(rows):
            dirty=self._is_dirty(pk,row[1])
            tag="dirty" if dirty else ("even" if i%2==0 else "odd")
            self.mct.insert_row(row, tag)
        if pk in self._col_w_cache:
            for ci,w in enumerate(self._col_w_cache[pk]):
                if ci<len(self.mct._col_w): self.mct._col_w[ci]=w
            self.mct._recalc_stretch()
        else:
            self.mct.fit_col_widths(rows)
        self.mct.commit()
        self._update_dirty_label()

    def _on_col_resize(self,widths):
        self._col_w_cache[self.current_page]=widths

    def _get_rows(self,pk):
        v=self.data[pk]; rows=[]
        if pk=="a0":
            defs=[
                (0,1,"SFF8024Identifier"),(1,1,"CmisRevision"),(2,1,"MemoryModel"),
                (3,1,"ModuleState"),
                (4,1,"FlagsSummaryBank0"),(5,1,"FlagsSummaryBank1"),
                (6,1,"FlagsSummaryBank2"),(7,1,"FlagsSummaryBank3"),
                (8,1,"ModuleFlags [CDB/FW]"),
                (9,1,"ModuleFlags [Vcc/Temp]"),
                (10,1,"ModuleFlags [Tx Lane]"),
                (11,1,"ModuleFlags [Rx Lane]"),
                (12,1,"ModuleFlags [Lane Adv]"),
                (13,1,"ModuleFlags [6]"),
                (14,2,"TempMon"),(16,2,"VccMon"),
                (18,2,"Aux1Mon"),(20,2,"Aux2Mon"),
                (22,2,"Aux3Mon"),(24,2,"CustomMon"),
                (26,1,"ModuleControls"),
                (27,1,"Reserved [0x1B~0x24]"),
                (28,8,"ModuleLevelMasks [0x1C~0x23]"),
                (36,1,"Reserved [0x24]"),
                (37,1,"CdbStatus1"),(38,1,"CdbStatus2"),
                (39,2,"ActiveFWVersion"),
                (41,1,"ModuleFaultCause"),
                (42,1,"Reserved [0x2A]"),
                (43,8,"Reserved [0x2B~0x32]"),
                (51,8,"Reserved [0x33~0x3A]"),
                (59,1,"Reserved [0x3B]"),(60,5,"Reserved [0x3C~0x40]"),
                (65,1,"OutputDisableTx"),(66,1,"OutputDisableRx"),
                (67,1,"OutputSquelchForceTx"),(68,4,"Reserved [0x44~0x47]"),
                (72,8,"TxBiasMon [Lane1~8]"),
                (80,4,"AppSelCodeLane [Lane1~4]"),
                (84,1,"AppSelCodeLane [Lane5~8]"),
                (85,4,"AppDescriptor[0]"),
                (89,4,"AppDescriptor[1]"),
                (93,4,"AppDescriptor[2]"),
                (97,4,"AppDescriptor[3]"),
                (101,4,"AppDescriptor[4]"),
                (105,4,"AppDescriptor[5]"),
                (109,4,"AppDescriptor[6]"),
                (113,4,"AppDescriptor[7]"),
                (117,8,"DPStateHostLane [Lane1~8]"),
                (125,1,"Reserved [0x7D]"),
                (126,1,"BankSelect"),(127,1,"PageSelect"),
            ]
        elif pk=="p00":
            defs=[
                (0,1,"SFF8024IdentifierCopy"),(1,16,"VendorName"),
                (17,3,"VendorOUI"),(20,16,"VendorPN"),(36,2,"VendorRev"),
                (38,16,"VendorSN"),(54,8,"DateCode"),(62,10,"CLEICode"),
                (72,1,"ModulePowerClass"),(73,1,"MaxPower"),
                (74,1,"CableLinkLength"),(75,1,"MediaConnectorType"),
                (76,8,"MaxMediaLaneCount [App0~7]"),
                (84,1,"MediaInterfaceTechnology"),
                (85,4,"AppDescriptor[0]"),
                (89,4,"AppDescriptor[1]"),
                (93,1,"Reserved [0xDD]"),
                (94,1,"PageChecksum"),
                (95,4,"AppDescriptor[2]"),
                (99,4,"AppDescriptor[3]"),
                (103,4,"AppDescriptor[4]"),
                (107,4,"AppDescriptor[5]"),
                (111,4,"AppDescriptor[6]"),
                (115,4,"AppDescriptor[7]"),
                (119,8,"Reserved [0xF7~0xFE]"),
                (127,1,"Reserved [0xFF]"),
            ]
        elif pk=="p01":
            defs=[
                (0,2,"InactiveFWVersion"),(2,2,"HardwareRevision"),
                (4,2,"LengthSMF"),(6,1,"LengthOM5"),(7,1,"LengthOM4"),
                (8,1,"LengthOM3"),(9,1,"LengthOM2"),
                (10,2,"NominalWavelength"),(12,2,"WavelengthTolerance"),
                (14,1,"ImplementedMemoryPages"),(15,2,"DurationAdvertisement"),
                (17,10,"ModuleCharacteristics"),(27,2,"SupportedControls"),
                (29,2,"SupportedFlags"),(31,2,"SupportedMonitors"),
                (33,2,"SupportedSIControls"),(35,4,"SupportedCDBCommands"),
                (39,2,"MaxDurationInit"),(41,2,"MaxDurationResetInit"),
                (43,2,"MaxDurationReset"),(45,2,"MaxDurationLPInit"),
                (47,2,"MaxDurationHPInit"),(49,2,"MaxDurationTxTurnOn"),
                (51,2,"MaxDurationTxTurnOff"),
                (53,2,"MaxDurationRxTurnOn"),(55,2,"MaxDurationRxTurnOff"),
                (57,2,"MaxDurationTxDisToLOSTx"),(59,4,"SupportedCDBCommands2"),
                (63,1,"Reserved [0xBF]"),
                (64,8,"MediaLaneCharacteristics [Lane1~8]"),
                (72,8,"MediaLaneSupportedFlags [Lane1~8]"),
                (80,16,"TxSIControlAdvertisement"),(96,16,"RxSIControlAdvertisement"),
                (112,15,"Reserved [0xF0~0xFE]"),
                (127,1,"PageChecksum"),
            ]
        elif pk=="p03":
            defs=[
                (0,2,"Reserved [0x80~0x81]"),
                (2,2,"Tx Power Lane 1"),(4,2,"Tx Power Lane 2"),
                (6,2,"Tx Power Lane 3"),(8,2,"Tx Power Lane 4"),
                (10,2,"Tx Power Lane 5"),(12,2,"Tx Power Lane 6"),
                (14,2,"Tx Power Lane 7"),(16,2,"Tx Power Lane 8"),
                (18,2,"Rx Power Lane 1"),(20,2,"Rx Power Lane 2"),
                (22,2,"Rx Power Lane 3"),(24,2,"Rx Power Lane 4"),
                (26,2,"Rx Power Lane 5"),(28,2,"Rx Power Lane 6"),
                (30,2,"Rx Power Lane 7"),(32,2,"Rx Power Lane 8"),
                (34,2,"Tx Bias Lane 1"),(36,2,"Tx Bias Lane 2"),
                (38,2,"Tx Bias Lane 3"),(40,2,"Tx Bias Lane 4"),
                (42,2,"Tx Bias Lane 5"),(44,2,"Tx Bias Lane 6"),
                (46,2,"Tx Bias Lane 7"),(48,2,"Tx Bias Lane 8"),
                (50,2,"Media Lane SNR Lane 1"),(52,2,"Media Lane SNR Lane 2"),
                (54,2,"Media Lane SNR Lane 3"),(56,2,"Media Lane SNR Lane 4"),
                (58,2,"Media Lane SNR Lane 5"),(60,2,"Media Lane SNR Lane 6"),
                (62,2,"Media Lane SNR Lane 7"),(64,2,"Media Lane SNR Lane 8"),
                (66,61,"Reserved [0xC2~0xFE]"),
                (127,1,"Page Checksum"),
            ]
        elif pk=="p02":
            defs=[
                (0,2,"TempHighAlarm"),(2,2,"TempLowAlarm"),
                (4,2,"TempHighWarning"),(6,2,"TempLowWarning"),
                (8,2,"VccHighAlarm"),(10,2,"VccLowAlarm"),
                (12,2,"VccHighWarning"),(14,2,"VccLowWarning"),
                (16,2,"Aux1HighAlarm"),(18,2,"Aux1LowAlarm"),
                (20,2,"Aux1HighWarning"),(22,2,"Aux1LowWarning"),
                (24,2,"Aux2HighAlarm"),(26,2,"Aux2LowAlarm"),
                (28,2,"Aux2HighWarning"),(30,2,"Aux2LowWarning"),
                (32,2,"Aux3HighAlarm"),(34,2,"Aux3LowAlarm"),
                (36,2,"Aux3HighWarning"),(38,2,"Aux3LowWarning"),
                (40,2,"CustomMonHighAlarm"),(42,2,"CustomMonLowAlarm"),
                (44,2,"CustomMonHighWarning"),(46,2,"CustomMonLowWarning"),
                (48,2,"OpticalPowerHighAlarmTx"),(50,2,"OpticalPowerLowAlarmTx"),
                (52,2,"OpticalPowerHighWarningTx"),(54,2,"OpticalPowerLowWarningTx"),
                (56,2,"LaserBiasHighAlarmTx"),(58,2,"LaserBiasLowAlarmTx"),
                (60,2,"LaserBiasHighWarningTx"),(62,2,"LaserBiasLowWarningTx"),
                (64,2,"OpticalPowerHighAlarmRx"),(66,2,"OpticalPowerLowAlarmRx"),
                (68,2,"OpticalPowerHighWarningRx"),(70,2,"OpticalPowerLowWarningRx"),
                (72,2,"LaneOpticalPowerHighAlarmTx"),(74,2,"LaneOpticalPowerLowAlarmTx"),
                (76,2,"LaneOpticalPowerHighWarningTx"),(78,2,"LaneOpticalPowerLowWarningTx"),
                (80,2,"LaneLaserBiasHighAlarmTx"),(82,2,"LaneLaserBiasLowAlarmTx"),
                (84,2,"LaneLaserBiasHighWarningTx"),(86,2,"LaneLaserBiasLowWarningTx"),
                (88,2,"LaneOpticalPowerHighAlarmRx"),(90,2,"LaneOpticalPowerLowAlarmRx"),
                (92,2,"LaneOpticalPowerHighWarningRx"),(94,2,"LaneOpticalPowerLowWarningRx"),
                (96,31,"Reserved [0xE0~0xFE]"),
                (127,1,"PageChecksum"),
            ]
        elif pk=="p10":
            defs=[
                (0,1,"DPDeinit"),
                (1,1,"OutputDisableTx"),
                (2,1,"OutputSquelchForceTx"),
                (3,1,"OutputPolarityFlipTx"),
                (4,1,"OutputDisableRx"),
                (5,1,"OutputPolarityFlipRx"),
                (6,2,"Reserved [0x8E~0x8F]"),
                (8,1,"TxInputAmplitudeLane1"),(9,1,"TxInputAmplitudeLane2"),
                (10,1,"TxInputAmplitudeLane3"),(11,1,"TxInputAmplitudeLane4"),
                (12,1,"TxInputAmplitudeLane5"),(13,1,"TxInputAmplitudeLane6"),
                (14,1,"TxInputAmplitudeLane7"),(15,1,"TxInputAmplitudeLane8"),
                (16,1,"TxInputEqPreCursorLane1"),(17,1,"TxInputEqPreCursorLane2"),
                (18,1,"TxInputEqPreCursorLane3"),(19,1,"TxInputEqPreCursorLane4"),
                (20,1,"TxInputEqPreCursorLane5"),(21,1,"TxInputEqPreCursorLane6"),
                (22,1,"TxInputEqPreCursorLane7"),(23,1,"TxInputEqPreCursorLane8"),
                (24,1,"TxInputEqPostCursorLane1"),(25,1,"TxInputEqPostCursorLane2"),
                (26,1,"TxInputEqPostCursorLane3"),(27,1,"TxInputEqPostCursorLane4"),
                (28,1,"TxInputEqPostCursorLane5"),(29,1,"TxInputEqPostCursorLane6"),
                (30,1,"TxInputEqPostCursorLane7"),(31,1,"TxInputEqPostCursorLane8"),
                (32,1,"RxOutputEqLane1"),(33,1,"RxOutputEqLane2"),
                (34,1,"RxOutputEqLane3"),(35,1,"RxOutputEqLane4"),
                (36,1,"RxOutputEqLane5"),(37,1,"RxOutputEqLane6"),
                (38,1,"RxOutputEqLane7"),(39,1,"RxOutputEqLane8"),
                (40,1,"CDREnableRx"),
                (41,1,"CDREnableTx"),
                (42,85,"Reserved [0xAA~0xFE]"),
                (127,1,"PageChecksum"),
            ]
        elif pk=="p11":
            defs=[
                (0,1,"SCS1::AppSelCodeLane1"),(1,1,"SCS1::AppSelCodeLane2"),
                (2,1,"SCS1::AppSelCodeLane3"),(3,1,"SCS1::AppSelCodeLane4"),
                (4,1,"SCS1::AppSelCodeLane5"),(5,1,"SCS1::AppSelCodeLane6"),
                (6,1,"SCS1::AppSelCodeLane7"),(7,1,"SCS1::AppSelCodeLane8"),
                (8,1,"SCS1::DPConfigLane1"),(9,1,"SCS1::DPConfigLane2"),
                (10,1,"SCS1::DPConfigLane3"),(11,1,"SCS1::DPConfigLane4"),
                (12,1,"SCS1::DPConfigLane5"),(13,1,"SCS1::DPConfigLane6"),
                (14,1,"SCS1::DPConfigLane7"),(15,1,"SCS1::DPConfigLane8"),
                (16,1,"SCS1::TxInputAmplitudeLane1"),(17,1,"SCS1::TxInputAmplitudeLane2"),
                (18,1,"SCS1::TxInputAmplitudeLane3"),(19,1,"SCS1::TxInputAmplitudeLane4"),
                (20,1,"SCS1::TxInputAmplitudeLane5"),(21,1,"SCS1::TxInputAmplitudeLane6"),
                (22,1,"SCS1::TxInputAmplitudeLane7"),(23,1,"SCS1::TxInputAmplitudeLane8"),
                (24,2,"Reserved [0x98~0x99]"),
                (26,1,"SCS1::TxInputEqPreCursorLane1"),(27,1,"SCS1::TxInputEqPreCursorLane2"),
                (28,1,"SCS1::TxInputEqPreCursorLane3"),(29,1,"SCS1::TxInputEqPreCursorLane4"),
                (30,1,"SCS1::TxInputEqPreCursorLane5"),(31,1,"SCS1::TxInputEqPreCursorLane6"),
                (32,1,"SCS1::TxInputEqPreCursorLane7"),(33,1,"SCS1::TxInputEqPreCursorLane8"),
                (34,1,"SCS1::TxInputEqPostCursorLane1"),(35,1,"SCS1::TxInputEqPostCursorLane2"),
                (36,1,"SCS1::TxInputEqPostCursorLane3"),(37,1,"SCS1::TxInputEqPostCursorLane4"),
                (38,1,"SCS1::TxInputEqPostCursorLane5"),(39,1,"SCS1::TxInputEqPostCursorLane6"),
                (40,1,"SCS1::TxInputEqPostCursorLane7"),(41,1,"SCS1::TxInputEqPostCursorLane8"),
                (42,1,"SCS1::RxOutputEqLane1"),(43,1,"SCS1::RxOutputEqLane2"),
                (44,1,"SCS1::RxOutputEqLane3"),(45,1,"SCS1::RxOutputEqLane4"),
                (46,1,"SCS1::RxOutputEqLane5"),(47,1,"SCS1::RxOutputEqLane6"),
                (48,1,"SCS1::RxOutputEqLane7"),(49,1,"SCS1::RxOutputEqLane8"),
                (50,9,"Reserved [0xB2~0xBA]"),
                (59,1,"SCS1::CDREnableRxLane1"),(60,1,"SCS1::CDREnableRxLane2"),
                (61,1,"SCS1::CDREnableRxLane3"),(62,1,"SCS1::CDREnableRxLane4"),
                (63,1,"SCS1::CDREnableRxLane5"),(64,1,"SCS1::CDREnableRxLane6"),
                (65,1,"SCS1::CDREnableRxLane7"),(66,1,"SCS1::CDREnableRxLane8"),
                (67,33,"Reserved [0xC3~0xE3]"),
                (100,27,"Reserved [0xE4~0xFE]"),
                (127,1,"PageChecksum"),
            ]
        else:
            base=0x80
            for i in range(128):
                if v[i]:
                    rows.append((str(i+base),f"{i+base:02X}h",h2(v[i]),
                                  h2(v[i]),f"{pk.upper()}:{i+base:02X}h",""))
            return rows
        base=0x80 if pk!="a0" else 0
        for start,ln,field in defs:
            ad=(str(start+base) if ln==1 else f"{start+base}~{start+ln-1+base}")
            ah=(f"{start+base:02X}h" if ln==1 else
                f"{start+base:02X}~{start+ln-1+base:02X}h")
            # Value[hex]: 전체 바이트 표시
            vs=" ".join(h2(v[start+i]) for i in range(ln))
            if pk=="a0":
                decoded=self._decoded_a0(v,start)
                desc=self._desc_a0(v,start)
            elif pk=="p00":
                decoded=self._decoded_p00(v,start)
                desc=self._desc_p00(v,start)
            elif pk=="p01":
                decoded=self._decoded_p01(v,start)
                desc=self._desc_p01(start)
            elif pk=="p02":
                decoded=self._decoded_p02(v,start)
                desc=self._desc_p02(start)
            elif pk=="p03":
                decoded=self._decoded_p03(v,start)
                desc=self._desc_p03(start)
            elif pk=="p10":
                decoded=self._decoded_p10(v,start)
                desc=self._desc_p10(start)
            elif pk=="p11":
                decoded=self._decoded_p11(v,start)
                desc=self._desc_p11(start)
            else:
                decoded=""; desc=""
            rows.append((ad,ah,vs,decoded,field,desc))
        return rows

    def _decoded_a0(self,v,i):
        """Value 열 옆 Decoded 열: 숫자/상태를 간결하게"""
        if i==0:  return SFF8024.get(v[0],f"?({h2(v[0])})")
        if i==1:
            rev = _CMIS_REV_MAP.get(v[1], f"v{(v[1]>>4)&0xF}.{v[1]&0xF}")
            return rev
        if i==2:
            mci=["≤400kHz","≤1MHz","Rsvd","Rsvd"][(v[2]>>2)&3]
            return f"{'Flat' if v[2]>>7 else 'Paged'} MCI={mci}"
        if i==3:
            ms=(v[3]>>1)&7
            st=["Rsvd","LowPwr","PwrUp","Ready","PwrDn","Fault","Rsvd","Rsvd"]
            return f"{st[ms]} Int={'No' if v[3]&1 else 'ASSERT'}"
        if i>=4 and i<=7:
            b=i-4; flags=[]
            if v[i]&8: flags.append(f"B{b}P2C")
            if v[i]&4: flags.append(f"B{b}P14")
            if v[i]&2: flags.append(f"B{b}P12")
            if v[i]&1: flags.append(f"B{b}P11")
            return ",".join(flags) if flags else "—"
        if i==8:
            fl=[]
            if v[8]&0x80: fl.append("CdbCmp2")
            if v[8]&0x40: fl.append("CdbCmp1")
            if v[8]&0x04: fl.append("DPFWErr")
            if v[8]&0x02: fl.append("ModFWErr")
            if v[8]&0x01: fl.append("StateChg")
            return ",".join(fl) if fl else "—"
        if i==9:
            fl=[]
            if v[9]&0x80: fl.append("VccLowW")
            if v[9]&0x40: fl.append("VccHiW")
            if v[9]&0x20: fl.append("VccLowA")
            if v[9]&0x10: fl.append("VccHiA")
            if v[9]&0x08: fl.append("TmpLowW")
            if v[9]&0x04: fl.append("TmpHiW")
            if v[9]&0x02: fl.append("TmpLowA")
            if v[9]&0x01: fl.append("TmpHiA")
            return ",".join(fl) if fl else "N/A"
        if i==14:
            return f"{s16(v[14],v[15])/256:.2f} °C"
        if i==16:
            return f"{u16(v[16],v[17])*0.0001:.4f} V"
        if i==18: return f"{s16(v[18],v[19])} (raw S16)"
        if i==20: return f"{s16(v[20],v[21])} (raw S16)"
        if i==22: return f"{s16(v[22],v[23])} (raw S16)"
        if i==24: return f"{s16(v[24],v[25])} (raw S16)"
        if i==26:
            return (f"LowPwrHW={(v[26]>>6)&1} "
                    f"Squelch={(v[26]>>5)&1} "
                    f"LowPwrSW={(v[26]>>4)&1} "
                    f"Reset={(v[26]>>3)&1}")
        if i==27:
            fc=v[27]&0x1F
            names=["","TxFault","RxLOS","MemError","ThermalShutdown","FWFault"]
            return names[fc] if 0<fc<len(names) else f"Code={fc}"
        if i==37: return f"Busy={(v[37]>>7)&1} Fail={(v[37]>>6)&1} Res={h2(v[37]&0x3F)}"
        if i==38: return f"Busy={(v[38]>>7)&1} Fail={(v[38]>>6)&1} Res={h2(v[38]&0x3F)}"
        if i==39: return f"v{v[39]}.{v[40]}"
        if i==65: return f"TxDis=0b{v[65]:08b}"
        if i==66: return f"RxOutDis=0b{v[66]:08b}"
        if i==67: return f"TxForcedSql=0b{v[67]:08b}"
        if i==72:
            biases=[u16(v[72+j*2],v[73+j*2])*2 for j in range(min(4,56//2))]
            return " ".join(f"L{j+1}:{b}µA" for j,b in enumerate(biases))
        for ap in range(8):
            base_ap=85+ap*4
            if i==base_ap:
                return (f"Host={h2(v[base_ap])} "
                        f"Media={h2(v[base_ap+1])} "
                        f"HLane={(v[base_ap+2]>>4)&0xF} "
                        f"MLane={v[base_ap+2]&0xF} "
                        f"LaneAssign={h2(v[base_ap+3])}")
        if i==117:
            return " ".join(f"L{j+1}:{h2(v[117+j])}" for j in range(8))
        if i==126: return f"Bank={v[126]}"
        if i==127: return f"Page={h2(v[127])}"
        return "—"

    def _desc_a0(self,v,i):
        """Description 열: 필드 의미 설명"""
        if i==0:  return f"Identifier: Module type per SFF-8024 Table 4-1"
        if i==1:  return f"CMIS Revision: [7:4]=Major [3:0]=Minor | Ref: {SPEC_CMIS}"
        if i==2:  return "Memory Model [7]=Paged/Flat, Max MCI Speed [3:2]"
        if i==3:  return "Module State [3:1]=ModuleState, [0]=IntL asserted"
        if i>=4 and i<=7: return f"IntL Flag Summary Bank{i-4}: Interrupt summary per page group"
        if i==8:  return "Module Interrupt Flags 1: CDB Complete / FW Error Flags"
        if i==9:  return "Module Interrupt Flags 2: Vcc/Temp Alarm & Warning Flags"
        if i==10: return "Module Interrupt Flags 3: Tx Lane Fault/Squelch Flags"
        if i==11: return "Module Interrupt Flags 4: Rx Lane LOS/LOL Flags"
        if i==12: return "Module Interrupt Flags 5: Additional Lane Flags"
        if i==13: return "Module Interrupt Flags 6"
        if i==14: return "Module Temperature (S16/256 °C)"
        if i==16: return "Supply Voltage (U16×100µV)"
        if i==18: return "Aux1 Monitor (TEC Current or Custom, S16)"
        if i==20: return "Aux2 Monitor (Laser Temp or TEC Current, S16)"
        if i==22: return "Aux3 Monitor (Laser Temp or Aux Voltage, S16)"
        if i==24: return "Custom Monitor (S16)"
        if i==26: return "Module Global Controls: LowPwrHW[6] Squelch[5] LowPwrSW[4] Reset[3]"
        if i==27: return "Module Fault Cause (CMIS 5.0+): TxFault/RxLOS/MemError/Thermal/FWFault"
        if i==37: return "CDB Status 1: Busy[7] Fail[6]"
        if i==38: return "CDB Status 2: Busy[7] Fail[6]"
        if i==39: return "Active FW Version (Major.Minor)"
        if i==41: return "Lane IntL Flag Summary [Tx]: Tx Lane interrupt summary"
        if i==42: return "Lane IntL Flag Summary [Rx]: Rx Lane interrupt summary"
        if i==43: return "Tx Fault / Forced Squelch Flags per Lane (Lane1~8)"
        if i==51: return "Rx LOS / LOL Flags per Lane (Lane1~8)"
        if i==65: return "Tx Disable Status: bit N-1 = Lane N (1=Disabled)"
        if i==66: return "Rx Output Disable Status: bit N-1 = Lane N"
        if i==67: return "Tx Forced Squelch Status: bit N-1 = Lane N"
        if i==72: return "Tx Bias Monitor [Lane1~8] (U16×2µA per Lane)"
        if i==80: return "Active DataPath Config [Lane1~4]: [7:4]=Host AppSel [3:0]=Media AppSel"
        if i==84: return "Active DataPath Config [Lane5~8]"
        for ap in range(8):
            ai=[85,89,93,97,101,105,109,113][ap]
            if i==ai: return f"Application Descriptor [{ap}]: HostIfID / MediaIfID / LaneCnt / LaneAssign"
        if i==117: return "DataPath State per Lane: [7:4]=AppSel [3:0]=DataPathID (Lane1~8)"
        if i==126: return "Bank Select: Upper memory bank selection"
        if i==127: return "Page Select: Upper memory page selection"
        return ""

    def _decoded_p00(self,v,i):
        """Decoded 열: 값을 사람이 읽을 수 있는 형태로"""
        if i==0:  return SFF8024.get(v[0],f"?({h2(v[0])})")
        if i==1:  return asc_display(v,1,16)          # VendorName
        if i==17:
            oui=f"{h2(v[17])}-{h2(v[18])}-{h2(v[19])}"
            # 알려진 OUI
            known={"CC-03-88":"MangoBoost","00-17-6A":"Cisco","00-90-65":"Finisar",
                   "00-02-C9":"Mellanox","00-00-5A":"3Com"}
            name=known.get(oui,"")
            return f"{oui}" + (f" ({name})" if name else "")  # OUI
        if i==20: return asc_display(v,20,35)         # VendorPN
        if i==36: return asc_display(v,36,37)         # VendorRev
        if i==38:
            raw=asc_display(v,38,53)
            # 실제 의미있는 내용 앞부분 추출 (공백 제거)
            stripped=raw.strip()
            if stripped:
                return stripped
            return "(space padding only)"  # VendorSN
        if i==54:
            dc=asc_display(v,54,61)
            if all(v[54+j]==0 for j in range(8)):
                return "(미설정)"
            return dc
        if i==62: return asc_display(v,62,71)         # CLEI
        if i==72:
            pc=(v[72]>>5)&7
            pw=["≤1.5W","≤2.0W","≤2.5W","≤3.5W","≤4.0W","≤4.5W","≤5.0W",">5W"]
            return f"Class{pc+1} {pw[pc]}"
        if i==73: return f"{v[73]*0.25:.2f} W"
        if i==74: return "N/A(cable)" if v[74]==0 else f"Mult={v[74]>>6} Val={v[74]&0x3F}"
        if i==75: return f"Connector={h2(v[75])}"
        if i==76:  # MediaLanes per AppDescriptor
            return " ".join(f"A{j}:{v[76+j]}" for j in range(8))
        if i==84:
            return {0:"850nm VCSEL",1:"1310nm VCSEL",2:"1550nm VCSEL",
                    3:"1310nm FP",4:"1310nm DFB",5:"1490nm DFB",
                    6:"1310nm EML",7:"1550nm EML"}.get(v[84],h2(v[84]))
        # AppDescriptor[0~7]: 각 4바이트 (85,89,93제외,95,99,103,107,111,115)
        for ap,(ai) in enumerate([85,89,95,99,103,107,111,115]):
            if i==ai:
                return (f"Host={h2(v[ai])} "
                        f"Media={h2(v[ai+1])} "
                        f"HLane={(v[ai+2]>>4)&0xF} "
                        f"MLane={v[ai+2]&0xF} "
                        f"LaneAssign={h2(v[ai+3])}")
        if i==93: return f"Reserved={h2(v[93])}"
        if i==94: return f"0x{h2(v[94])}"
        return "N/A"

    def _desc_p00(self,v,i):
        """Description 열: 필드 의미"""
        if i==0:  return "Module Type Identifier (SFF-8024 Table 4-1 copy)"
        if i==1:  return "Vendor Name: Manufacturer name (ASCII 16 chars, right space padded)"
        if i==17: return "Vendor OUI: Manufacturer IEEE OUI 3 bytes"
        if i==20: return "Vendor Part Number: Part number (ASCII 16 chars)"
        if i==36: return "Vendor Revision: Part number revision (ASCII 2 chars)"
        if i==38: return "Vendor Serial Number: Serial number (ASCII 16 chars)"
        if i==54: return "Date Code: Manufacturing date (YYYYMMDD, ASCII 8 chars)"
        if i==62: return "CLEI Code: Common Language Equipment Identifier (ASCII 10 chars, optional)"
        if i==72: return "Module Power Class [7:5] / Max Power multiplier [4:0]"
        if i==73: return "Max Power: Maximum power consumption (0.25W step)"
        if i==74: return "Cable Assembly Length"
        if i==75: return "Connector Type (SFF-8024 Table 4-3)"
        if i==76: return "Max Media Lane Count per Application Descriptor (App0~7, CMIS 5.2)"
        if i==84: return "Media Interface Technology (SFF-8024 Table 4-8: laser wavelength/type)"
        for ap,(ai) in enumerate([85,89,95,99,103,107,111,115]):
            if i==ai: return f"Application Descriptor [{ap}]: HostIfID / MediaIfID / LaneCnt / LaneAssign"
        if i==93: return "Reserved (0xDD)"
        if i==94: return "Page Checksum: Bytes 0x80~0xDD sum mod 256 (at 0xDE)"
        if i==119: return "Reserved (0xF7~0xFE)"
        if i==127: return "Reserved (0xFF)"
        return ""

    # ── P01h (Advertising) ─────────────────────────────
    def _decoded_p01(self,v,i):
        """P01h Decoded: 광고 필드 해석"""
        if i==0:  # InactiveFW Major.Minor
            return f"v{v[0]}.{v[1]}"
        if i==2:  # HW Major.Minor
            return f"HW v{v[2]}.{v[3]}"
        if i==4:  # LengthSMF
            mult=[0.1,1,10,0][v[4]>>6]
            base=v[4]&0x3F
            km=mult*base
            return f"{km:.1f} km" if km else "N/A"
        if i==6:  return f"{v[6]*2} m" if v[6] else "N/A"  # OM5
        if i==7:  return f"{v[7]*2} m" if v[7] else "N/A"  # OM4
        if i==8:  return f"{v[8]*2} m" if v[8] else "N/A"  # OM3
        if i==9:  return f"{v[9]} m"   if v[9] else "N/A"  # OM2
        if i==10: # NominalWavelength (U16 x 0.05nm)
            wl=u16(v[10],v[11])*0.05
            return f"{wl:.2f} nm" if wl else "N/A"
        if i==12: # WavelengthTolerance (U16 x 0.005nm)
            tol=u16(v[12],v[13])*0.005
            return f"±{tol:.3f} nm" if tol else "N/A"
        if i==14: # SupportedPages 비트 플래그
            bits=[]
            if v[14]&0x80: bits.append("P01h")
            if v[14]&0x40: bits.append("P02h")
            if v[14]&0x20: bits.append("P10h")
            if v[14]&0x10: bits.append("P11h")
            return ",".join(bits) if bits else "—"
        if i==127: return f"0x{h2(v[127])}"
        # 타이밍 특성 (ms, U16)
        _timing={39:"T_init",41:"T_reset",43:"T_resetInit",45:"T_lpInit",
                 47:"T_hpInit",49:"T_txTurnOn",51:"T_txTurnOff",
                 53:"T_rxTurnOn",55:"T_rxTurnOff",57:"T_txDisToLos"}
        if i in _timing:
            return f"{_timing[i]}={u16(v[i],v[i+1])} ms"
        if i==59: return f"CDB_Adv={h2(v[59])} {h2(v[60])} {h2(v[61])} {h2(v[62])}"
        if i==64: return " ".join(f"L{j+1}:{h2(v[64+j])}" for j in range(8))
        if i==72: return " ".join(f"L{j+1}:{h2(v[72+j])}" for j in range(8))
        if i==80: return f"SICtrl[TX] {' '.join(h2(v[80+j]) for j in range(8))}..."
        if i==96: return f"SICtrl[RX] {' '.join(h2(v[96+j]) for j in range(8))}..."
        return "—"

    def _desc_p01(self,i):
        """P01h Description"""
        if i==0:  return "Inactive FW Version: Inactive firmware version (Major.Minor)"
        if i==2:  return "Hardware Version: Hardware revision (Major.Minor)"
        if i==4:  return "Link Length SMF: Max SMF link length ([7:6]=multiplier×{0.1,1,10}, [5:0]=value km)"
        if i==6:  return "Link Length OM5: Max OM5 link length (×2m)"
        if i==7:  return "Link Length OM4: Max OM4 link length (×2m)"
        if i==8:  return "Link Length OM3: Max OM3 link length (×2m)"
        if i==9:  return "Link Length OM2: Max OM2 link length (×1m)"
        if i==10: return "Nominal Wavelength: Nominal center wavelength (U16×0.05nm)"
        if i==12: return "Wavelength Tolerance: Wavelength tolerance ±(U16×0.005nm)"
        if i==14: return "Supported Pages Adv.: Supported upper pages bitmap (P01h/P02h/P10h/P11h etc.)"
        if i==15: return "Duration Advertising: State machine transition time advertising (ms)"
        if i==17: return "Module Characteristics Adv.: CDR/TX/RX feature support advertising (10 bytes)"
        if i==27: return "Supported Controls Adv.: Supported control features bitmap"
        if i==29: return "Supported Flags Adv.: Supported flags bitmap"
        if i==31: return "Supported Monitors Adv.: Supported monitors bitmap"
        if i==33: return "Supported SI Controls Adv.: Supported Signal Integrity Controls bitmap"
        if i==35: return "Supported CDB Commands: Supported CDB features advertising (4 bytes)"
        if i==39: return "t_init_time: Max time from power-on to Ready state (U16 ms)"
        if i==41: return "t_reset_time: Max time from ResetL de-assert to Ready state (U16 ms)"
        if i==43: return "t_resetInit_time: Max time in MgmtInit state (U16 ms)"
        if i==45: return "t_lpInit_time: Max LowPwr to Init transition time (U16 ms)"
        if i==47: return "t_hpInit_time: Max HighPwr Init time (U16 ms)"
        if i==49: return "t_txTurnOn_time: Max Tx turn-on time (U16 ms)"
        if i==51: return "t_txTurnOff_time: Max Tx turn-off time (U16 ms)"
        if i==53: return "t_rxTurnOn_time: Max Rx turn-on time (U16 ms)"
        if i==55: return "t_rxTurnOff_time: Max Rx turn-off time (U16 ms)"
        if i==57: return "t_txDisToLos_time: Max time from TxDisable to LOS assert (U16 ms)"
        if i==59: return "CDB Feature Advertising: Supported CDB feature details (4 bytes)"
        if i==64: return "Media Lane Characteristics per Lane (Lane1~8, 1 byte each)"
        if i==72: return "Media Lane Supported Flags per Lane (Lane1~8, 1 byte each)"
        if i==80: return "Tx SI Control Advertising: Tx SI Controls support range (16 bytes)"
        if i==96: return "Rx SI Control Advertising: Rx SI Controls support range (16 bytes)"
        if i==112: return "Reserved (0xF0~0xFE)"
        if i==127: return "Page Checksum: Bytes 0x82~0xFE sum mod 256"
        return ""

    # ── P02h (Module and Lane Thresholds) ──────────────
    def _decoded_p02(self,v,i):
        """P02h Decoded: 임계값 실제 단위 변환"""
        # Temp: S16 / 256 °C
        if 0<=i<=6 and i%2==0:
            val=s16(v[i],v[i+1])/256.0
            return f"{val:+.2f} °C"
        # Vcc: U16 × 100µV → V
        if 8<=i<=14 and i%2==0:
            val=u16(v[i],v[i+1])*0.0001
            return f"{val:.4f} V"
        # Aux1/2/3/Custom: S16, 단위 모듈마다 다름
        if 16<=i<=46 and i%2==0:
            val=s16(v[i],v[i+1])
            return f"{val} (raw S16)"
        # Tx Power: U16 × 0.1µW → dBm
        if 48<=i<=54 and i%2==0:
            return uw_to_dbm(u16(v[i],v[i+1]))
        # Tx Bias: U16 × 2µA
        if 56<=i<=62 and i%2==0:
            val=u16(v[i],v[i+1])*2
            return f"{val} µA"
        # Rx Power: U16 × 0.1µW → dBm
        if 64<=i<=70 and i%2==0:
            return uw_to_dbm(u16(v[i],v[i+1]))
        # Per-Lane Tx Power (72~79)
        if 72<=i<=78 and i%2==0:
            return uw_to_dbm(u16(v[i],v[i+1]))
        # Per-Lane Tx Bias (80~87)
        if 80<=i<=86 and i%2==0:
            return f"{u16(v[i],v[i+1])*2} µA"
        # Per-Lane Rx Power (88~95)
        if 88<=i<=94 and i%2==0:
            return uw_to_dbm(u16(v[i],v[i+1]))
        if i==127: return f"0x{h2(v[127])}"
        return "—"

    def _desc_p02(self,i):
        """P02h Description — 모듈 레벨 _P02H_DESC 상수 사용"""
        if i in _P02H_DESC:
            name, unit = _P02H_DESC[i]
            return f"{name} ({unit})"
        return ""

    def _decoded_p03(self,v,i):
        """P03h Decoded — 레인별 Tx/Rx 파워·바이어스·SNR 실시간 모니터"""
        if 2<=i<=16 and i%2==0:
            lane=(i-2)//2+1
            return f"L{lane}: {uw_to_dbm(u16(v[i],v[i+1]))}"
        if 18<=i<=32 and i%2==0:
            lane=(i-18)//2+1
            return f"L{lane}: {uw_to_dbm(u16(v[i],v[i+1]))}"
        if 34<=i<=48 and i%2==0:
            lane=(i-34)//2+1
            raw=u16(v[i],v[i+1])
            return f"L{lane}: {raw*2} µA = {raw*2/1000:.3f} mA"
        if 50<=i<=64 and i%2==0:
            lane=(i-50)//2+1
            raw=u16(v[i],v[i+1])
            return f"L{lane}: {raw/256:.2f} dB"
        if i==127: return f"0x{h2(v[127])}"
        return "—"

    def _desc_p03(self,i):
        """P03h Description"""
        if 2<=i<=16 and i%2==0:
            lane=(i-2)//2+1
            return f"Tx Power Lane {lane}: Media Tx optical output power (U16×0.1µW → dBm, RO)"
        if 18<=i<=32 and i%2==0:
            lane=(i-18)//2+1
            return f"Rx Power Lane {lane}: Media Rx optical input power (U16×0.1µW → dBm, RO)"
        if 34<=i<=48 and i%2==0:
            lane=(i-34)//2+1
            return f"Tx Bias Lane {lane}: Media Tx laser bias current (U16×2µA, RO)"
        if 50<=i<=64 and i%2==0:
            lane=(i-50)//2+1
            return f"Media Lane SNR Lane {lane}: Signal-to-Noise Ratio (U8.8 dB, RO)"
        if i==127: return "Page Checksum: Bytes 0x80~0xFE sum mod 256"
        return ""

    # ── P10h (Lane SI Controls) ────────────────────────────
    def _decoded_p10(self,v,i):
        if i in (0,1,2,3,4,5,40,41): return f"0b{v[i]:08b}"
        if 8<=i<=39: return f"{v[i]}"
        if i==127: return f"0x{h2(v[127])}"
        return "—"

    def _desc_p10(self,i):
        _LANE_CTRL={
            0:"DataPath Deinit [Lane8:Lane1]: bit N-1 = Lane N (1=Deinit DataPath)",
            1:"Tx Disable [Lane8:Lane1]: bit N-1 = Lane N (1=Tx disabled)",
            2:"Tx Forced Squelch [Lane8:Lane1]: bit N-1 = Lane N (1=Squelch forced)",
            3:"Tx Polarity Flip [Lane8:Lane1]: bit N-1 = Lane N (1=polarity inverted)",
            4:"Rx Output Disable [Lane8:Lane1]: bit N-1 = Lane N (1=Rx output disabled)",
            5:"Rx Output Polarity Flip [Lane8:Lane1]: bit N-1 = Lane N (1=polarity inverted)",
            40:"Rx CDR Bypass [Lane8:Lane1]: bit N-1 = Lane N (1=CDR Bypass)",
            41:"Tx CDR Bypass [Lane8:Lane1]: bit N-1 = Lane N (1=CDR Bypass)"}
        if i in _LANE_CTRL: return _LANE_CTRL[i]
        if 8<=i<=15:  return f"Tx Output Amplitude Lane {i-7}: 4-bit unsigned (0=min)"
        if 16<=i<=23: return f"Tx Pre-Cursor Lane {i-15}: 5-bit signed (two's complement)"
        if 24<=i<=31: return f"Tx Post-Cursor Lane {i-23}: 5-bit signed (two's complement)"
        if 32<=i<=39: return f"Rx Input Equalization Lane {i-31}: 4-bit unsigned"
        if i==127: return "Page Checksum: Bytes 0x80~0xFE sum mod 256"
        return ""

    # ── P11h (SCS1 — Staged Control Set 1) ────────────────
    def _decoded_p11(self,v,i):
        if 0<=i<=7:
            host=(v[i]>>4)&0xF; media=v[i]&0xF
            if v[i]==0: return "—"
            return f"Host={host} Media={media}"
        if 8<=i<=15: return f"0b{v[i]:08b}"
        if 16<=i<=49 and i not in (24,25): return f"{v[i]}"
        if 59<=i<=66: return f"0b{v[i]:08b}"
        if i==127: return f"0x{h2(v[127])}"
        return "—"

    def _desc_p11(self,i):
        if 0<=i<=7:   return f"SCS1 AppSel Code Lane {i+1}: Host AppSel[7:4] / Media AppSel[3:0]"
        if 8<=i<=15:  return f"SCS1 Data Path Configuration Code Lane {i-7}: RxCDR Bypass[7] TxCDR Bypass[4] DataPath ID[3:0]"
        if 16<=i<=23: return f"SCS1 Tx Output Amplitude Lane {i-15}: 4-bit unsigned (0=min)"
        if 26<=i<=33: return f"SCS1 Tx Pre-Cursor Lane {i-25}: 5-bit signed"
        if 34<=i<=41: return f"SCS1 Tx Post-Cursor Lane {i-33}: 5-bit signed"
        if 42<=i<=49: return f"SCS1 Rx Input Equalization Lane {i-41}: 4-bit unsigned"
        if 59<=i<=66: return f"SCS1 Rx CDR Bypass Lane {i-58}: 1=Bypass, 0=Enable CDR"
        if i==127: return "Page Checksum: Bytes 0x80~0xFE sum mod 256"
        return ""

    def _is_dirty(self,pk,addr_hex):
        try:
            s=int(addr_hex.replace("h","").split("~")[0],16)
            if pk=="p00": s-=0x80
            orig=self.orig_data[pk][s]
            return 0<=s<128 and self.data[pk][s]!=orig
        except: return False

    def _update_dirty_label(self):
        # 버그 수정: "for pk ... for i ... or condition" 은
        # (for i in range(128) or condition) 으로 파싱되어 항상 0 반환했던 문제 수정
        n=sum(1 for pk in PAGE_KEYS
                for i in range(128)
                if self.data[pk][i]!=self.orig_data[pk][i])
        self.dirty_label.config(text=f"  ✎ {n}개 수정됨" if n else "",
                                 fg=self.theme["dirty"])

    def _decoded_preview_cb(self, pk, abs_addr, byte_list):
        """팝업 Decoded 미리보기 콜백
        pk: 현재 페이지 키, abs_addr: 절대 주소(hex), byte_list: 현재 입력값
        """
        try:
            idx = abs_addr if pk=="a0" else abs_addr - 0x80
            if idx < 0 or idx > 127: return "—"
            # 임시 배열: 현재 페이지 데이터 복사 후 편집중 바이트 반영
            tmp = list(self.data[pk])
            for i, b in enumerate(byte_list):
                if 0 <= idx+i < 128:
                    tmp[idx+i] = b
            if pk=="a0":   return self._decoded_a0(tmp, idx) or "—"
            if pk=="p00":  return self._decoded_p00(tmp, idx) or "—"
            if pk=="p01":  return self._decoded_p01(tmp, idx) or "—"
            if pk=="p02":  return self._decoded_p02(tmp, idx) or "—"
            if pk=="p03":  return self._decoded_p03(tmp, idx) or "—"
        except Exception:
            pass
        return "—"

    def _on_inline_edit(self,ri,vals,byte_offset=0):
        """CanvasTable Value[hex] 바이트팝업 commit 콜백
        vals: 바이트 리스트
        byte_offset: 필드 내 시작 오프셋 (단일 바이트 편집 시)
        """
        if ri<0 or ri>=len(self.mct._rows): return
        row=self.mct._rows[ri][0]
        pk=self.current_page
        try:
            base_idx=int(row[1].replace("h","").split("~")[0],16)
            if pk!="a0": base_idx-=0x80
            addr_str=row[1].replace("h","")
            if "~" in addr_str:
                s,e=addr_str.split("~")
                field_len=int(e,16)-int(s,16)+1
            else:
                field_len=1
            write_len=min(len(vals), field_len-byte_offset)
            for i in range(write_len):
                target=base_idx+byte_offset+i
                if 0<=target<128:
                    self.data[pk][target]=vals[i]
            self._calc_checksum(pk)
            self._show_page(pk)
            val_str=" ".join(h2(v) for v in vals[:write_len])
            self._log(f"편집: {pk} [{row[1]}]+{byte_offset} ← {val_str}")
        except Exception as e:
            logging.error("인라인 편집 오류", exc_info=True)
            messagebox.showerror("오류",str(e))

    def _calc_checksum(self, pk):
        """페이지 체크섬 실시간 계산
        CMIS 5.2 기준:
          P00h: Byte 128~221 (idx 0~93) 합산 mod256 → idx 94
          P01h: Byte 130~254 (idx 2~126) 합산 mod256 → idx 127
          P02h: Byte 128~254 (idx 0~126) 합산 mod256 → idx 127
          P03h: Byte 128~254 (idx 0~126) 합산 mod256 → idx 127
          P10h/P11h: 표준 체크섬 없음 → 계산 안 함
        """
        v=self.data[pk]
        if pk=="p00":
            self.data[pk][94]=sum(v[0:94])&0xFF
        elif pk=="p01":
            # InactiveFW(idx 0~1) 제외, idx 2~126 합산
            self.data[pk][127]=sum(v[2:127])&0xFF
        elif pk in ("p02","p03"):
            self.data[pk][127]=sum(v[0:127])&0xFF
        # p10/p11: CMIS 표준 체크섬 미정의 → 계산하지 않음

    def _on_decoded_edit(self,ri):
        """Decoded 열 더블클릭: 드롭다운 또는 텍스트 편집 팝업"""
        if ri<0 or ri>=len(self.mct._rows): return
        row=self.mct._rows[ri][0]
        pk=self.current_page
        if pk!="p00": return  # P00h만 Decoded 편집 지원
        field=str(row[4]) if len(row)>4 else ""
        try:
            idx=int(str(row[1]).replace("h","").split("~")[0],16)-0x80
        except: return
        t=self.theme
        acc=t.get("acc","#185FA5"); bg=t.get("bg1","#f0f0f0")
        fg=t.get("t1","#111111")
        # ── 드롭다운 필드 (SFF8024, MediaTech)
        if idx==0:   # SFF8024Identifier
            self._show_dropdown(ri, idx, SFF8024_LIST,
                "모듈 타입 선택", pk)
            return
        if idx==84:  # MediaInterfaceTech
            self._show_dropdown(ri, idx, MEDIA_TECH_LIST,
                "미디어 인터페이스 기술 선택", pk)
            return
        # ── ASCII 텍스트 직접 편집 (idx 1~93, 체크섬 제외)
        ascii_ranges=[
            (1,16,"VendorName"),(20,16,"VendorPN"),(36,2,"VendorRev"),
            (38,16,"VendorSN"),(54,8,"DateCode"),(62,10,"CLEICode"),
        ]
        for start,ln,name in ascii_ranges:
            if start<=idx<start+ln:
                cur="".join(
                    chr(self.data[pk][start+j])
                    if 0x20<=self.data[pk][start+j]<=0x7E else " "
                    for j in range(ln))
                self._show_text_edit(ri, start, ln, cur.rstrip(), pk)
                return

    def _show_dropdown(self, ri, idx, items, title, pk):
        """드롭다운 팝업으로 값 선택"""
        t=self.theme
        acc=t.get("acc","#185FA5"); bg=t.get("bg1","#f0f0f0")
        fg=t.get("t1","#111111")
        pop=tk.Toplevel(self)
        pop.title(title)
        pop.resizable(False,False)
        pop.attributes("-topmost",True)
        pop.config(bg=bg)
        tk.Label(pop,text=title,bg=acc,fg="white",
                 font=("",9,"bold")).pack(fill=tk.X,padx=0,pady=0)
        var=tk.StringVar()
        cur_val=self.data[pk][idx]
        cur_label=next((lbl for v,lbl in items if v==cur_val),f"0x{cur_val:02X}")
        var.set(cur_label)
        cb=ttk.Combobox(pop,textvariable=var,
                        values=[lbl for _,lbl in items],
                        state="readonly",width=28,font=("Consolas",10))
        cb.pack(padx=10,pady=8)
        def apply():
            sel=var.get()
            for v,lbl in items:
                if lbl==sel:
                    self.data[pk][idx]=v
                    self._calc_checksum(pk)
                    self._show_page(pk)
                    self._log(f"편집(Decoded): {pk} [{idx+0x80:02X}h] ← 0x{v:02X} {sel}")
                    break
            pop.destroy()
        bf=tk.Frame(pop,bg=bg); bf.pack(pady=(0,8))
        tk.Button(bf,text="적용",width=8,cursor="hand2",
                  bg=acc,fg="white",relief=tk.FLAT,command=apply).pack(side=tk.LEFT,padx=4)
        tk.Button(bf,text="취소",width=8,cursor="hand2",
                  command=pop.destroy).pack(side=tk.LEFT,padx=4)
        cb.bind("<Return>",lambda e:apply())

    def _show_text_edit(self, ri, start, ln, cur_text, pk):
        """텍스트 직접 편집 팝업 → Value[hex] 업데이트"""
        t=self.theme
        acc=t.get("acc","#185FA5"); bg=t.get("bg1","#f0f0f0")
        fg=t.get("t1","#111111")
        pop=tk.Toplevel(self)
        pop.title(f"Decoded 편집 (최대 {ln}자)")
        pop.resizable(False,False)
        pop.attributes("-topmost",True)
        pop.config(bg=bg)
        tk.Label(pop,text=f"ASCII 입력 (최대 {ln}자, 우측 공백 패딩)",
                 bg=acc,fg="white",font=("",9,"bold")).pack(fill=tk.X)
        var=tk.StringVar(value=cur_text)
        e=tk.Entry(pop,textvariable=var,width=max(ln+4,20),
                   font=("Consolas",11),bg=bg,fg=fg,
                   relief="solid",bd=1)
        e.pack(padx=10,pady=8)
        e.select_range(0,"end"); e.focus_set()
        def apply():
            text=var.get()[:ln]  # 최대 ln자 자름
            padded=text.ljust(ln)  # 우측 공백 패딩
            for j,ch in enumerate(padded):
                self.data[pk][start+j]=ord(ch)&0xFF
            self._calc_checksum(pk)
            self._show_page(pk)
            self._log(f"편집(Decoded): {pk} [{start+0x80:02X}~{start+ln-1+0x80:02X}h] ← {repr(text)}")
            pop.destroy()
        bf=tk.Frame(pop,bg=bg); bf.pack(pady=(0,8))
        tk.Button(bf,text="적용",width=8,cursor="hand2",
                  bg=acc,fg="white",relief=tk.FLAT,command=apply).pack(side=tk.LEFT,padx=4)
        tk.Button(bf,text="취소",width=8,cursor="hand2",
                  command=pop.destroy).pack(side=tk.LEFT,padx=4)
        e.bind("<Return>",lambda ev:apply())
        e.bind("<Escape>",lambda ev:pop.destroy())

    def _on_row_dblclick(self,idx):
        pass  # Decoded열 편집은 _on_decoded_edit에서 처리

    # ── EEPROM 읽기/쓰기 ─────────────────────────────────
    # _read_eeprom : 별도 스레드에서 전체 페이지 읽기
    #   Lower Memory(A0h 00~7F) → 각 Upper Page(80~FF)
    #   페이지 전환: set_page(pn) → Byte 7Fh에 페이지 번호 쓰기
    # _start_write : 선택된 페이지/범위만 쓰기 (dirty_only 옵션)
    # _verify_write: 쓰기 후 read-back 비교 검증
    def _read_eeprom(self):
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        pages_total = 7  # a0 + p00~p03,p10,p11 6개
        def task():
            conn=None
            try:
                conn=self._open_conn()
                i2c=i2c_8bit(self.i2c_addr_var.get())
                self._log("━━━ EEPROM 읽기 시작 ━━━")
                # 상태 표시
                self.after(0,lambda:self.conn_status.config(
                    text="● 읽기 중...", fg=self.theme["amb"]))
                # Lower Memory (A0h)
                self._log(f"[1/7] Lower Memory (A0h) 읽기...")
                data=conn.read_page(i2c, 0x00, 128)
                for reg,val in enumerate(data): self.data["a0"][reg]=val
                self._log(f"  ✓ A0h 완료 ({len(data)}바이트)")
                # Upper Memory Pages
                page_info=list(zip(
                    ["p00","p01","p02","p03","p10","p11"],
                    [0,1,2,3,0x10,0x11],
                    ["P00h","P01h","P02h","P03h","P10h","P11h"]))
                for n,(pk,pn,lbl) in enumerate(page_info,2):
                    self._log(f"[{n}/7] {lbl} (Page 0x{pn:02X}) 읽기...")
                    conn.set_page(pn)
                    data=conn.read_page(i2c, 0x80, 128)
                    for reg,val in enumerate(data): self.data[pk][reg]=val
                    self._log(f"  ✓ {lbl} 완료 ({len(data)}바이트)")
                self.orig_data=copy.deepcopy(self.data)
                self._log("━━━ ✓ 읽기 완료 (7/7 페이지) ━━━")
                self.after(0,lambda:self._show_page(self.current_page))
            except Exception as e:
                logging.error("EEPROM 읽기 오류", exc_info=True)
                self._log(f"읽기 오류: {e} (crash.log 참조)")
                self.after(0,lambda e=e:messagebox.showerror("읽기 오류",str(e)))
            finally:
                self._close_conn(conn)
        threading.Thread(target=task,daemon=True).start()


    def _start_write(self):
        """EEPROM 쓰기 — dirty 바이트만 전송 (SFF-8636 방식)
        변경된 바이트(data != orig_data)만 전송 → 미변경 RO/Reserved 바이트 자동 제외
        """
        if self._device_index < 0:
            messagebox.showwarning("연결 필요","먼저 연결하세요."); return
        page_map={"p00":0,"p01":1,"p02":2,"p03":3,"p10":0x10,"p11":0x11}
        sel_pks=[pk for pk in PAGE_KEYS if self.page_sel_vars[pk].get()]
        if not sel_pks:
            messagebox.showwarning("페이지 미선택","쓸 페이지를 하나 이상 선택하세요."); return

        # ── dirty 바이트 수집 ─────────────────────────────
        targets=[]
        for pk in sel_pks:
            base=0 if pk=="a0" else 0x80
            for i in range(128):
                if self.data[pk][i] != self.orig_data[pk][i]:
                    targets.append((pk, i, base+i, self.data[pk][i]))

        # ── P00h: dirty 있으면 RW 범위 전체 쓰기로 확장 ──────
        # CMIS 모듈은 P00h를 page 단위로 NVM 커밋하는 경우가 있어
        # dirty 바이트만 쓰면 checksum 불일치로 NVM 저장 안 될 수 있음.
        # RO 바이트(Identifier, AppDesc 등)는 _RW_P00에서 이미 제외됨.
        if "p00" in sel_pks and any(pk_=="p00" for pk_,_,_,_ in targets):
            existing_p00={idx_ for pk_,idx_,_,_ in targets if pk_=="p00"}
            for i in sorted(_RW_P00):   # RW 바이트만
                if i not in existing_p00:
                    targets.append(("p00", i, i+0x80, self.data["p00"][i]))
            targets.sort(key=lambda x:(
                ["a0","p00","p01","p02","p10","p11"].index(x[0]), x[2]))

        # ── 체크섬 강제 포함 (P01h/P02h만) ─────────────────
        # P00h checksum(idx 94): 모듈이 자체 계산하므로 호스트 write 무시 → 제외
        cs_map={"p01":127,"p02":127}
        existing={(pk_,idx_) for pk_,idx_,_,_ in targets}
        for pk,cs_idx in cs_map.items():
            if pk in sel_pks:
                has_change=any(pk_==pk for pk_,_,_,_ in targets)
                if has_change and (pk,cs_idx) not in existing:
                    targets.append((pk,cs_idx,cs_idx+0x80,self.data[pk][cs_idx]))

        if not targets:
            messagebox.showinfo("변경 없음","수정된 바이트가 없습니다."); return

        # ── 확인 팝업 (dirty 바이트만 표시, 최대 20행) ──────
        dirty_only=[(pk,idx,addr,val) for pk,idx,addr,val in targets
                    if self.data[pk][idx] != self.orig_data[pk][idx]]
        summary=["[변경된 바이트 쓰기 확인]\n"]
        # P00h 전체 쓰기 안내
        if any(pk_=="p00" for pk_,_,_,_ in targets):
            p00_total=sum(1 for pk_,_,_,_ in targets if pk_=="p00")
            summary.append(f"  ※ P00h: NVM 커밋 보장을 위해 idx 0~94 전체 전송 ({p00_total}바이트)")
            summary.append("")
        cur_pk=None; shown=0
        for pk,idx,addr,val in dirty_only:
            if pk!=cur_pk:
                summary.append(f"  [{pk.upper()}]"); cur_pk=pk
            orig=self.orig_data[pk][idx]
            summary.append(f"    {addr:02X}h  {orig:02X}h → {val:02X}h")
            shown+=1
            if shown>=20:
                remain=len(dirty_only)-shown
                if remain>0:
                    summary.append(f"    ... 외 {remain}개 변경")
                break
        summary.append(f"\n실제 전송: {len(targets)}바이트  되돌릴 수 없습니다. 계속?")
        if not messagebox.askyesno("쓰기 확인","\n".join(summary)): return

        # ── 실제 쓰기 스레드 ─────────────────────────────────
        def task():
            conn=None
            try:
                conn=self._open_conn()
                i2c=i2c_8bit(self.i2c_addr_var.get())
                NVM_DELAY=0.020   # NVM tWR: burst 후 20ms 대기
                self._log(f"━━━ EEPROM 쓰기 시작 (burst write, {len(targets)}바이트) ━━━")
                self.after(0,lambda:self.conn_status.config(
                    text="● 쓰기 중...",fg=self.theme["amb"]))

                # ── targets를 연속 주소 burst 그룹으로 분할 ──
                # (pk, addr) 연속이면 하나의 I2C 트랜잭션으로 묶음
                bursts=[]  # [(pk, start_idx, start_addr, [val0, val1, ...])]
                for pk,idx,addr,val in targets:
                    if (bursts and bursts[-1][0]==pk
                            and addr==bursts[-1][2]+len(bursts[-1][3])):
                        bursts[-1][3].append(val)
                    else:
                        bursts.append([pk, idx, addr, [val]])

                total_bursts=len(bursts); done_bytes=0
                cur_page=None   # 페이지 변경 시에만 set_page 호출
                for bi,(pk,idx,start_addr,vals) in enumerate(bursts):
                    # 페이지 선택 — 페이지가 바뀔 때만 한 번 호출
                    if pk!="a0" and pk!=cur_page:
                        conn.set_page(page_map[pk])
                        cur_page=pk
                    n=len(vals)
                    abs_str=(f"{start_addr:02X}h" if n==1
                             else f"{start_addr:02X}~{start_addr+n-1:02X}h")
                    self._log(f"  → {pk.upper()} [{abs_str}] "
                              f"← {' '.join(h2(v) for v in vals)} "
                              f"({bi+1}/{total_bursts})")
                    conn.write_burst(i2c, start_addr, vals)
                    done_bytes+=n
                    pct=done_bytes/len(targets)*100
                    msg=f"{pk.upper()} [{abs_str}] ({done_bytes}/{len(targets)})"
                    self.after(0,lambda p=pct,m=msg:(
                        self.progress_var.set(p),
                        self.progress_label.config(text=m)))

                conn.set_page(0)
                self.orig_data=copy.deepcopy(self.data)
                self._log(f"━━━ ✓ 쓰기 완료 ({done_bytes}바이트, {total_bursts}개 burst) ━━━")

                # ── Read-back 검증 (페이지 단위 read) ──────────
                self._log("━━━ Read-back 검증 시작 ━━━")
                mismatches=[]
                # 검증 대상 페이지만 한 번씩 읽어 캐시
                verify_pks={pk_ for pk_,_,_,_ in targets}
                rb_cache={}
                for pk in verify_pks:
                    if pk!="a0": conn.set_page(page_map[pk])
                    base=0 if pk=="a0" else 0x80
                    rb_cache[pk]=conn.read_page(i2c, base, 128)
                conn.set_page(0)
                # 캐시와 비교
                for pk,idx,addr,wrote_val in targets:
                    rb=rb_cache.get(pk)
                    read_val=rb[idx] if rb and idx<len(rb) else None
                    if read_val != wrote_val:
                        mismatches.append(
                            (pk.upper(), addr, wrote_val, read_val))
                # verify read 결과로 data/orig_data 동기화
                # → 모듈이 자체 계산한 CC 등 반영, dirty 표시 방지
                # ※ CC(idx 94)는 모듈 무시로 우리 계산값 유지 → 제외
                CC_SKIP = {"p00": {94}}  # 각 페이지별 동기화 제외 idx
                for pk, rb in rb_cache.items():
                    skip = CC_SKIP.get(pk, set())
                    for i, v in enumerate(rb[:128]):
                        if i in skip: continue
                        self.data[pk][i]      = v
                        self.orig_data[pk][i] = v
                # CC는 현재 data 기준으로 재계산, orig도 동기화 (dirty 제거)
                for pk in rb_cache:
                    self._calc_checksum(pk)
                    # p00 CC: orig도 재계산값으로 동기화 → dirty 표시 제거
                    if pk=="p00":
                        self.orig_data[pk][94]=self.data[pk][94]

                if mismatches:
                    self._log(f"━━━ ✗ 검증 실패: {len(mismatches)}개 불일치 ━━━")
                    for pg,addr,wv,rv in mismatches:
                        self._log(f"  [{pg} {addr:02X}h] 썼음={h2(wv)} 읽힘={h2(rv)}")
                    detail="\n".join(
                        f"  [{pg} {addr:02X}h] 썼음={h2(wv)} 읽힘={h2(rv)}"
                        for pg,addr,wv,rv in mismatches[:10])
                    if len(mismatches)>10:
                        detail+=f"\n  ... 외 {len(mismatches)-10}개"
                    self.after(0,lambda detail=detail,n=len(mismatches):(
                        self.progress_var.set(0),
                        self.progress_label.config(text=f"✗ 검증 실패 {n}개"),
                        self._update_dirty_label(),
                        self._show_page(self.current_page),
                        messagebox.showwarning("검증 실패",
                            f"쓰기 완료 ({done_bytes}바이트)\n\n"
                            f"Read-back 불일치: {n}개\n{detail}")))
                else:
                    self._log(f"━━━ ✓ 검증 통과 ({done_bytes}바이트 모두 일치) ━━━")
                    self.after(0,lambda:(
                        self.progress_var.set(0),
                        self.progress_label.config(text=f"✓ {done_bytes}바이트 쓰기+검증 완료"),
                        self._update_dirty_label(),
                        self._show_page(self.current_page),
                        messagebox.showinfo("완료",
                            f"쓰기 완료 ({done_bytes}바이트)\n✅ Read-back 검증 통과")))
            except Exception as ex:
                logging.error("EEPROM 쓰기 오류",exc_info=True)
                self._log(f"쓰기 오류: {ex} (crash.log 참조)")
                self.after(0,lambda ex=ex:messagebox.showerror("쓰기 오류",str(ex)))
            finally:
                self._close_conn(conn)
        threading.Thread(target=task,daemon=True).start()
    def _verify_write(self):
        """쓰기 후 검증 — _start_write 내부에 inline verify가 포함되어 있으므로
        별도 수동 검증은 현재 미사용 (버튼에서 호출 시 안내 메시지 표시)"""
        messagebox.showinfo("검증 안내",
            "쓰기 완료 후 Read-back 검증이 자동으로 실행됩니다.\n"
            "별도 수동 검증은 필요하지 않습니다.")

    # ── 로그 ─────────────────────────────────────────────
    def _log(self,msg):
        ts=time.strftime("%H:%M:%S")
        def _ins():
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END,f"[{ts}] {msg}\n")
            self.log_text.see(tk.END); self.log_text.config(state=tk.DISABLED)
        self.after(0,_ins)

    def _clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete("1.0",tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _save_log(self):
        path=filedialog.asksaveasfilename(defaultextension=".txt",
                                           filetypes=[("Text","*.txt")])
        if not path: return
        with open(path,"w") as f: f.write(self.log_text.get("1.0",tk.END))


# ── 실행 ──────────────────────────────────────────────────
if __name__=="__main__":
    logging.info(f"=== {APP_NAME} v{APP_VERSION} 시작 ===")
    try:
        app=App()
        app.mainloop()
        logging.info(f"=== 정상 종료 ===")
    except Exception:
        logging.error("치명적 오류로 종료", exc_info=True)
        print(f"오류 발생. 상세: {CRASH_LOG}"); raise
